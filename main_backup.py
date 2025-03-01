from datetime import datetime, timedelta
import random
import os
import pandas
import pyodbc
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from tqdm import tqdm
import shutil
import json

def Rest_Qty(Quantity,Delivery):
    if Quantity % Delivery > 0:
        Generated_Qty = (Quantity // Delivery) * (Delivery - 1)
        return Quantity - Generated_Qty
    else:
        return Quantity // Delivery

def test_integer(inset_value):
    while True:
        try:
            number = int(inset_value)
            return number
        except ValueError:
            inset_value = input(f"The {inset_value} is not integer number, test again: ")

def test_float(inset_value):
    while True:
        try:
            number = float(inset_value)
            number = round(number, 2)
            return number
        except ValueError:
            inset_value = input(f"The {inset_value} is not float number, test again: ")

def test_date(inset_value, format):
    while True:
        try:
            if inset_value == "t":
                today = datetime.today()
                today2 = today.strftime(format)
                return today2
            elif inset_value == "l":
                today = datetime.strptime("31.12.2099",format)
                today2 = today.strftime(format)
                return today2
            elif inset_value == "f":
                today = datetime.strptime("01.01.1755",format)
                today2 = today.strftime(format)
                return today2
            else:
                date = datetime.strptime(inset_value,format)
                date2 = date.strftime(format)
                return date2
        except:
            inset_value = input(f"The {inset_value} is not date, test again: ")


def update_string(inset_value):
    try:
        inset_value = str(inset_value).replace(" ", "_")
        return inset_value
    except:
        return inset_value

def excel_colum_def(COL_ID):
    div1 = COL_ID
    column_label = str()
    while div1:
        (div1, mod1) = divmod(div1, 26)
        column_label = chr(mod1 + 64) + column_label
    return column_label

def concatenate_dataframe(dataframe_name, JSON_help_path):
    os.chdir(f"{JSON_help_path}")
    Read_file = open(f"{dataframe_name}.json", "rt", encoding="Windows-1250", errors='ignore')
    lines = Read_file.readlines()
    file.write(f"""\t\t"{dataframe_name}": \n""")
    file.write("""\t\t{\n""")  
    for line in lines:
        line_len = int(len(line.rstrip("\n")))
        if (line_len != 0) and (line != "    }\n") and (line != "    {\n"):
            file.write("\t"+f"{line}")
    if dataframe_name == "HQ_Packg_Track_Setup_df":
        file.write("""\t\t}\n""")    
        file.write("""\t}\n""") 
        file.write("""}""") 
    else:
        file.write("""\t\t},\n""")    
    file.write("\n") 

# Defaults
print("""--------------------------------------------------------------------------------------------------------------------------
|   Designed by: Jan Vaško                                                                                               |
|   Contact: Jan.Vasko@konicaminolta.eu                                                                                  |
|   Phone: + 420 601 383 301                                                                                             |
|------------------------------------------------------------------------------------------------------------------------|""")
print("""|  Before you run questionary you have to prepare:                                                                       |
|    Number Series for:                                                                                                  |
|        1) Purchase Order                                                                                               |
|        2) HQ Item Transport Register - first available (not used) Regsiter ID                                          |
|        3) BEU Confirmation Number                                                                                      |
|        4) BEU PreAdvice Number                                                                                         |
|        5) BEU Delivery Number                                                                                          |
|        6) BEU Invoice Number                                                                                           |
|        7) NUS3: HQ ATP Check Register - first available (not used) Regsiter ID                                         |
|        8) NUS3: HQ Substitution Register - first available (not used) Regsiter ID                                      |
|        9) NUS3: HQ SerialNumber Register - first available (not used) Regsiter ID                                      |
|        10) Serial Numbes No. Series                                                                                    |
|        11) NUS3: HQ Delivery Tracking Register - first available (not used) Regsiter ID                                |
|            A) Package Numbers                                                                                          |
|            B) Bill of Landing Numbes                                                                                   |
|            C) EXIDV2 Numbers                                                                                           |
|        12) NUS3: HQ Package Tracking Register - first available (not used) Regsiter ID                                 |
|                                                                                                                        |
|    Orther Information needs to be collected before run:                                                                |
|        1) Vendor Number                                                                                                |
|        2) NUS3: HQ Logistic Process                                                                                    |
|        3) HQ Order Type                                                                                                |
|        4) HQ Shipping Condition                                                                                        |
|        5) Shipment Method                                                                                              |
|        6) Shipping Agent                                                                                               |
|        7) Shipping Agent Service Code                                                                                  |
|        8) Location Code                                                                                                |
|        9) Currency Code                                                                                                |
|        10) Country/Region of Origin Code                                                                               |
|        11) Plant Numbers                                                                                               |
|        12) Tariff Numbers                                                                                              |
|        13) ATP Stock                                                                                                   |
|        14) Weight Unit of Measure                                                                                      |
|        15) Volume Unit of Measure                                                                                      |
|                                                                                                                        |
|    Other info:                                                                                                         |
|       Items                                                                                                            |
|        - must be specified all of them from the perspective of HQ Communciation on Confimraiton file                   |
|        - Free of charge, BOM ... must be added and address to proper main Items                                        |
|        - program will exclude them on Purhcase Line --> will be updated later by "HQ Update" (as it is in reality)     |
|        - Items Randomisation is funciotn to create houndreds of PO lines (mainly in replenishment) without need of     |
|          specify all of them. You have to just define few of them and function will add another lines randomly from    |
|          your inserted ones.                                                                                           |""")
print("""|------------------------------------------------------------------------------------------------------------------------|    
|  JSON setups implementation                                                                                            |
|    - you can have all setup prepared in .json format, json provided into NUS3 for:                                     |
|        1) Replenishment                                                                                                |
|        2) Replenishment with cancellation Items on confimation                                                         |
|        3) DDL Tower                                                                                                    |
|        4) DDL Dealer                                                                                                   |
|        5) DDL Consumables                                                                                              |
|        6) BEU Set                                                                                                      |
|                                                                                                                        |
|    - you can safe your questionary choises as .json for future use (without questionay)                                |""")
print("""|------------------------------------------------------------------------------------------------------------------------|    
|  Navision Version:                                                                                                     |
|    - all data are prepared for related HQ integration tables which has some influence on process                       |
|    NUS2 Tables:                                                                                                        |
|        1) Purhcase Header                                                                                              |
|        2) Purchase Line                                                                                                |
|        3) HQ Item Transport Register                                                                                   |
|        4) HQ Transfer Tracking No.                                                                                     |
|    NUS3:                                                                                                               |
|        1) Purhcase Header                                                                                              |
|        2) Purchase Line                                                                                                |
|        3) HQ Item Transport Register                                                                                   |
|        4) HQ ATP Check Register                                                                                        |
|        5) HQ Substitution Register                                                                                     |
|        6) HQ PreAdvice Register                                                                                        |
|        7) HQ SerialNumber Register                                                                                     |
|        8) HQ Delivery Tracking Register                                                                                |
|        9) HQ Package Tracking Register                                                                                 |""")
print("""|------------------------------------------------------------------------------------------------------------------------|    
|  SAP documents:                                                                                                        |
|    NUS2: Confirmation, PreAdvice, Delivery and Invoice                                                                 |
|    NUS3: Export, Confirmation, PreAdvice, Delivery and Invoice                                                         |""")
print("""|------------------------------------------------------------------------------------------------------------------------|    
|  Data downloaded - only from NUS2                                                                                      |
|    - prepared for download values for several entities from NUS2 DBS QA and PRD                                        |
|    Downloadable entity:                                                                                                |
|        1) Shipment Methods                                                                                             |
|        2) Shipping Agents                                                                                              |
|        3) Shipping Agens Service Codes                                                                                 |
|        4) Country Region of Origin Codes                                                                               |
|        5) Pland No.                                                                                                    |
|        6) Tariff Numbers                                                                                               |
|        7) HQ Shipping Condition                                                                                        |""")
print("""|------------------------------------------------------------------------------------------------------------------------|  
|  Option Values:                                                                                                        |
|    1) General_Setup_df                                                                                                 |
|        1) Navision - "NUS2", "NUS3"                                                                                    |     
|    2) PO_Document_Header_Setup_df                                                                                      |
|        1) Document_Type - "Order"                                                                                      |
|        2) HQ_Order_Type - "standard","standard_overnight","express","overnight","normal","directdelivery",             |
|                           "directdelivery_overnight","return_order","return_order_spare_parts","release",              |
|                           "consignment","fieldservice","airlens"                                                       |
|        3) HQ_Shipping_Condition - "01", "02", "03", "04", "05", "14", "18"                                             |
|    3) Items_df                                                                                                         |
|        1) Line_Type - "TEXT", "ITEM"                                                                                   |
|        2) HQ_Confirmation_Line_Flag - "Substituted", "Cancelled", "Label", "Finished", "To Parent"                     |
|    4) HQ_Delivery_Setup_df                                                                                             |
|        1) HQ_Delviery_E_Com_Process_Stat - "Sent", "Ready To Send"                                                     |
|    5) HQ_Invoice_Setup_df                                                                                              |
|        1) HQ_Invoice_Plant - "1000", "1002", "1004"                                                                    |
|    6) HQ_ATPR_Setup_df                                                                                                 |
|        1) HQ_ATP_Stock - "ONH", "ONB", "BACK"                                                                          |""")
print("""|------------------------------------------------------------------------------------------------------------------------|
|  Dates Values:                                                                                                         |
|    - all dates relates to Date Format, must be updated by country specific ...                                         |
|    - all dates also allow Today option just with "t" --> like in NAV                                                   |
|    1) General_Setup_df                                                                                                 |
|        1) HQ_Date_format - %d.%m.%Y [15.06.2022] --> NUS3                                                              |
|        2) HQ_Date_format - %d.%m.%y [15.06.22]   --> NUS2                                                              |""")
print("""|------------------------------------------------------------------------------------------------------------------------|
|   Numeric Values:                                                                                                      |
|    - all Numeric Values are treated as "Integer" --> without decimal space excep:                                      |
|    1) Items_df                                                                                                         |
|        1) Item_Unit_Price --> can has 2 decimal space                                                                  |""")
print("""|------------------------------------------------------------------------------------------------------------------------|
|   Bolean values:                                                                                                       |
|    - usually used in questionary to determin if you want to use specific opetion (like random weight, Free of charge   |
|      ...) valuees are  [Y/N]                                                                                           |
|    Y - Yes                                                                                                             |
|    N - No                                                                                                              |""")
print("""|------------------------------------------------------------------------------------------------------------------------|
|    Functions backlog (to be done):                                                                                     |
|    1) Corect selection of "Shipping Agent Service Code" based on "Shipping Agent Code" whem randomization of Agent used|
|    2) Plant No. --> make same for Invoice and HQ Delivery Tracking Register                                            |
|    3) Multiple Confirmaiton --> to check Cancellation                                                                  |
|    4) .xsd application to created Excel --> not need to copy / paste to template                                       |
|    5) Create XML files (to simulate HQ Import Purchase Orders) instead of Import Excel                                 |
|    6) Fields definition for NUS2                                                                                       |
--------------------------------------------------------------------------------------------------------------------------
""")

JSON_Load = ""
while JSON_Load != "Y" and JSON_Load != "N":
    JSON_Load = update_string(input("Do you want to load setups from JSON file? [Y/N]: ")).upper() 
if JSON_Load == "Y":
    while True:
        JSON_Path = str(input("Put full path without file name [string]: "))
        JSON_Load_File_Name = str(input("Put file name [string]: "))
        data = pandas.read_json(f"{JSON_Path}{JSON_Load_File_Name}")
        # Define each settings from JSON 
        General_Setup_df = pandas.DataFrame(data["Setup"]["General_Setup_df"], columns=["Navision", "HQ_Date_format", "Conf_Package", "Export_File_name", "Export_File_Location_as_Current", "Export_File_path", "Export_File_Type"], index=[0])
        Questions_df = pandas.DataFrame(data["Setup"]["Questions_df"], columns=["HQ_Export_Quest", "HQ_Conf_Quest", "HQ_PreAdv_Quest", "HQ_Del_Quest", "HQ_DeL_Track_Quest", "HQ_Packg_Track_Quest", "HQ_Inv_Quest", "Record_link_Quest"], index=[0])
        PO_Document_Header_Setup_df = pandas.DataFrame(data["Setup"]["PO_Document_Header_Setup_df"], columns=["Document_Type", "Document_Number_prefix", "Document_Number_suffix", "Document_Number_Increment", "Documents_Count", "Buy_from_Vendor_No", "Document_Location","HQ_Logistic_Process", "HQ_Random_Order_Type", "HQ_Order_Type", "HQ_Shipping_Condition", "Shipment_Methods", "Shipping_Agent", "Shipping_Agent_Service"], index=[0])
        Item_Index_count = list(range(len(data["Setup"]["Items_df"]["Item_No"])))   # Helps to count number of Items in the list
        Items_df = pandas.DataFrame(data["Setup"]["Items_df"], columns=["Item_No", "Line_Type", "Item_Line_Quantity", "Item_Unit_of_Measure", "Item_Unit_Price", "Main_BOM_Item", "Item_Connected_to_BOM", "BOM_Item_Relation", "Item_Free_Of_Charge", "Item_Free_Of_Charge_Relation", "Item_SN_Tracking", "HQ_Confirmation_Line_Flag_Use", "HQ_Confirmation_Line_Flag", "HQ_SUB_New_Item"],  index=Item_Index_count)
        HQ_General_Setup_df = pandas.DataFrame(data["Setup"]["HQ_General_Setup_df"], columns=["HQ_HQITR_Register_No_Start", "HQ_HQITR_Register_No_Increment", "HQ_Order_Date", "HQ_Delivery_Start_and_End_Date", "HQ_Currency_Code", "HQ_Location_Code", "HQ_Document_Order_No_Start", "HQ_Document_Order_No_Increment", "HQ_Vendor_Line_No_Start", "HQ_Vendor_Line_No_Increment"], index=[0])
        HQ_Export_Setup_df = pandas.DataFrame(data["Setup"]["HQ_Export_Setup_df"], columns=["HQ_Export_Vendor_Line_No"], index=[0])
        HQ_Confirmation_Setup_df = pandas.DataFrame(data["Setup"]["HQ_Confirmation_Setup_df"], columns=["HQ_Confirmation_prefix", "HQ_Confirmation_suffix", "HQ_Confirmation_Increment", "HQ_Conf_Vendor_Document_Created_Date"], index=[0])
        HQ_PreAdvice_Setup_df = pandas.DataFrame(data["Setup"]["HQ_PreAdvice_Setup_df"], columns=["HQ_PreAdvice_prefix", "HQ_PreAdvice_suffix", "HQ_PreAdvice_Increment", "HQ_PreAdvice_Vendor_Document_Created_Date"], index=[0])
        HQ_Delivery_Setup_df = pandas.DataFrame(data["Setup"]["HQ_Delivery_Setup_df"], columns=["HQ_Delivery_prefix", "HQ_Delivery_suffix", "HQ_Delivery_Increment", "Delivery_Count_per_Order", "Delivery_Random_Select", "HQ_Delivery_Receipt_Date", "HQ_Delivery_Vendor_Document_Created_Date", "HQ_Delviery_E_Com_Process_Stat", "HQ_Delviery_To_Post_Auto"], index=[0])
        HQ_Invoice_Setup_df = pandas.DataFrame(data["Setup"]["HQ_Invoice_Setup_df"], columns=["HQ_Invoice_prefix", "HQ_Invoice_suffix", "HQ_Invoice_Increment", "HQ_Invoice_Vendor_Document_Created_Date", "HQ_Invoice_Random_Count_Reg_Origin", "HQ_Invoice_Count_Reg_Origin", "HQ_Invoice_Random_Plant", "HQ_Invoice_Plant", "HQ_Invoice_Random_Tariff_No", "HQ_Invoice_Tariff", "HQ_Invoice_To_Post_Auto"], index=[0])
        HQ_ATPR_Setup_df = pandas.DataFrame(data["Setup"]["HQ_ATPR_Setup_df"], columns=["HQ_ATP_Register_No_Start", "HQ_ATP_Register_No_Increment", "HQ_ATP_Random_Stock", "HQ_ATP_Stock", "HQ_ATP_Multiple_Lines_Check", "HQ_ATP_Multiple_Lines", "HQ_ATP_First_Scheduled_Date", "HQ_ATP_Sched_Date_Increments_day", "HQ_ATP_Zero_Date"], index=[0])
        HQ_SUB_Setup_df = pandas.DataFrame(data["Setup"]["HQ_SUB_Setup_df"], columns=["HQ_SUB_Register_No_Start", "HQ_SUB_Register_No_Increment"], index=[0])
        HQ_HQPAR_Setup_df = pandas.DataFrame(data["Setup"]["HQ_HQPAR_Setup_df"], columns=["HQ_HQPAR_Register_No_Start", "HQ_HQPAR_Register_No_Increment"], index=[0])
        HQ_HQSNR_Setup_df = pandas.DataFrame(data["Setup"]["HQ_HQSNR_Setup_df"], columns=["HQ_HQSNR_Register_No_Start", "HQ_HQSNR_Register_No_Increment", "HQ_SN_SN_prefix", "HQ_SN_SN_suffix", "HQ_SN_SN_Increment"], index=[0])
        HQ_DEL_Track_Setup_df = pandas.DataFrame(data["Setup"]["HQ_DEL_Track_Setup_df"], columns=["HQ_HQDTR_Register_No_Start", "HQ_HQDTR_Register_No_Increment", "HQ_Del_Track_Packa_prefix", "HQ_Del_Track_Packa_suffix", "HQ_Del_Track_Packa_Increment", "HQ_Del_Track_Bill_prefix", "HQ_Del_Track_Bill_suffix", "HQ_Del_Track_Bill_Increment", "HQ_Del_Track_EXIDV_prefix", "HQ_Del_Track_EXIDV_suffix", "HQ_Del_Track_EXIDV_Increment", "HQ_Del_Track_Random_Shipment_Method", "HQ_Del_Track_Shipment_Method", "HQ_Del_Track_Random_Shipping_Agent", "HQ_Del_Track_Shipping_Agent", "HQ_Del_Track_Random_Weight", "HQ_Del_Track_Weight", "HQ_Del_Track_Weight_UOM", "HQ_Del_Track_Random_Volume", "HQ_Del_Track_Volume", "HQ_Del_Track_Volume_UOM"], index=[0])
        HQ_Packg_Track_Setup_df = pandas.DataFrame(data["Setup"]["HQ_Packg_Track_Setup_df"], columns=["HQ_HQPTR_Register_No_Start", "HQ_HQPTR_Register_No_Increment", "HQ_Packg_Track_Packages_per_Delivery"], index=[0])
        Record_Links_Setup_df = pandas.DataFrame(data["Setup"]["Record_Links_Setup_df"], columns=["NOC", "Server", "Server_link", "User ID", "Company"], index=[0])

        General_Setup_df.Name = "General_Setup_df"
        Questions_df.Name = "Questions_df"
        PO_Document_Header_Setup_df.Name = "PO_Document_Header_Setup_df"
        Items_df.Name = "Items_df"
        HQ_General_Setup_df.Name = "HQ_General_Setup_df"
        HQ_Export_Setup_df.Name = "HQ_Export_Setup_df"
        HQ_Confirmation_Setup_df.Name = "HQ_Confirmation_Setup_df"
        HQ_PreAdvice_Setup_df.Name = "HQ_PreAdvice_Setup_df"
        HQ_Delivery_Setup_df.Name = "HQ_Delivery_Setup_df"
        HQ_Invoice_Setup_df.Name = "HQ_Invoice_Setup_df"
        HQ_ATPR_Setup_df.Name = "HQ_ATPR_Setup_df"
        HQ_SUB_Setup_df.Name = "HQ_SUB_Setup_df"
        HQ_HQPAR_Setup_df.Name = "HQ_HQPAR_Setup_df"
        HQ_HQSNR_Setup_df.Name = "HQ_HQSNR_Setup_df"
        HQ_DEL_Track_Setup_df.Name = "HQ_DEL_Track_Setup_df"
        HQ_Packg_Track_Setup_df.Name = "HQ_Packg_Track_Setup_df"
        Record_Links_Setup_df.Name = "Record_Links_Setup_df"


        # print Setup
        print("\n-section summary-")
        print(General_Setup_df.transpose())

        print("\n-section summary-")
        print(Questions_df.transpose())

        print("\n-section summary-")
        print(PO_Document_Header_Setup_df.transpose())

        print("\n-section summary-")
        print(Items_df.transpose())

        print("\n-section summary-")
        print(HQ_General_Setup_df.transpose())

        print("\n-section summary-")
        print(HQ_Export_Setup_df.transpose())

        print("\n-section summary-")
        print(HQ_Confirmation_Setup_df.transpose())

        print("\n-section summary-")
        print(HQ_PreAdvice_Setup_df.transpose())

        print("\n-section summary-")
        print(HQ_Delivery_Setup_df.transpose())

        print("\n-section summary-")
        print(HQ_Invoice_Setup_df.transpose())

        print("\n-section summary-")
        print(HQ_ATPR_Setup_df.transpose())

        print("\n-section summary-")
        print(HQ_SUB_Setup_df.transpose())

        print("\n-section summary-")
        print(HQ_HQPAR_Setup_df.transpose())

        print("\n-section summary-")
        print(HQ_HQSNR_Setup_df.transpose())

        print("\n-section summary-")
        print(HQ_DEL_Track_Setup_df.transpose())

        print("\n-section summary-")
        print(HQ_Packg_Track_Setup_df.transpose())

        print("\n-section summary-")
        print(Record_Links_Setup_df.transpose())

        # Section checker
        Section_check = ""
        while Section_check != "Y" and Section_check != "N":
            Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
        if Section_check == "Y":
            print("Data Loaded ...")
            break
        elif Section_check == "N":
            pass  
    
if JSON_Load == "N":
    # General Informaiton
    print("\n-----------------------------------------------------")
    print("---General information---")

    # Questionary base definition:
    Questions_dict = {
        "HQ_Export_Quest": "",
        "HQ_Conf_Quest": "",
        "HQ_PreAdv_Quest": "",
        "HQ_Del_Quest": "",
        "HQ_DeL_Track_Quest": "",
        "HQ_Packg_Track_Quest": "",
        "HQ_Inv_Quest": "",
        "Record_link_Quest": ""}
    Questions_df = pandas.DataFrame(Questions_dict, columns=Questions_dict.keys(), index=[0])
    Questions_df.Name = "Questions_df"

    while True:
        # Defined by Export Type
        Export_File_Type = ""
        while Export_File_Type != "EXCEL" and Export_File_Type != "XML":
            Export_File_Type = update_string(input("Define to which format to export (Excel - One file for Data import; XML - multiple file per Vendor Document No)[EXCEL/XML]: ")).upper() 
        if Export_File_Type == "EXCEL":
            Export_File_Location_as_Current = ""
            while Export_File_Location_as_Current != "Y" and Export_File_Location_as_Current != "N":
                Export_File_Location_as_Current = update_string(input("Do you want to create export in the same directory as current script[Y/N]: ")).upper() 
            if Export_File_Location_as_Current == "Y":
                Export_File_path = os.path.dirname(os.path.abspath(__file__))
            elif Export_File_Location_as_Current == "N":
                Export_File_path = str(input("Please define full path where to save Excel file [string]: "))
            Export_File_name = update_string(input("Define your file name[string]: "))
        elif Export_File_Type == "XML":
            Export_File_Location_as_Current = "N"
            Export_File_path = str(input("Please define full path to HQ Import folder '../TONAV/' (Program will automacically find sub folders for each type) [string]: "))
            Export_File_name = ""
            
        # Navisoin Version + Date Firmat
        Navision = ""
        while Navision != "NUS3" and Navision != "NUS2":
            Navision = update_string(input("For which NAV version you prepare data? [NUS2 / NUS3][string]: "))
        if Navision == "NUS3":
            Conf_Package = "HQ_DATA_GENERATOR"
            print(f"Generator prepare data for Configuraiton Package {Conf_Package}.")
        if Navision == "NUS3":
            HQ_Date_format_check = "Long"
        elif Navision == "NUS2":
            HQ_Date_format_check = "Short"
        else:
            HQ_Date_format_check = ""
        while HQ_Date_format_check != "Long" and HQ_Date_format_check != "Short":
            HQ_Date_format_check = update_string(input("Which date format to use? [Long - 15.06.2022 (NUS3) / Short - 15.06.22 (NUS2)][string]: "))
        if HQ_Date_format_check == "Long":
            HQ_Date_format = "%d.%m.%Y"
        else:
            HQ_Date_format = "%d.%m.%y"

        General_Setup_Dict = {
            "Navision": Navision,
            "HQ_Date_format": HQ_Date_format,
            "Conf_Package": Conf_Package,
            "Export_File_name": Export_File_name,
            "Export_File_Location_as_Current": Export_File_Location_as_Current,
            "Export_File_path": Export_File_path,
            "Export_File_Type": Export_File_Type}
            
        General_Setup_df = pandas.DataFrame(General_Setup_Dict, columns=General_Setup_Dict.keys(), index=[0])
        General_Setup_df.Name = "General_Setup_df"
        print("\n-section summary-")
        print(General_Setup_df.transpose())

        # Section checker
        Section_check = ""
        while Section_check != "Y" and Section_check != "N":
            Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
        if Section_check == "Y":
            break
        elif Section_check == "N":
            pass

    # Data downloaded from DB
    Download_Data_NOC_list = ["BBG", "BBL", "BCZ", "BFI", "BGR", "BHN", "BNO", "BPT", "BRO", "BSK", "BSW", "BTR", "BUR"]
    Download_Data = ""
    while Download_Data != "Y" and Download_Data != "N":
        Download_Data = update_string(input("""Do you want to download data from NUS2 QA or PRD DB for some or all of these tables:
            1) Shipment_Methods list
            2) Shipping Agents
            3) Shipping Agens Service Codes
            4) Country Region of Origin Codes
            5) Pland No.
            6) Tariff Numbers 
            7) HQ Shipping Condition
                [Y/N]: """))
    if Download_Data == "Y":
        while True:
            Download_Data_System = ""
            while Download_Data_System != "QA" and Download_Data_System != "PRD":
                Download_Data_System = update_string(input("Which DB you want to select as data source? [QA/PRD]: "))
            Download_Data_NOC = ""
            while Download_Data_NOC not in Download_Data_NOC_list:
                Download_Data_NOC = update_string(input(f"Which NOC you want to select? {Download_Data_NOC_list} [Code]: "))
            Download_Data_Shipment_Methods = update_string(input("Do you want download data for this table [Shipment Methods]? [Y/N]: ")).upper() 
            Download_Data_Shipping_Agents = update_string(input("Do you want download data for this table [Shipping Agents]? [Y/N]: ")).upper() 
            if Download_Data_Shipping_Agents == "Y":
                Download_Data_Shipping_Agens_Service_Codes = update_string(input("Do you want download data for this table [Shipping Agent Service Code] [Y/N]?: "))
            elif Download_Data_Shipping_Agents == "N":
                Download_Data_Shipping_Agens_Service_Codes = "N"
            Download_Data_Country_Region_of_Origin_Codes = update_string(input("Do you want download data for this table [Country/Region Codes]? [Y/N]: ")).upper() 
            Download_Data_Pland_No = update_string(input("Do you want download data for this table [Plant No.]? [Y/N]: ")).upper() 
            Download_Data_Tariff_Numbers = update_string(input("Do you want download data for this table [Tariff Numbers]? [Y/N]: ")).upper() 
            Download_Data_HQ_Shipping_Condition = update_string(input("Do you want download data for this table [HQ Shipping Condition]? [Y/N]: ")).upper() 

            # Download Setup DataFrame
            Download_Setup_dict = {
                "Download_Data_System": Download_Data_System,
                "Download_Data_NOC": Download_Data_NOC,
                "Download_Data_Shipment_Methods": Download_Data_Shipment_Methods,
                "Download_Data_Shipping_Agents": Download_Data_Shipping_Agents,
                "Download_Data_Shipping_Agens_Service_Codes": Download_Data_Shipping_Agens_Service_Codes,
                "Download_Data_Pland_No": Download_Data_Pland_No,
                "Download_Data_Tariff_Numbers": Download_Data_Tariff_Numbers,
                "Download_Data_HQ_Shipping_Condition": Download_Data_HQ_Shipping_Condition}
            Download_Setup_df = pandas.DataFrame(Download_Setup_dict, columns=Download_Setup_dict.keys(), index=[0])
            Download_Setup_df.Name = "Download_Setup_df"
            print("\n-section summary-")
            print(Download_Setup_df.transpose())

            # Section checker
            Section_check = ""
            while Section_check != "Y" and Section_check != "N":
                Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
            if Section_check == "Y":
                break
            elif Section_check == "N":
                pass

        # Database connection preparation
        if Download_Setup_df.iloc[0]["Download_Data_System"] == "QA":
            server = 'kmnus2qadbs01.erp.kme.intern'
        else:
            server = 'kmnus2prddbs01.erp.kme.intern'
        database = update_string(Download_Setup_df.iloc[0]["Download_Data_NOC"]+Download_Setup_df.iloc[0]["Download_Data_System"])
        dabase_schama = "dbo"
        Connection_string = f'DRIVER=SQL Server;SERVER={server};DATABASE={database};Trusted_Connection=yes;'
        cnxn = pyodbc.connect(Connection_string, autocommit=True)
        cursor = cnxn.cursor()
        
        # Company Selection
        company_list = []
        cursor.execute(f'SELECT [Name] FROM [{database}].[{dabase_schama}].[Company]')
        for row in cursor.fetchall():
            company_list.append(row[0])
        Company = ""
        while Company not in company_list:
            Company = str(input(f"Select Company you want {company_list}: "))

        # Data Download
        Downloaded_Shipment_Method_List = []
        Downloaded_Shipping_Agents_list = []
        Downloaded_Shipping_Agens_Service_Codes_list  = []
        Downloaded_Pland_No_list = []
        Downloaded_Tariff_Numbers_list = []
        Downloaded_HQ_Shipping_Condition_list = []

        # Shipment Methods - done
        if Download_Setup_df.iloc[0]["Download_Data_Shipment_Methods"] == "Y":
            cursor.execute(f'SELECT [Code] FROM [{database}].[{dabase_schama}].[{Company}$Shipment Method]')
            for row in cursor.fetchall():
                Downloaded_Shipment_Method_List.append(row[0])
            Downloaded_Shipment_Method_List = list(set(Downloaded_Shipment_Method_List))

        # Shipping Agents
        if Download_Setup_df.iloc[0]["Download_Data_Shipping_Agents"] == "Y":
            cursor.execute(f'SELECT [Code] FROM [{database}].[{dabase_schama}].[{Company}$Shipping Agent]')
            for row in cursor.fetchall():
                Downloaded_Shipping_Agents_list.append(row[0])
            Downloaded_Shipping_Agents_list = list(set(Downloaded_Shipping_Agents_list))

        # Shipping_Agens_Service_Codes_list
        if Download_Setup_df.iloc[0]["Download_Data_Shipping_Agents"] == "Y" and Download_Setup_df.iloc[0]["Download_Data_Shipping_Agens_Service_Codes"] == "Y":
            cursor.execute(f'SELECT [Shipping Agent Code],[Code] FROM [{database}].[{dabase_schama}].[{Company}$Shipping Agent Services]')
            for row in cursor.fetchall():
                Downloaded_Shipping_Agens_Service_Codes_list.append(row[0])

        # Plant No - Done
        if Download_Setup_df.iloc[0]["Download_Data_Pland_No"] == "Y":
            cursor.execute(f'SELECT [Plant No_] FROM [{database}].[{dabase_schama}].[{Company}$Country_Region]')
            for row in cursor.fetchall():
                Downloaded_Pland_No_list.append(row[0])
            Downloaded_Pland_No_list = list(set(Downloaded_Pland_No_list))

        # Tariff Numbers - Done
        if Download_Setup_df.iloc[0]["Download_Data_Tariff_Numbers"] == "Y":
            cursor.execute(f'SELECT [No_] FROM [{database}].[{dabase_schama}].[{Company}$Tariff Number]')
            for row in cursor.fetchall():
                Downloaded_Tariff_Numbers_list.append(row[0])
            Downloaded_Tariff_Numbers_list = list(set(Downloaded_Tariff_Numbers_list))

        # HQ Shipping Conditions - Done
        if Download_Setup_df.iloc[0]["Download_Data_HQ_Shipping_Condition"] == "Y":
            cursor.execute(f'SELECT [HQ Shipping Condition] FROM [{database}].[{dabase_schama}].[{Company}$HQ Shipping Condition Setup]')
            for row in cursor.fetchall():
                Downloaded_HQ_Shipping_Condition_list.append(row[0])
            Downloaded_HQ_Shipping_Condition_list = list(set(Downloaded_HQ_Shipping_Condition_list))

        cnxn.commit()
        cursor.close()
        
        # Variable to Delete
        Var_to_del = ["Export_File_Location_as_Current", "Download_Data_NOC_list", "Download_Data", "Download_Data_System", "Download_Data_NOC", "Download_Data_Shipment_Methods", "Download_Data_Shipping_Agents", "Download_Data_Shipping_Agens_Service_Codes", "Download_Data_Country_Region_of_Origin_Codes", "Download_Data_Pland_No", "Download_Data_Tariff_Numbers", "Download_Data_HQ_Shipping_Condition", "Download_Setup_dict", "server", "database", "dabase_schama", "Connection_string", "cnxn", "cursor", "row", "company_list", "Company", "General_Setup_Dict"]
        for Variable in Var_to_del:      
            try:
                exec(f'del {Variable}')
            except:
                pass
        try:
            del Variable
            del Var_to_del
        except:
            pass
        
    elif Download_Data == "N":
        print("You will have to select from defaults used in this program")
        # Download Setup DataFrame
        Download_Setup_dict = {
            "Download_Data_Shipment_Methods": "N",
            "Download_Data_Shipping_Agents": "N",
            "Download_Data_Shipping_Agens_Service_Codes": "N",
            "Download_Data_Pland_No": "N",
            "Download_Data_Tariff_Numbers": "N",
            "Download_Data_HQ_Shipping_Condition": "N"}
        Download_Setup_df = pandas.DataFrame(Download_Setup_dict, columns=Download_Setup_dict.keys(), index=[0])
        Download_Setup_df.Name = "Download_Setup_df"
        print("\n-section summary-")
        print(Download_Setup_df.transpose())

    # Purchase Header
    print("\n-----------------------------------------------------")
    print("---Purchase Header Definition---")
    while True:
        Document_Type = "Order"
        # Document Number
        Document_Number_prefix = update_string(input("Define Purchase No. Series [Code]: "))                            
        Document_Number_suffix = update_string(input("Define Purchase No. countable suffix [integer]: "))                   # "str" because of pandas Dataframe
        Document_Number_Increment = update_string(test_integer(input("Define Purchase No. Series increment [integer]: ")))
        Documents_Count = update_string(test_integer(input("How many Purchase Orders you want to create [integer]: ")))

        # Other important
        Buy_from_Vendor_No = update_string(input("Define Vendor Code [Code]: "))  
        if General_Setup_df.iloc[0]["Navision"] == "NUS3": 
            HQ_Logistic_Process = update_string(input("Define HQ Logistic Process Code [Code]: "))   
        else:
            HQ_Logistic_Process = ""
        Document_Location= update_string(input("Define Location Code, used on PO Header [Code]: "))  

        # Order Type
        HQ_Order_Type_list = ["standard", "standard_overnight", "express", "overnight", "normal", "directdelivery", "directdelivery_overnight", "return_order", "return_order_spare_parts", "release", "consignment", "fieldservice", "airlens"]
        HQ_Random_Order_Type = ""
        while HQ_Random_Order_Type != "Y" and HQ_Random_Order_Type != "N":
            HQ_Random_Order_Type = update_string(input(f"Do you want to use random Order Types? Random between these values: {HQ_Order_Type_list}, [Y/N]: ")).upper() 
        if HQ_Random_Order_Type == "Y":
            HQ_Order_Type = ",".join(HQ_Order_Type_list)
        elif HQ_Random_Order_Type == "N":
            HQ_Manual_Order_Type = ""
            while HQ_Manual_Order_Type not in HQ_Order_Type_list:
                HQ_Manual_Order_Type = update_string(input("Select one from above Order Type Code you want to use [Code]: "))
            HQ_Order_Type = list([HQ_Manual_Order_Type])

        # Shipping Condition
        if Download_Setup_df.iloc[0]["Download_Data_HQ_Shipping_Condition"] == "Y":
            HQ_Shipping_Condition = update_string(input(f"Define Shipping Condition {Downloaded_HQ_Shipping_Condition_list} [Code]: "))
        else:
            HQ_Shipping_Condition = update_string(input("Define Shipping Condition [Code]: "))

        # Shipment Method
        if Download_Setup_df.iloc[0]["Download_Data_Shipment_Methods"] == "Y":
            Shipment_Methods = update_string(input(f"Define Shipment Method {Downloaded_Shipment_Method_List} [Code]: "))
        else:
            Shipment_Methods = update_string(input("Define Shipment Method [Code]: "))
        
        # Shipping Agent + Service
        if Download_Setup_df.iloc[0]["Download_Data_Shipping_Agents"] == "Y":
            Shipping_Agent = update_string(input(f"Define Shipping Agent {Downloaded_Shipping_Agents_list} [Code]: "))
        else:
            Shipping_Agent = update_string(input("Define Shipping Agent [Code]: ")) 
            
        if Download_Setup_df.iloc[0]["Download_Data_Shipping_Agens_Service_Codes"] == "Y":
            Shipping_Agent_Service = update_string(input(f"Define Shipping Agent Service {Downloaded_Shipping_Agens_Service_Codes_list} [Code]: "))
        else:
            Shipping_Agent_Service = update_string(input("Define Shipping Agent Service [Code]: "))  

        # Setup Dataframe definition - Purchase Header (+HQ)
        PO_Document_Header_Setup_dict = {
            "Document_Type": Document_Type,
            "Document_Number_prefix": Document_Number_prefix,
            "Document_Number_suffix": Document_Number_suffix,
            "Document_Number_Increment": Document_Number_Increment,
            "Documents_Count": Documents_Count,
            "Buy_from_Vendor_No": Buy_from_Vendor_No,
            "Document_Location": Document_Location,
            "HQ_Logistic_Process": HQ_Logistic_Process,
            "HQ_Random_Order_Type": HQ_Random_Order_Type,
            "HQ_Order_Type": HQ_Order_Type,
            "HQ_Shipping_Condition": HQ_Shipping_Condition,
            "Shipment_Methods": Shipment_Methods,
            "Shipping_Agent": Shipping_Agent,
            "Shipping_Agent_Service": Shipping_Agent_Service}
        PO_Document_Header_Setup_df = pandas.DataFrame(PO_Document_Header_Setup_dict, index=[0])
        PO_Document_Header_Setup_df.Name = "PO_Document_Header_Setup_df"
        print("\n-section summary-")
        print(PO_Document_Header_Setup_df.transpose())

        # Variable to Delete
        Var_to_del = ["Document_Type", "Document_Number_prefix", "Document_Number_suffix", "Document_Number_Increment", "Documents_Count", "Buy_from_Vendor_No", "Document_Location", "HQ_Logistic_Process", "HQ_Order_Type", "HQ_Random_Order_Type", "HQ_Manual_Order_Type", "HQ_Order_Type_list", "HQ_Shipping_Condition", "Shipment_Methods", "Shipping_Agent", "Shipping_Agent_Service", "PO_Document_Header_Setup_dict", "Export_File_Location_as_Current", "Export_File_name", "Navision", "HQ_Date_format", "Questions_dict"]
        for Variable in Var_to_del:      
            try:
                exec(f'del {Variable}')
            except:
                pass
        try:
            del Variable
            del Var_to_del
        except:
            pass

        # Section checker
        Section_check = ""
        while Section_check != "Y" and Section_check != "N":
            Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
        if Section_check == "Y":
            break
        elif Section_check == "N":
            pass

    # Purchase (+HQ) Line
    print("\n-----------------------------------------------------")
    print("---Purchase Line Definition---")
    PO_Lines_Donwload = ""
    while PO_Lines_Donwload != "Y" and PO_Lines_Donwload != "N":
            PO_Lines_Donwload = update_string(input(f"Do you want to donwload Items from concreate Purchase Order? [Y/N]: ")).upper() 
    if PO_Lines_Donwload == "Y":
            import Lib.PO_Line_Donwload as PO_Line_Donwload
            Items_Dict = PO_Line_Donwload.Get_Lines(pandas, pyodbc)
            Items_df = pandas.DataFrame(Items_Dict, columns=Items_Dict.keys())
            Items_df.Name = "Items_df"

            print("\n-section summary-")
            print(Items_df.transpose())

    elif PO_Lines_Donwload == "N":
        while True:
            Item_Random_add = ""
            while Item_Random_add != "Y" and Item_Random_add != "N":
                    Item_Random_add = str(input(f"Do you want to use function of auto add Items from your selection? (Will be add after you confirm completnes of Items) [Y/N]: ")).upper() 
            if Item_Random_add == "Y":
                Items_Random_count = test_integer(input(f"How many Items you want to add at the end of a Dataframe? [integer]: "))
            Item_No = []
            Line_Type = []
            Item_Line_Quantity = []
            Item_Unit_of_Measure = []
            Item_Unit_Price = []
            Item_Free_Of_Charge = []
            Item_Free_Of_Charge_Relation = []
            Item_SN_Tracking = []
            HQ_Confirmation_Line_Flag_Use = []
            HQ_Confirmation_Line_Flag = []
            HQ_SUB_New_Item = []
            Main_BOM_Item = []
            Item_Connected_to_BOM = []
            BOM_Item_Relation = []

            Add_Item = "Y"  
            Item_Counter = 1
            while Add_Item == "Y":
                print(f"\n---Item{Item_Counter}---")
                Item_No.append(str(input(f"Item{Item_Counter}: Define Item No. for every line [Code]: ")))
                Line_Type.append("Item"  )
                Item_Line_Quantity.append(update_string(test_integer(input(f"Item{Item_Counter}: Define Quantity per line [integer]: "))))
                Item_Unit_of_Measure.append(str(input(f"Item{Item_Counter}: Define Unit of Measure for Item [Used in HQ] [Code]: ")))
                Price_text = test_float(str(input(f"Item{Item_Counter}: Define Unit Price for Item [Used in HQ] [float]: ")).replace(",","."))
                Item_Unit_Price.append(Price_text)

                # BOM Item
                if Item_Random_add == "N":
                    Main_BOM_Item_check = ""
                    while Main_BOM_Item_check != "Y" and Main_BOM_Item_check != "N":
                        Main_BOM_Item_check = str(input(f"Item{Item_Counter}: Is Item Main BOM Item which should be ordered from BEU as single one [Y/N]: ")).upper() 
                    if Main_BOM_Item_check == "Y":
                        Main_BOM_Item.append("Y")
                        Item_Connected_to_BOM.append("N")
                        BOM_Item_Relation.append("")
                    if Main_BOM_Item_check == "N":
                        Main_BOM_Item.append("N")
                        Item_Connected_to_BOM_check = ""
                        while Item_Connected_to_BOM_check != "Y" and Item_Connected_to_BOM_check != "N":
                            Item_Connected_to_BOM_check = str(input(f"Item{Item_Counter}: Define if Item is connected to Main BOM Item [Y/N]: ")).upper() 
                        if Item_Connected_to_BOM_check == "Y":
                            Item_Connected_to_BOM.append("Y")
                            BOM_Item_Relation.append(str(input(f"Item{Item_Counter}: To which Item is this Item related {Item_No[:-1]} [code]: ")))
                        if Item_Connected_to_BOM_check == "N":
                            Item_Connected_to_BOM.append("N")
                            BOM_Item_Relation.append("")
                else:
                    Main_BOM_Item.append("N")
                    Item_Connected_to_BOM.append("N")
                    BOM_Item_Relation.append("")

                # Free Of Charge + Tracking
                if Item_Random_add == "N":
                    if Main_BOM_Item_check == "Y":
                        Item_Free_Of_Charge_Check = "N"
                    else:
                        Item_Free_Of_Charge_Check = ""
                    while Item_Free_Of_Charge_Check != "Y" and Item_Free_Of_Charge_Check != "N":
                        Item_Free_Of_Charge_Check = str(input(f"Item{Item_Counter}: Define if Item is Free of Charge [Y/N]: ")).upper() 
                    if Item_Free_Of_Charge_Check == "N":
                        Item_Free_Of_Charge.append("N")
                        Item_Free_Of_Charge_Relation.append("")
                        Item_SN_Tracking_Check = ""
                        while Item_SN_Tracking_Check != "Y" and Item_SN_Tracking_Check != "N":
                            Item_SN_Tracking_Check = str(input(f"Item{Item_Counter}: Define if Item is tracked by SN [Y/N]: ")).upper() 
                        if Item_SN_Tracking_Check == "Y":
                            Item_SN_Tracking.append("Y")
                        elif Item_SN_Tracking_Check == "N":
                            Item_SN_Tracking.append("N")
                    elif Item_Free_Of_Charge_Check == "Y":
                        Item_Free_Of_Charge.append("Y")
                        Item_Free_Con_Check = ""
                        while Item_Free_Con_Check not in Item_No:
                            Item_Free_Con_Check = str(input(f"Item{Item_Counter}: To which Item is this Free of Charge related {Item_No[:-1]} [code]: "))
                        Item_Free_Of_Charge_Relation.append(Item_Free_Con_Check)
                        Item_SN_Tracking.append("N")
                        print("This Item is not going to be placed into Purchase Line --> be updated by Confirmation")
                else:
                    Item_Free_Of_Charge.append("N")
                    Item_Free_Of_Charge_Relation.append("")
                    Item_SN_Tracking_Check = ""
                    while Item_SN_Tracking_Check != "Y" and Item_SN_Tracking_Check != "N":
                        Item_SN_Tracking_Check = str(input(f"Item{Item_Counter}: Define if Item is tracked by SN [Y/N]: ")).upper() 
                    if Item_SN_Tracking_Check == "Y":
                        Item_SN_Tracking.append("Y")
                    elif Item_SN_Tracking_Check == "N":
                        Item_SN_Tracking.append("N")

                # HQ Line Flag
                if Item_Random_add == "N":
                    if Main_BOM_Item_check == "Y":
                        HQ_Confirmation_Line_Flag_Use.append("Y")
                        HQ_Confirmation_Line_Flag.append("Label")
                        HQ_SUB_New_Item.append(str(""))
                    else:
                        HQ_Confirmaiton_Line_Flag_List = ["Substituted", "Cancelled", "Label", "Finished", "To Parent"]
                        HQ_Confirmation_Line_Flag_Use_Check = ""
                        while HQ_Confirmation_Line_Flag_Use_Check != "Y" and HQ_Confirmation_Line_Flag_Use_Check != "N":
                            HQ_Confirmation_Line_Flag_Use_Check = update_string(input(f"Do you want to use Line Flag {HQ_Confirmaiton_Line_Flag_List} [HQ only] [Y/N]: ")).upper() 
                        if HQ_Confirmation_Line_Flag_Use_Check == "Y":
                            HQ_Confirmation_Line_Flag_Use.append(str("Y"))
                            HQ_Confirmation_Line_Flag_set = ""
                            while HQ_Confirmation_Line_Flag_set not in HQ_Confirmaiton_Line_Flag_List:
                                HQ_Confirmation_Line_Flag_set = str(input("Select one from above list [Code]: "))
                            HQ_Confirmation_Line_Flag.append(str(HQ_Confirmation_Line_Flag_set))
                            if HQ_Confirmation_Line_Flag_set == "Substituted":
                                HQ_SUB_New_Item.append(update_string(input(f"Item{Item_Counter}: Define New Item which substitute exported one [Code]: ")))
                            elif HQ_Confirmation_Line_Flag != "Substituted":
                                HQ_SUB_New_Item.append(str(""))
                        elif HQ_Confirmation_Line_Flag_Use_Check == "N":
                            HQ_Confirmation_Line_Flag_Use.append(str("N"))
                            HQ_Confirmation_Line_Flag.append("")
                            HQ_SUB_New_Item.append(str(""))
                else:
                    HQ_Confirmation_Line_Flag_Use.append("N")
                    HQ_Confirmation_Line_Flag.append("")
                    HQ_SUB_New_Item.append("")


                # Add Another Item
                Add_Item = ""
                while Add_Item != "Y" and Add_Item != "N":
                    Add_Item = update_string(input("Do you want to add another Item [Y/N]: ")).upper() 
                if Add_Item == "Y":
                    Item_Counter += 1

            # Setup Dataframe definition - Items
            Items_Dict = {
                "Item_No": Item_No,
                "Line_Type": Line_Type,
                "Item_Line_Quantity": Item_Line_Quantity,
                "Item_Unit_of_Measure": Item_Unit_of_Measure,
                "Item_Unit_Price": Item_Unit_Price,
                "Main_BOM_Item": Main_BOM_Item,
                "Item_Connected_to_BOM": Item_Connected_to_BOM,
                "BOM_Item_Relation": BOM_Item_Relation,
                "Item_Free_Of_Charge": Item_Free_Of_Charge,
                "Item_Free_Of_Charge_Relation": Item_Free_Of_Charge_Relation,
                "Item_SN_Tracking": Item_SN_Tracking,
                "HQ_Confirmation_Line_Flag_Use": HQ_Confirmation_Line_Flag_Use,
                "HQ_Confirmation_Line_Flag": HQ_Confirmation_Line_Flag,
                "HQ_SUB_New_Item": HQ_SUB_New_Item}
            Items_df = pandas.DataFrame(Items_Dict, columns=Items_Dict.keys())
            Items_df.Name = "Items_df"

            # Items randomisation adder
            if Item_Random_add == "Y":
                for i in range(int(Items_Random_count)):
                    index_selection = random.randint(0, Items_df.shape[0] - 1)
                    Item_Series = Items_df.iloc[[index_selection]]
                    Items_df = Items_df.append(Item_Series, ignore_index = True)
            else:
                pass
                    
            print("\n-section summary-")
            print(Items_df.transpose())

            # Variable to Delete
            Var_to_del = ["Add_Item", "Item_Counter", "Item_No", "Line_Type", "Item_Line_Quantity", "Item_Unit_of_Measure", "Price_text", "Item_Unit_Price", "Item_Free_Of_Charge", "Item_Free_Of_Charge_Relation", "Items_Dict", "Item_Free_Of_Charge_Check", "Item_SN_Tracking_Check", "HQ_SUB_New_Item"]
            for Variable in Var_to_del:      
                try:
                    exec(f'del {Variable}')
                except:
                    pass
            try:
                del Variable
                del Var_to_del
            except:
                pass

            # Section checker
            Section_check = ""
            while Section_check != "Y" and Section_check != "N":
                Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
            if Section_check == "Y":
                break
            elif Section_check == "N":
                pass

    print("\n-----------------------------------------------------")
    # HQ Item Transport Registes List
    # Valid for all Document Type
    print("---HQ Global Informaiton Definition---")
    print("---HQ Item Transport Register Definition---")
    while True:
        HQ_HQITR_Register_No_Start = update_string(test_integer(input("Which HQ Register start from (Has to be used the actual max +1 --> because of SQL auto-increment) [integer]: ")))
        HQ_HQITR_Register_No_Increment = "1"
        HQ_Order_Date = test_date((str(input("Define HQ Order Date [DD.MM.YYYY][t - today][ex:28.04.2022]: "))), General_Setup_df.iloc[0]["HQ_Date_format"])
        HQ_Delivery_Start_and_End_Date = test_date((str(input("Define HQ Delivery Start and End Dates [DD.MM.YYYY][t - today][ex:28.04.2022]: "))), General_Setup_df.iloc[0]["HQ_Date_format"])
        HQ_Currency_Code = update_string(input("Define HQ Currency Code [Code]: ")) 
        HQ_Location_Code = update_string(input("Define HQ Location Code [Code]: ")) 
        HQ_Document_Order_No_Start = "10"     # Important for NUS2 too, as it helps define first row of Delivery and import to HQ Tracking Register table
        HQ_Document_Order_No_Increment = "10"
        HQ_Vendor_Line_No_Start = "10"
        HQ_Vendor_Line_No_Increment = "10"

        # Setup Dataframe definition - General
        HQ_General_Setup_Dict = {
            "HQ_HQITR_Register_No_Start": HQ_HQITR_Register_No_Start,
            "HQ_HQITR_Register_No_Increment": HQ_HQITR_Register_No_Increment,
            "HQ_Order_Date": HQ_Order_Date,
            "HQ_Delivery_Start_and_End_Date": HQ_Delivery_Start_and_End_Date,
            "HQ_Currency_Code": HQ_Currency_Code,
            "HQ_Location_Code": HQ_Location_Code,
            "HQ_Document_Order_No_Start": HQ_Document_Order_No_Start,
            "HQ_Document_Order_No_Increment": HQ_Document_Order_No_Increment,
            "HQ_Vendor_Line_No_Start": HQ_Vendor_Line_No_Start,
            "HQ_Vendor_Line_No_Increment": HQ_Vendor_Line_No_Increment}
        HQ_General_Setup_df = pandas.DataFrame(HQ_General_Setup_Dict, index=[0])
        HQ_General_Setup_df.Name = "HQ_General_Setup_df"
        print("\n-section summary-")
        print(HQ_General_Setup_df.transpose())

        # Variable to Delete
        Var_to_del = ["HQ_Order_Date", "HQ_Delivery_Start_and_End_Date", "HQ_Currency_Code", "HQ_Location_Code", "HQ_Document_Order_No_Start", "HQ_Document_Order_No_Increment", "HQ_HQITR_Register_No_Start", "HQ_HQITR_Register_No_Increment", "HQ_Vendor_Line_No_Start", "HQ_Vendor_Line_No_Increment", "HQ_General_Setup_Dict"]
        for Variable in Var_to_del:      
            try:
                exec(f'del {Variable}')
            except:
                pass
        try:
            del Variable
            del Var_to_del
        except:
            pass
        
        # Section checker
        Section_check = ""
        while Section_check != "Y" and Section_check != "N":
            Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
        if Section_check == "Y":
            break
        elif Section_check == "N":
            pass

    # Export
    print("\n-----------------------------------------------------")
    if General_Setup_df.iloc[0]["Navision"] == "NUS3": 
        while Questions_df.iloc[0]["HQ_Export_Quest"] != "Y" and Questions_df.iloc[0]["HQ_Export_Quest"] != "N":
            Questions_df.iloc[0]["HQ_Export_Quest"] = update_string(input("\nDo you want to prepare Expot data [Y/N]: ")).upper() 
        if Questions_df.iloc[0]["HQ_Export_Quest"] == "Y":
            # Export number
            while True:
                HQ_Export_Vendor_Line_No = "0"   

                # Setup Dataframe definition - Export
                HQ_Export_Setup_Dict = {
                    "HQ_Export_Vendor_Line_No": HQ_Export_Vendor_Line_No}
                HQ_Export_Setup_df = pandas.DataFrame(HQ_Export_Setup_Dict, index=[0])
                HQ_Export_Setup_df.Name = "HQ_Export_Setup_df"
                print("\n-section summary-")
                print(HQ_Export_Setup_df.transpose())

                # Variable to Delete
                Var_to_del = ["HQ_Export_Vendor_Line_No", "HQ_Export_Setup_Dict"]
                for Variable in Var_to_del:      
                    try:
                        exec(f'del {Variable}')
                    except:
                        pass
                try:
                    del Variable
                    del Var_to_del
                except:
                    pass

                # Section checker
                Section_check = ""
                while Section_check != "Y" and Section_check != "N":
                    Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
                if Section_check == "Y":
                    break
                elif Section_check == "N":
                    pass
        else:
            HQ_Export_Setup_Dict = {
                "HQ_Export_Vendor_Line_No": "0"}
            HQ_Export_Setup_df = pandas.DataFrame(HQ_Export_Setup_Dict, index=[0])
            HQ_Export_Setup_df.Name = "HQ_Export_Setup_df"
            print("You did't selected Export to be created - skipping")
    elif General_Setup_df.iloc[0]["Navision"] == "NUS2": 
        HQ_Export_Setup_Dict = {
            "HQ_Export_Vendor_Line_No": "0"}
        HQ_Export_Setup_df = pandas.DataFrame(HQ_Export_Setup_Dict, index=[0])
        HQ_Export_Setup_df.Name = "HQ_Export_Setup_df"
        print("NUS2 doesnt have a Export lines in HQ Item Transport Register table")
    else:
        pass

    # Confimraiton
    print("\n-----------------------------------------------------")
    while Questions_df.iloc[0]["HQ_Conf_Quest"] != "Y" and Questions_df.iloc[0]["HQ_Conf_Quest"] != "N":
        Questions_df.iloc[0]["HQ_Conf_Quest"] = update_string(input("Do you want to prepare Confirmation data [Y/N]: ")).upper() 
    if Questions_df.iloc[0]["HQ_Conf_Quest"] == "Y":
        while True:
            # Confirmation Number
            HQ_Confirmation_prefix = update_string(input("Define Vendor Document No. - Prefix for Confirmaiton [Code]: "))
            HQ_Confirmation_suffix = update_string(input("Define Vendor Document No. countable suffix for Confirmaiton [integer]: "))  
            HQ_Confirmation_Increment = update_string(test_integer(input("Define Vendor Document No. Series increment for Confirmaiton [integer]: ")))

            # Confirmation Dates
            HQ_Conf_Vendor_Document_Created_Date = test_date((str(input("Define Vendor Document Created Date for Confirmation [DD.MM.YYYY][t - today][ex:28.04.2022]: "))), General_Setup_df.iloc[0]["HQ_Date_format"])

            # Setup Dataframe definition - Confirmation
            HQ_Confirmation_Setup_Dict = {
                "HQ_Confirmation_prefix": HQ_Confirmation_prefix,
                "HQ_Confirmation_suffix": HQ_Confirmation_suffix,
                "HQ_Confirmation_Increment": HQ_Confirmation_Increment,
                "HQ_Conf_Vendor_Document_Created_Date": HQ_Conf_Vendor_Document_Created_Date} 
            HQ_Confirmation_Setup_df = pandas.DataFrame(HQ_Confirmation_Setup_Dict, index=[0])
            HQ_Confirmation_Setup_df.Name = "HQ_Confirmation_Setup_df"
            print("\n-section summary-")
            print(HQ_Confirmation_Setup_df.transpose())

            # Variable to Delete
            Var_to_del = ["HQ_Confirmation_prefix", "HQ_Confirmation_suffix", "HQ_Confirmation_Increment", "HQ_Conf_Vendor_Document_Created_Date", "HQ_Confirmation_Setup_Dict"]
            for Variable in Var_to_del:      
                try:
                    exec(f'del {Variable}')
                except:
                    pass
            try:
                del Variable
                del Var_to_del
            except:
                pass

            # Section checker
            Section_check = ""
            while Section_check != "Y" and Section_check != "N":
                Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
            if Section_check == "Y":
                break
            elif Section_check == "N":
                pass
    else:
        # Setup Dataframe definition - Confirmation
        HQ_Confirmation_Setup_Dict = {
            "HQ_Confirmation_prefix": "",
            "HQ_Confirmation_suffix": "",
            "HQ_Confirmation_Increment": "",
            "HQ_Conf_Vendor_Document_Created_Date": ""} 
        HQ_Confirmation_Setup_df = pandas.DataFrame(HQ_Confirmation_Setup_Dict, index=[0])
        HQ_Confirmation_Setup_df.Name = "HQ_Confirmation_Setup_df"
        print("You did't selected Confirmation to be created - skipping")

    # PreAdvice
    print("\n-----------------------------------------------------")
    while Questions_df.iloc[0]["HQ_PreAdv_Quest"] != "Y" and Questions_df.iloc[0]["HQ_PreAdv_Quest"] != "N":
        Questions_df.iloc[0]["HQ_PreAdv_Quest"] = update_string(input("Do you want to prepare PreAdvice data [Y/N]: ")).upper() 
    if Questions_df.iloc[0]["HQ_PreAdv_Quest"] == "Y":
        while True:
            # Pre-Advice
            HQ_PreAdvice_prefix = update_string(input("Define Vendor Document No. - Prefix for PreAdvice [Code]: "))
            HQ_PreAdvice_suffix = update_string(input("Define Vendor Document No. countable suffix for PreAdvice [integer]: "))  
            HQ_PreAdvice_Increment = update_string(test_integer(input("Define Vendor Document No. Series increment for PreAdvice [integer]: ")))

            # Pre-Advice Dates
            HQ_PreAdvice_Vendor_Document_Created_Date = test_date((str(input("Define Vendor Document Created Date for PreAdvice [DD.MM.YYYY][t - today][ex:28.04.2022]: "))), General_Setup_df.iloc[0]["HQ_Date_format"])
            print(f"Picking Date is take from SAP Creation Date: {HQ_PreAdvice_Vendor_Document_Created_Date} + 3D for each Deliver created. Picking Date, Trans. Plannig Date, Loading Date, Planned GI Date and Delivery Date will be set +1D from each other.")

            # Setup Dataframe definition - PreAdvice
            HQ_PreAdvice_Setup_Dict = {
                "HQ_PreAdvice_prefix": HQ_PreAdvice_prefix,
                "HQ_PreAdvice_suffix": HQ_PreAdvice_suffix,
                "HQ_PreAdvice_Increment": HQ_PreAdvice_Increment,
                "HQ_PreAdvice_Vendor_Document_Created_Date": HQ_PreAdvice_Vendor_Document_Created_Date}
            HQ_PreAdvice_Setup_df = pandas.DataFrame(HQ_PreAdvice_Setup_Dict, index=[0])
            HQ_PreAdvice_Setup_df.Name = "HQ_PreAdvice_Setup_df"
            print("\n-section summary-")
            print(HQ_PreAdvice_Setup_df.transpose())

            # Variable to Delete
            Var_to_del = ["HQ_PreAdvice_prefix", "HQ_PreAdvice_suffix", "HQ_PreAdvice_Increment", "HQ_PreAdvice_Vendor_Document_Created_Date", "HQ_PreAdvice_Setup_Dict"]
            for Variable in Var_to_del:      
                try:
                    exec(f'del {Variable}')
                except:
                    pass
            try:
                del Variable
                del Var_to_del
            except:
                pass

            # Section checker
            Section_check = ""
            while Section_check != "Y" and Section_check != "N":
                Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
            if Section_check == "Y":
                break
            elif Section_check == "N":
                pass
    else:
        HQ_PreAdvice_Setup_Dict = {
            "HQ_PreAdvice_prefix": "",
            "HQ_PreAdvice_suffix": "",
            "HQ_PreAdvice_Increment": "",
            "HQ_PreAdvice_Vendor_Document_Created_Date": ""}
        HQ_PreAdvice_Setup_df = pandas.DataFrame(HQ_PreAdvice_Setup_Dict, index=[0])
        HQ_PreAdvice_Setup_df.Name = "HQ_PreAdvice_Setup_df"
        print("You did't selected PreAdvice to be created - skipping")

    # Delivery
    print("\n-----------------------------------------------------")
    while Questions_df.iloc[0]["HQ_Del_Quest"] != "Y" and Questions_df.iloc[0]["HQ_Del_Quest"] != "N":
        Questions_df.iloc[0]["HQ_Del_Quest"] = update_string(input("Do you want to prepare Delivery data [Y/N]: ")).upper() 
    if Questions_df.iloc[0]["HQ_Del_Quest"] == "Y":
        while True:
            # Delivery number
            HQ_Delivery_prefix = update_string(input("Define Vendor Document No. - Prefix for Delivery [Code]: "))
            HQ_Delivery_suffix = update_string(input("Define Vendor Document No. countable suffix for Delivery [integer]: "))  
            HQ_Delivery_Increment = update_string(test_integer(input("Define Vendor Document No. Series increment for Delivery [integer]: ")))

            # Delivery Dates
            HQ_Delivery_Receipt_Date = test_date((str(input("Define HQ Receipt Date for Delivery [DD.MM.YYYY][t - today][ex:28.04.2022]: "))), General_Setup_df.iloc[0]["HQ_Date_format"])
            HQ_Delivery_Vendor_Document_Created_Date = test_date((str(input("Define Vendor Document Created Date for Delivery [DD.MM.YYYY][t - today][ex:28.04.2022]: "))), General_Setup_df.iloc[0]["HQ_Date_format"])
            print(f"Picking Date is take from SAP Creation Date: {HQ_Delivery_Vendor_Document_Created_Date} + 3D for each Deliver created. Picking Date, Trans. Plannig Date, Loading Date, Planned GI Date and Delivery Date will be set +1D from each other.")

            # E-Communciation
            HQ_Delviery_E_Com_Process_Stat = ""
            while HQ_Delviery_E_Com_Process_Stat != "Y" and HQ_Delviery_E_Com_Process_Stat != "N":
                HQ_Delviery_E_Com_Process_Stat = update_string(input("Define if E-Comm should be createdfor all Deliveries: [Y/N]: ")).upper() 
            if HQ_Delviery_E_Com_Process_Stat == "Y":
                HQ_Delviery_E_Com_Process_Stat = "Ready To Send"
            elif HQ_Delviery_E_Com_Process_Stat == "N":
                HQ_Delviery_E_Com_Process_Stat = "Sent"

            # Multiple Deliveries + Randome Item and Qty assignment to Delivery
            HQ_Delivery_Multiple_Del = ""
            while HQ_Delivery_Multiple_Del != "Y" and HQ_Delivery_Multiple_Del != "N":
                HQ_Delivery_Multiple_Del = update_string(input("Do you want multiple Deliveries per Order? [Y/N]: ")).upper()  
            if HQ_Delivery_Multiple_Del == "Y":
                # Maximum Delivery Counter = Count of Items and its qty
                Delivery_max = Items_df["Item_Line_Quantity"].apply(int)
                Delivery_Count_per_Order = update_string(test_integer(input(f"How many Deliveries per each Order you want? Max = {Delivery_max} [integer]: ")))
                Delivery_Random_Select = ""
                while Delivery_Random_Select != "Y" and Delivery_Random_Select != "N":
                    Delivery_Random_Select = update_string(input("Do you want randomize Items assignment to delivery (Delivery will contain random Items and Random Qty, Dont use it when BOM or Tower is used)? [Y/N]: ")).upper()  
                print("This will also cause split of Pre-Advice and Invoices.")
            elif HQ_Delivery_Multiple_Del == "N":
                Delivery_Count_per_Order = "1"
                Delivery_Random_Select = "N"
                print("There will be create only one Pre-Advice, Delivery and Invoice per Order.")

            # To-Post: Auto
            if General_Setup_df.iloc[0]["Navision"] == "NUS3": 
                HQ_Delivery_To_Post_Auto = ""
                while HQ_Delivery_To_Post_Auto != "Y" and HQ_Delivery_To_Post_Auto != "N":
                    HQ_Delivery_To_Post_Auto = update_string(input("Should be all deliveries set as To-post - Auto? [Y/N]: ")).upper()  
                if HQ_Delivery_To_Post_Auto == "Y":
                    HQ_Delivery_To_Post_Auto = "True"
                else:
                    HQ_Delivery_To_Post_Auto = "False"
            elif General_Setup_df.iloc[0]["Navision"] == "NUS2": 
                HQ_Delivery_To_Post_Auto = "False"
            else:
                pass

            # Setup Dataframe definition - Delivery
            HQ_Delivery_Setup_Dict = {
                "HQ_Delivery_prefix": HQ_Delivery_prefix,
                "HQ_Delivery_suffix": HQ_Delivery_suffix,
                "HQ_Delivery_Increment": HQ_Delivery_Increment,
                "Delivery_Count_per_Order": Delivery_Count_per_Order,
                "Delivery_Random_Select": Delivery_Random_Select,
                "HQ_Delivery_Receipt_Date": HQ_Delivery_Receipt_Date,
                "HQ_Delivery_Vendor_Document_Created_Date": HQ_Delivery_Vendor_Document_Created_Date,
                "HQ_Delviery_E_Com_Process_Stat": HQ_Delviery_E_Com_Process_Stat,
                "HQ_Delviery_To_Post_Auto": HQ_Delivery_To_Post_Auto}
            HQ_Delivery_Setup_df = pandas.DataFrame(HQ_Delivery_Setup_Dict, index=[0])
            HQ_Delivery_Setup_df.Name = "HQ_Delivery_Setup_df"
            print("\n-section summary-")
            print(HQ_Delivery_Setup_df.transpose())

            # Variable to Delete
            Var_to_del = ["HQ_Delivery_prefix", "HQ_Delivery_suffix", "HQ_Delivery_Increment", "Delivery_Count_per_Order", "Delivery_Random_Select", "HQ_Delivery_Receipt_Date", "HQ_Delivery_Vendor_Document_Created_Date", "HQ_Delviery_E_Com_Process_Stat", "HQ_Delivery_To_Post_Auto", "HQ_Delivery_Setup_Dict"]
            for Variable in Var_to_del:      
                try:
                    exec(f'del {Variable}')
                except:
                    pass
            try:
                del Variable
                del Var_to_del
            except:
                pass

            # Section checker
            Section_check = ""
            while Section_check != "Y" and Section_check != "N":
                Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
            if Section_check == "Y":
                break
            elif Section_check == "N":
                pass
    else:
        HQ_Delivery_Setup_Dict = {
            "HQ_Delivery_prefix": "",
            "HQ_Delivery_suffix": "",
            "HQ_Delivery_Increment": "",
            "Delivery_Count_per_Order": "1",
            "Delivery_Random_Select": "N",
            "HQ_Delivery_Receipt_Date": "",
            "HQ_Delivery_Vendor_Document_Created_Date": "",
            "HQ_Delviery_E_Com_Process_Stat": "",
            "HQ_Delviery_To_Post_Auto": "False"}
        HQ_Delivery_Setup_df = pandas.DataFrame(HQ_Delivery_Setup_Dict, index=[0])
        HQ_Delivery_Setup_df.Name = "HQ_Delivery_Setup_df"
        print("You did't selected Delivery to be created - skipping")

    # Invoice
    print("\n-----------------------------------------------------")
    while Questions_df.iloc[0]["HQ_Inv_Quest"] != "Y" and Questions_df.iloc[0]["HQ_Inv_Quest"] != "N":
        Questions_df.iloc[0]["HQ_Inv_Quest"] = update_string(input("Do you want to prepare Invoice data [Y/N]: ")).upper() 
    if Questions_df.iloc[0]["HQ_Inv_Quest"] == "Y":
        while True:
            # HQ Invoice No
            HQ_Invoice_prefix = update_string(input("Define Vendor Document No. - Prefix for Invoice [Code]: "))
            HQ_Invoice_suffix = update_string(input("Define Vendor Document No. countable suffix for Invoice [integer]: "))  
            HQ_Invoice_Increment = update_string(test_integer(input("Define Vendor Document No. Series increment for Invoice [integer]: ")))

            # Invoice Dates
            HQ_Invoice_Vendor_Document_Created_Date = test_date((str(input("Define Vendor Document Created Date (SAP Posting date) for Invoice [DD.MM.YYYY][t - today][ex:28.04.2022]: "))), General_Setup_df.iloc[0]["HQ_Date_format"])
            
            # Country/Region of Origin 
            HQ_Count_Reg_Origin_list = ["JP", "CN", "RO", "PH", "NO"]
            HQ_Invoice_Random_Count_Reg_Origin = ""
            while HQ_Invoice_Random_Count_Reg_Origin != "Y" and HQ_Invoice_Random_Count_Reg_Origin != "N":
                HQ_Invoice_Random_Count_Reg_Origin = update_string(input(f"Do you want to use random Country/Region of Origin Code? Random between these values: {HQ_Count_Reg_Origin_list}, [Y/N]: ")).upper() 
            if HQ_Invoice_Random_Count_Reg_Origin == "Y":
                HQ_Invoice_Count_Reg_Origin = ",".join(HQ_Count_Reg_Origin_list)
            elif HQ_Invoice_Random_Count_Reg_Origin == "N":
                HQ_Invoice_Manual_Count_Reg_Origin = ""
                while HQ_Invoice_Manual_Count_Reg_Origin not in HQ_Count_Reg_Origin_list:
                    HQ_Invoice_Manual_Count_Reg_Origin = update_string(input(f"Write one Country / Region of Origin Code you want to use {HQ_Count_Reg_Origin_list} [Code]: "))
                HQ_Invoice_Count_Reg_Origin = list([HQ_Invoice_Manual_Count_Reg_Origin])

            # Plant Number
            if Download_Setup_df.iloc[0]["Download_Data_Pland_No"] == "Y":
                HQ_Invoice_Plant_list = Downloaded_Pland_No_list
            else:
                HQ_Invoice_Plant_list = ["1000", "1002", "1004"]
            HQ_Invoice_Random_Plant = ""
            while HQ_Invoice_Random_Plant != "Y" and HQ_Invoice_Random_Plant != "N":
                HQ_Invoice_Random_Plant = update_string(input(f"Do you want to use random Plant? {HQ_Invoice_Plant_list} [Y/N]: ")).upper() 
            if HQ_Invoice_Random_Plant == "Y":
                HQ_Invoice_Plant = ",".join(HQ_Invoice_Plant_list)
            elif HQ_Invoice_Random_Plant == "N":
                HQ_Invoice_Manual_Plant = "9999"
                while HQ_Invoice_Manual_Plant not in HQ_Invoice_Plant_list:
                    HQ_Invoice_Manual_Plant = update_string(input(f"Write one Plant Code you want to use {HQ_Invoice_Plant_list} [Code]: "))
                HQ_Invoice_Plant = list([HQ_Invoice_Manual_Plant])

            # Tariff Numbers
            if General_Setup_df.iloc[0]["Navision"] == "NUS3": 
                if Download_Setup_df.iloc[0]["Download_Data_Tariff_Numbers"] == "Y":
                    HQ_Invoice_Tariff_list = Downloaded_Tariff_Numbers_list
                else:
                    HQ_Invoice_Tariff_list = ["32091000", "32151100", "32151900", "32159000", "34029090"]
                HQ_Invoice_Random_Tariff_No = ""
                while HQ_Invoice_Random_Tariff_No != "Y" and HQ_Invoice_Random_Tariff_No != "N":
                    HQ_Invoice_Random_Tariff_No = update_string(input(f"Do you want to use random Tarif Number Codes? Random between these values: {HQ_Invoice_Tariff_list}, [Y/N]: ")).upper() 
                if HQ_Invoice_Random_Tariff_No == "Y":
                    HQ_Invoice_Tariff = ",".join(HQ_Invoice_Tariff_list)
                elif HQ_Invoice_Random_Tariff_No == "N":
                    HQ_Invoice_Manual_Tariff_No = ""
                    while HQ_Invoice_Manual_Tariff_No not in HQ_Invoice_Tariff_list:
                        HQ_Invoice_Manual_Tariff_No = update_string(input(f"Write one Tariff Code you want to use {HQ_Invoice_Tariff_list} [Code]: "))
                    HQ_Invoice_Tariff = list([HQ_Invoice_Manual_Tariff_No])
            elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
                HQ_Invoice_Random_Tariff_No = "N"
                HQ_Invoice_Tariff = ""
            else:
                pass

            # To-Post: Auto
            if General_Setup_df.iloc[0]["Navision"] == "NUS3": 
                HQ_Invoice_To_Post_Auto = ""
                while HQ_Invoice_To_Post_Auto != "Y" and HQ_Invoice_To_Post_Auto != "N":
                    HQ_Invoice_To_Post_Auto = update_string(input("Should be all Invoices set as To-post - Auto? [Y/N]: ")).upper()  
                if HQ_Invoice_To_Post_Auto == "Y":
                    HQ_Invoice_To_Post_Auto = "True"
                else:
                    HQ_Invoice_To_Post_Auto = "False"
            elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
                 HQ_Invoice_To_Post_Auto = "False"
            else:
                pass

            # Setup Dataframe definition - Invoice
            HQ_Invoice_Setup_Dict = {
                "HQ_Invoice_prefix": HQ_Invoice_prefix,
                "HQ_Invoice_suffix": HQ_Invoice_suffix,
                "HQ_Invoice_Increment": HQ_Invoice_Increment,
                "HQ_Invoice_Vendor_Document_Created_Date": HQ_Invoice_Vendor_Document_Created_Date,
                "HQ_Invoice_Random_Count_Reg_Origin": HQ_Invoice_Random_Count_Reg_Origin,
                "HQ_Invoice_Count_Reg_Origin": HQ_Invoice_Count_Reg_Origin,
                "HQ_Invoice_Random_Plant": HQ_Invoice_Random_Plant,
                "HQ_Invoice_Plant": HQ_Invoice_Plant,
                "HQ_Invoice_Random_Tariff_No": HQ_Invoice_Random_Tariff_No,
                "HQ_Invoice_Tariff": HQ_Invoice_Tariff,
                "HQ_Invoice_To_Post_Auto": HQ_Invoice_To_Post_Auto}
            HQ_Invoice_Setup_df = pandas.DataFrame(HQ_Invoice_Setup_Dict, index=[0])
            HQ_Invoice_Setup_df.Name = "HQ_Invoice_Setup_df"
            print("\n-section summary-")
            print(HQ_Invoice_Setup_df.transpose())

            # Variable to Delete
            Var_to_del = ["HQ_Invoice_prefix", "HQ_Invoice_suffix", "HQ_Invoice_Increment", "HQ_Invoice_Vendor_Document_Created_Date", "HQ_Invoice_Count_Reg_Origin", "HQ_Invoice_Random_Count_Reg_Origin", "HQ_Invoice_Manual_Count_Reg_Origin", "HQ_Count_Reg_Origin_list", "HQ_Invoice_Plant", "HQ_Invoice_Random_Plant", "HQ_Invoice_Manual_Plant", "HQ_Invoice_Plant_list", "HQ_Invoice_Tariff", "HQ_Invoice_Random_Tariff_No", "HQ_Invoice_Manual_Tariff_No", "HQ_Invoice_Tariff_list", "HQ_Invoice_To_Post_Auto", "HQ_Invoice_Setup_Dict"]
            for Variable in Var_to_del:      
                try:
                    exec(f'del {Variable}')
                except:
                    pass
            try:
                del Variable
                del Var_to_del
            except:
                pass

            # Section checker
            Section_check = ""
            while Section_check != "Y" and Section_check != "N":
                Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
            if Section_check == "Y":
                break
            elif Section_check == "N":
                pass

    else:
        HQ_Invoice_Setup_Dict = {
            "HQ_Invoice_prefix":"",
            "HQ_Invoice_suffix": "",
            "HQ_Invoice_Increment": "",
            "HQ_Invoice_Vendor_Document_Created_Date": "",
            "HQ_Invoice_Random_Count_Reg_Origin": "",
            "HQ_Invoice_Count_Reg_Origin": "",
            "HQ_Invoice_Random_Plant": "",
            "HQ_Invoice_Plant": "",
            "HQ_Invoice_Random_Tariff_No": "",
            "HQ_Invoice_Tariff": "",
            "HQ_Invoice_To_Post_Auto": "False"}
        HQ_Invoice_Setup_df = pandas.DataFrame(HQ_Invoice_Setup_Dict, index=[0])
        HQ_Invoice_Setup_df.Name = "HQ_Invoice_Setup_df"
        print("You did't selected Invoice to be created - skipping")

    # HQ ATP check Register
    if Questions_df.iloc[0]["HQ_Conf_Quest"] == "Y":
        print("\n-----------------------------------------------------")
        print("---HQ ATP Check Register Definition---")
        while True:
            # Register ID
            if General_Setup_df.iloc[0]["Navision"] == "NUS3":
                HQ_ATP_Register_No_Start = update_string(test_integer(input("Which HQ Register start from (Has to be used the actual max +1 --> because of SQL auto-increment) [integer]: ")))
                HQ_ATP_Register_No_Increment = "1"
            elif General_Setup_df.iloc[0]["Navision"] == "NUS2": 
                HQ_ATP_Register_No_Start = ""
                HQ_ATP_Register_No_Increment = ""
            else:
                pass

            # Stock Codes
            if General_Setup_df.iloc[0]["Navision"] == "NUS3":
                HQ_ATP_Stock_list = ["ONH", "ONB", "BACK"]
            elif General_Setup_df.iloc[0]["Navision"] == "NUS2": 
                HQ_ATP_Stock_list = ["ONH", "ONB"]
            else:
                pass

            # Random selection
            HQ_ATP_Random_Stock = ""
            while HQ_ATP_Random_Stock != "Y" and HQ_ATP_Random_Stock != "N":
                HQ_ATP_Random_Stock = update_string(input(f"Do you want to use random Stock Codes? Random between these values: {HQ_ATP_Stock_list}, [Y/N]: ")).upper() 
            if HQ_ATP_Random_Stock == "Y":
                HQ_ATP_Stock = ",".join(HQ_ATP_Stock_list)
            elif HQ_ATP_Random_Stock == "N":
                HQ_ATP_Manual_Stock = ""
                while HQ_ATP_Manual_Stock not in HQ_ATP_Stock_list:
                    HQ_ATP_Manual_Stock = update_string(input(f"Write one Stock Code you want to use {HQ_ATP_Stock_list} [Code]: "))
                HQ_ATP_Stock = list([HQ_ATP_Manual_Stock])

            # Multiple records per Qty
            if General_Setup_df.iloc[0]["Navision"] == "NUS3":
                HQ_ATP_Multiple_Lines_Check = ""
                while HQ_ATP_Multiple_Lines_Check != "Y" and HQ_ATP_Multiple_Lines_Check != "N":
                    HQ_ATP_Multiple_Lines_Check = update_string(input(f"Do you want to use multiple lines for ATP Check Register for one Confirmation line (must be more Qty ordered) [Y/N]: ")).upper() 
                if HQ_ATP_Multiple_Lines_Check == "Y":
                    HQ_ATP_Multiple_Lines = update_string(test_integer(input("How many lines in HQ ATP Register for 1 Confirmed line (must be more Qty ordered) [integer]: ")))
                if HQ_ATP_Multiple_Lines_Check == "N":
                    HQ_ATP_Multiple_Lines = "1"
            elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
                HQ_ATP_Multiple_Lines_Check = "N"
                HQ_ATP_Multiple_Lines = "1"
            else:
                pass

            # Scheduled Dates
            HQ_ATP_First_Scheduled_Date = test_date((str(input("Define first Scheduled Date [DD.MM.YYYY][t - today][ex:28.04.2022]: "))), General_Setup_df.iloc[0]["HQ_Date_format"])   
            HQ_ATP_Sched_Date_Increments_day = timedelta(days=int(input("How many days will be used as increments for everyline in HQ ATP Check REgister [integer]: ")))
            HQ_ATP_Zero_Date = test_date((str(input("Define first Zero Date for BackOrdered [DD.MM.YYYY][t - today][f - 01.01.1755][l - 31.12.2099][ex:28.04.2022]: "))), General_Setup_df.iloc[0]["HQ_Date_format"])   
            print(f"Every line in HQ ATP Check Register will hase different Date based on formula: {HQ_ATP_First_Scheduled_Date} + {HQ_ATP_Sched_Date_Increments_day}")

            # Setup Dataframe definition - ATP Register
            HQ_ATP_Setup_Dict = {
                "HQ_ATP_Register_No_Start": HQ_ATP_Register_No_Start,
                "HQ_ATP_Register_No_Increment": HQ_ATP_Register_No_Increment,
                "HQ_ATP_Random_Stock": HQ_ATP_Random_Stock,
                "HQ_ATP_Stock": HQ_ATP_Stock,
                "HQ_ATP_Multiple_Lines_Check": HQ_ATP_Multiple_Lines_Check,
                "HQ_ATP_Multiple_Lines": HQ_ATP_Multiple_Lines,
                "HQ_ATP_First_Scheduled_Date": HQ_ATP_First_Scheduled_Date,
                "HQ_ATP_Sched_Date_Increments_day": HQ_ATP_Sched_Date_Increments_day,
                "HQ_ATP_Zero_Date": HQ_ATP_Zero_Date}
            HQ_ATPR_Setup_df = pandas.DataFrame(HQ_ATP_Setup_Dict, index=[0])
            HQ_ATPR_Setup_df.Name = "HQ_ATPR_Setup_df"
            print("\n-section summary-")
            print(HQ_ATPR_Setup_df.transpose())

            # Variable to Delete
            Var_to_del = ["HQ_ATP_Register_No_Start", "HQ_ATP_Register_No_Increment", "HQ_ATP_Random_Stock", "HQ_ATP_Manual_Stock", "HQ_ATP_Stock_list", "HQ_ATP_Stock", "HQ_ATP_Multiple_Lines", "HQ_ATP_First_Scheduled_Date", "HQ_ATP_Sched_Date_Increments_day", "HQ_ATP_Zero_Date,HQ_ATP_Setup_Dict"]
            for Variable in Var_to_del:      
                try:
                    exec(f'del {Variable}')
                except:
                    pass
            try:
                del Variable
                del Var_to_del
            except:
                pass

            # Section checker
            Section_check = ""
            while Section_check != "Y" and Section_check != "N":
                Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
            if Section_check == "Y":
                break
            elif Section_check == "N":
                pass
    else:
        HQ_ATP_Setup_Dict = {
            "HQ_ATP_Register_No_Start": "0",
            "HQ_ATP_Register_No_Increment": "1",
            "HQ_ATP_Random_Stock": "",
            "HQ_ATP_Stock": "",
            "HQ_ATP_Multiple_Lines_Check": "",
            "HQ_ATP_Multiple_Lines": "",
            "HQ_ATP_First_Scheduled_Date": "",
            "HQ_ATP_Sched_Date_Increments_day": "",
            "HQ_ATP_Zero_Date": ""}
        HQ_ATPR_Setup_df = pandas.DataFrame(HQ_ATP_Setup_Dict, index=[0])
        HQ_ATPR_Setup_df.Name = "HQ_ATPR_Setup_df"

    # HQ Substitution Register
    if General_Setup_df.iloc[0]["Navision"] == "NUS3":
        if Questions_df.iloc[0]["HQ_Conf_Quest"] == "Y":
            if "Substituted" in HQ_Confirmation_Line_Flag:
                print("\n-----------------------------------------------------")
                print("---HQ Substitution Register Definition---")
                while True:
                    pass
                    HQ_SUB_Register_No_Start = update_string(test_integer(input("Which HQ Register start from (Has to be used the actual max +1 --> because of SQL auto-increment) [integer]: ")))
                    HQ_SUB_Register_No_Increment = "1"
                    print("Field: [Time Stamp], will be taken from [HQ_Conf_Vendor_Document_Created_Date]")
                    print("Field: [Substitution - New Item], will be taken from Item which has Line Flag set")

                    # Setup Dataframe definition - Substitution Register
                    HQ_SUB_Setup_Dict = {
                        "HQ_SUB_Register_No_Start": HQ_SUB_Register_No_Start,
                        "HQ_SUB_Register_No_Increment": HQ_SUB_Register_No_Increment}
                    HQ_SUB_Setup_df = pandas.DataFrame(HQ_SUB_Setup_Dict, index=[0])
                    HQ_SUB_Setup_df.Name = "HQ_SUB_Setup_df"
                    print("\n-section summary-")
                    print(HQ_SUB_Setup_df.transpose())

                    # Variable to Delete
                    Var_to_del = ["HQ_SUB_Register_No_Start", "HQ_SUB_Register_No_Increment"]
                    for Variable in Var_to_del:      
                        try:
                            exec(f'del {Variable}')
                        except:
                            pass
                    try:
                        del Variable
                        del Var_to_del
                    except:
                        pass

                    # Section checker
                    Section_check = ""
                    while Section_check != "Y" and Section_check != "N":
                        Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
                    if Section_check == "Y":
                        break
                    elif Section_check == "N":
                        pass
            else:
                HQ_SUB_Setup_Dict = {
                "HQ_SUB_Register_No_Start": "0",
                "HQ_SUB_Register_No_Increment": "1"}
                HQ_SUB_Setup_df = pandas.DataFrame(HQ_SUB_Setup_Dict, index=[0])
                HQ_SUB_Setup_df.Name = "HQ_SUB_Setup_df"
        else:
            HQ_SUB_Setup_Dict = {
                "HQ_SUB_Register_No_Start": "0",
                "HQ_SUB_Register_No_Increment": "1"}
            HQ_SUB_Setup_df = pandas.DataFrame(HQ_SUB_Setup_Dict, index=[0])
            HQ_SUB_Setup_df.Name = "HQ_SUB_Setup_df"
    elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
        HQ_SUB_Setup_Dict = {
            "HQ_SUB_Register_No_Start": "0",
            "HQ_SUB_Register_No_Increment": "1"}
        HQ_SUB_Setup_df = pandas.DataFrame(HQ_SUB_Setup_Dict, index=[0])
        HQ_SUB_Setup_df.Name = "HQ_SUB_Setup_df"
    else:
        pass

    # HQ Pre-Advice Register
    if General_Setup_df.iloc[0]["Navision"] == "NUS3":
        if Questions_df.iloc[0]["HQ_PreAdv_Quest"] == "Y":
            print("\n-----------------------------------------------------")
            print("---HQ PreAdvice Register Definition---")

            while True:
                # Register ID
                HQ_HQPAR_Register_No_Start = update_string(test_integer(input("Which HQ Register start from (Has to be used the actual max +1 --> because of SQL auto-increment) [integer]: ")))
                HQ_HQPAR_Register_No_Increment = "1"

                # Setup Dataframe definition - Pre-Advice Register
                HQ_HQPAR_Setup_Dict = {
                    "HQ_HQPAR_Register_No_Start": HQ_HQPAR_Register_No_Start,
                    "HQ_HQPAR_Register_No_Increment": HQ_HQPAR_Register_No_Increment}
                HQ_HQPAR_Setup_df = pandas.DataFrame(HQ_HQPAR_Setup_Dict, index=[0])
                HQ_HQPAR_Setup_df.Name = "HQ_HQPAR_Setup_df"
                print("\n-section summary-")
                print(HQ_HQPAR_Setup_df.transpose())

                # Variable to Delete
                Var_to_del = ["HQ_HQPAR_Register_No_Start", "HQ_HQPAR_Register_No_Increment", "HQ_HQPAR_Setup_Dict"]
                for Variable in Var_to_del:      
                    try:
                        exec(f'del {Variable}')
                    except:
                        pass
                try:
                    del Variable
                    del Var_to_del
                except:
                    pass

                # Section checker
                Section_check = ""
                while Section_check != "Y" and Section_check != "N":
                    Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
                if Section_check == "Y":
                    break
                elif Section_check == "N":
                    pass
        else:
            HQ_HQPAR_Setup_Dict = {
                "HQ_HQPAR_Register_No_Start": "0",
                "HQ_HQPAR_Register_No_Increment": "1"}
            HQ_HQPAR_Setup_df = pandas.DataFrame(HQ_HQPAR_Setup_Dict, index=[0])
            HQ_HQPAR_Setup_df.Name = "HQ_HQPAR_Setup_df"
    elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
        HQ_HQPAR_Setup_Dict = {
            "HQ_HQPAR_Register_No_Start": "0",
            "HQ_HQPAR_Register_No_Increment": "1"}
        HQ_HQPAR_Setup_df = pandas.DataFrame(HQ_HQPAR_Setup_Dict, index=[0])
        HQ_HQPAR_Setup_df.Name = "HQ_HQPAR_Setup_df"
    else:
        pass

    # HQ Serial Number Register
    if General_Setup_df.iloc[0]["Navision"] == "NUS3":
        if Questions_df.iloc[0]["HQ_Del_Quest"] == "Y":
            if "Y" in Item_SN_Tracking:
                print("\n-----------------------------------------------------")
                print("---HQ Serial Number Register Definition---")
                print("One or more Item/s is defined as SN tracking, so you need to fill defails for HQ SerialNumber Tracking Register")
                while True:
                    # Register ID
                    HQ_HQSNR_Register_No_Start = update_string(test_integer(input("Which HQ Register start from (Has to be used the actual max +1 --> because of SQL auto-increment) [integer]: ")))
                    HQ_HQSNR_Register_No_Increment = "1"

                    # Serial Number No.
                    HQ_SN_SN_prefix = update_string(input("Define prefix SN Series [Code]: "))
                    HQ_SN_SN_suffix = update_string(input("Define SN countable suffix [integer]: "))   
                    HQ_SN_SN_Increment = update_string(test_integer(input("Define SN Series increment [integer]: ")))

                    # Setup Dataframe definition - Serial Number Register
                    HQ_HQSNR_Setup_Dict = {
                        "HQ_HQSNR_Register_No_Start": HQ_HQSNR_Register_No_Start,
                        "HQ_HQSNR_Register_No_Increment": HQ_HQSNR_Register_No_Increment,
                        "HQ_SN_SN_prefix": HQ_SN_SN_prefix,
                        "HQ_SN_SN_suffix": HQ_SN_SN_suffix,
                        "HQ_SN_SN_Increment": HQ_SN_SN_Increment}
                    HQ_HQSNR_Setup_df = pandas.DataFrame(HQ_HQSNR_Setup_Dict, index=[0])
                    HQ_HQSNR_Setup_df.Name = "HQ_HQSNR_Setup_df"
                    print("\n-section summary-")
                    print(HQ_HQSNR_Setup_df.transpose())

                    # Variable to Delete
                    Var_to_del = ["HQ_HQSNR_Register_No_Start", "HQ_HQSNR_Register_No_Increment", "HQ_SN_SN_prefix", "HQ_SN_SN_suffix", "HQ_SN_SN_Increment", "HQ_HQSNR_Setup_Dict"]
                    for Variable in Var_to_del:      
                        try:
                            exec(f'del {Variable}')
                        except:
                            pass
                    try:
                        del Variable
                        del Var_to_del
                    except:
                        pass

                    # Section checker
                    Section_check = ""
                    while Section_check != "Y" and Section_check != "N":
                        Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
                    if Section_check == "Y":
                        break
                    elif Section_check == "N":
                        pass
            else:
                HQ_HQSNR_Setup_Dict = {
                    "HQ_HQSNR_Register_No_Start": "0",
                    "HQ_HQSNR_Register_No_Increment": "1",
                    "HQ_SN_SN_prefix": "",
                    "HQ_SN_SN_suffix": "",
                    "HQ_SN_SN_Increment": ""}
                HQ_HQSNR_Setup_df = pandas.DataFrame(HQ_HQSNR_Setup_Dict, index=[0])
                HQ_HQSNR_Setup_df.Name = "HQ_HQSNR_Setup_df"
                print("No Item with SN tracking selected.")
        else:
            HQ_HQSNR_Setup_Dict = {
                "HQ_HQSNR_Register_No_Start": "0",
                "HQ_HQSNR_Register_No_Increment": "1",
                "HQ_SN_SN_prefix": "",
                "HQ_SN_SN_suffix": "",
                "HQ_SN_SN_Increment": ""}
            HQ_HQSNR_Setup_df = pandas.DataFrame(HQ_HQSNR_Setup_Dict, index=[0])
            HQ_HQSNR_Setup_df.Name = "HQ_HQSNR_Setup_df"
    elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
        if Questions_df.iloc[0]["HQ_Del_Quest"] == "Y":
            if "Y" in Item_SN_Tracking:
                print("\n-----------------------------------------------------")
                print("One or more Item/s is defined as SN tracking, so you need to fill defails for Serial numbers")
                while True:
                    # Serial Number No.
                    HQ_SN_SN_prefix = update_string(input("Define prefix SN Series [Code]: "))
                    HQ_SN_SN_suffix = update_string(input("Define SN countable suffix [integer]: "))   
                    HQ_SN_SN_Increment = update_string(test_integer(input("Define SN Series increment [integer]: ")))

                    # Setup Dataframe definition - Serial Number Register
                    HQ_HQSNR_Setup_Dict = {
                        "HQ_HQSNR_Register_No_Start": "0",
                        "HQ_HQSNR_Register_No_Increment": "1",
                        "HQ_SN_SN_prefix": HQ_SN_SN_prefix,
                        "HQ_SN_SN_suffix": HQ_SN_SN_suffix,
                        "HQ_SN_SN_Increment": HQ_SN_SN_Increment}
                    HQ_HQSNR_Setup_df = pandas.DataFrame(HQ_HQSNR_Setup_Dict, index=[0])
                    HQ_HQSNR_Setup_df.Name = "HQ_HQSNR_Setup_df"
                    print("\n-section summary-")
                    print(HQ_HQSNR_Setup_df.transpose())

                    # Variable to Delete
                    Var_to_del = ["HQ_HQSNR_Register_No_Start", "HQ_HQSNR_Register_No_Increment", "HQ_SN_SN_prefix", "HQ_SN_SN_suffix", "HQ_SN_SN_Increment", "HQ_HQSNR_Setup_Dict"]
                    for Variable in Var_to_del:      
                        try:
                            exec(f'del {Variable}')
                        except:
                            pass
                    try:
                        del Variable
                        del Var_to_del
                    except:
                        pass

                    # Section checker
                    Section_check = ""
                    while Section_check != "Y" and Section_check != "N":
                        Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
                    if Section_check == "Y":
                        break
                    elif Section_check == "N":
                        pass
            else:
                HQ_HQSNR_Setup_Dict = {
                    "HQ_HQSNR_Register_No_Start": "0",
                    "HQ_HQSNR_Register_No_Increment": "1",
                    "HQ_SN_SN_prefix": "",
                    "HQ_SN_SN_suffix": "",
                    "HQ_SN_SN_Increment": ""}
                HQ_HQSNR_Setup_df = pandas.DataFrame(HQ_HQSNR_Setup_Dict, index=[0])
                HQ_HQSNR_Setup_df.Name = "HQ_HQSNR_Setup_df"
                print("No Item with SN tracking selected.")
        else:
            HQ_HQSNR_Setup_Dict = {
                "HQ_HQSNR_Register_No_Start": "0",
                "HQ_HQSNR_Register_No_Increment": "1",
                "HQ_SN_SN_prefix": "",
                "HQ_SN_SN_suffix": "",
                "HQ_SN_SN_Increment": ""}
            HQ_HQSNR_Setup_df = pandas.DataFrame(HQ_HQSNR_Setup_Dict, index=[0])
            HQ_HQSNR_Setup_df.Name = "HQ_HQSNR_Setup_df"
    else:
        HQ_HQSNR_Setup_Dict = {
            "HQ_HQSNR_Register_No_Start": "0",
            "HQ_HQSNR_Register_No_Increment": "1",
            "HQ_SN_SN_prefix": "",
            "HQ_SN_SN_suffix": "",
            "HQ_SN_SN_Increment": ""}
        HQ_HQSNR_Setup_df = pandas.DataFrame(HQ_HQSNR_Setup_Dict, index=[0])
        HQ_HQSNR_Setup_df.Name = "HQ_HQSNR_Setup_df"

    # HQ Delivery Tracking Register
    if General_Setup_df.iloc[0]["Navision"] == "NUS3":
        if Questions_df.iloc[0]["HQ_Del_Quest"] == "Y":
            print("\n-----------------------------------------------------")
            while Questions_df.iloc[0]["HQ_DeL_Track_Quest"] != "Y" and Questions_df.iloc[0]["HQ_DeL_Track_Quest"] != "N":
                Questions_df.iloc[0]["HQ_DeL_Track_Quest"] = update_string(input("Do you want to prepare Delivery Tracking data [Y/N]: ")).upper() 
            if Questions_df.iloc[0]["HQ_DeL_Track_Quest"] == "Y":
                print("---HQ Delivery Tracking Register Definition---")

                while True:
                    # Register ID
                    HQ_HQDTR_Register_No_Start = update_string(test_integer(input("Which HQ Register start from (Has to be used the actual max +1 --> because of SQL auto-increment) [integer]: ")))
                    HQ_HQDTR_Register_No_Increment = "1"

                    # Package Trackin No.
                    HQ_Del_Track_Packa_prefix = update_string(input("Define prefix Package Trackin No. Series [Code]: "))
                    HQ_Del_Track_Packa_suffix = update_string(input("Define Package Trackin No. countable suffix [integer]: "))   
                    HQ_Del_Track_Packa_Increment = update_string(test_integer(input("Define Package Trackin No. Series increment [integer]: ")))

                    # Bill of Landing No.
                    HQ_Del_Track_Bill_prefix = update_string(input("Define prefix Bill of Landing No. Series [Code]: "))
                    HQ_Del_Track_Bill_suffix = update_string(input("Define Bill of Landing No. countable suffix [integer]: "))   
                    HQ_Del_Track_Bill_Increment = update_string(test_integer(input("Define Bill of Landing No. Series increment [integer]: ")))

                    #  EXIDV2
                    HQ_Del_Track_EXIDV_prefix = update_string(input("Define prefix EXIDV2 No. Series [Code]: "))
                    HQ_Del_Track_EXIDV_suffix = update_string(input("Define EXIDV2 No. countable suffix [integer]: "))   
                    HQ_Del_Track_EXIDV_Increment = update_string(test_integer(input("Define EXIDV2No. Series increment [integer]: ")))

                    # Shipment Method
                    if Download_Setup_df.iloc[0]["Download_Data_Shipment_Methods"] == "Y":
                        HQ_Del_Track_Shipment_Method_list = Downloaded_Shipment_Method_List
                    else:
                        HQ_Del_Track_Shipment_Method_list = ["EXW", "CIP", "DPU", "DDP"]
                    HQ_Del_Track_Random_Shipment_Method = ""
                    while HQ_Del_Track_Random_Shipment_Method != "Y" and HQ_Del_Track_Random_Shipment_Method != "N":
                        HQ_Del_Track_Random_Shipment_Method = update_string(input(f"Do you want to use random Icoterms Codes? Random between these values: {HQ_Del_Track_Shipment_Method_list}, [Y/N]: ")).upper() 
                    if HQ_Del_Track_Random_Shipment_Method == "Y":
                        HQ_Del_Track_Shipment_Method = ",".join(HQ_Del_Track_Shipment_Method_list)
                    elif HQ_Del_Track_Random_Shipment_Method == "N":
                        HQ_Del_Track_Manual_Shipment_Method = ""
                        while HQ_Del_Track_Manual_Shipment_Method not in HQ_Del_Track_Shipment_Method_list:
                            HQ_Del_Track_Manual_Shipment_Method = update_string(input(f"Write one Shipment_Method Code you want to use {HQ_Del_Track_Shipment_Method_list} [Code]: "))
                        HQ_Del_Track_Shipment_Method = list([HQ_Del_Track_Manual_Shipment_Method])

                    # Shipping Agents
                    if Download_Setup_df.iloc[0]["Download_Data_Shipping_Agents"] == "Y":
                        HQ_Del_Track_Shipping_Agent_list = Downloaded_Shipping_Agents_list
                    else:
                        HQ_Del_Track_Shipping_Agent_list = ["DHL", "GLS", "SUUS", "GW"]
                    HQ_Del_Track_Random_Shipping_Agent = ""
                    while HQ_Del_Track_Random_Shipping_Agent != "Y" and HQ_Del_Track_Random_Shipping_Agent != "N":
                        HQ_Del_Track_Random_Shipping_Agent = update_string(input(f"Do you want to use random Shipping Agent? Random between these values: {HQ_Del_Track_Shipping_Agent_list}, [Y/N]: ")).upper() 
                    if HQ_Del_Track_Random_Shipping_Agent == "Y":
                        HQ_Del_Track_Shipping_Agent = ",".join(HQ_Del_Track_Shipping_Agent_list)
                    elif HQ_Del_Track_Random_Shipping_Agent == "N":
                        HQ_Del_Track_Manual_Shipping_Agent = ""
                        while HQ_Del_Track_Manual_Shipping_Agent not in HQ_Del_Track_Shipping_Agent_list:
                            HQ_Del_Track_Manual_Shipping_Agent = update_string(input(f"Write one Shipping Agent Code you want to use {HQ_Del_Track_Shipping_Agent_list} [Code]: "))
                        HQ_Del_Track_Shipping_Agent = list([HQ_Del_Track_Manual_Shipping_Agent])

                    # Weight
                    HQ_Del_Track_Random_Weight = ""
                    while HQ_Del_Track_Random_Weight != "Y" and HQ_Del_Track_Random_Weight != "N":
                        HQ_Del_Track_Random_Weight = update_string(input("Do you want to use random Delivery weight? Random between these values:(1 - 100), [Y/N]: ")).upper() 
                    if HQ_Del_Track_Random_Weight == "Y":
                        HQ_Del_Track_Weight = "100"
                    elif HQ_Del_Track_Random_Weight == "N":
                        HQ_Del_Track_Weight = "999"
                        while HQ_Del_Track_Weight == "999":
                            HQ_Del_Track_Weight = update_string(test_integer(input("Set Weight for every Delivery up to 100 [integer]: ")))
                    HQ_Del_Track_Weight_UOM = update_string(input("Set Weight Unit of Measure for every Delivery [Code]: "))

                    # Volume
                    HQ_Del_Track_Random_Volume = ""
                    while HQ_Del_Track_Random_Volume != "Y" and HQ_Del_Track_Random_Volume != "N":
                        HQ_Del_Track_Random_Volume = update_string(input("Do you want to use random Delivery volume? Random between these values:(1 - 100), [Y/N]: ")).upper() 
                    if HQ_Del_Track_Random_Volume == "Y":
                        HQ_Del_Track_Volume = "100"
                    elif HQ_Del_Track_Random_Volume == "N":
                        HQ_Del_Track_Volume = "999"
                        while HQ_Del_Track_Volume == "999":
                            HQ_Del_Track_Volume = update_string(test_integer(input("Set Volume for every Delivery [integer]: ")))
                    HQ_Del_Track_Volume_UOM = update_string(input("Set Volume Unit of Measure for every Delivery [Code]: "))

                    # Setup Dataframe definition - Delivery Tracking Register
                    HQ_DEL_Track_Setup_Dict = {
                        "HQ_HQDTR_Register_No_Start": HQ_HQDTR_Register_No_Start,
                        "HQ_HQDTR_Register_No_Increment": HQ_HQDTR_Register_No_Increment,
                        "HQ_Del_Track_Packa_prefix": HQ_Del_Track_Packa_prefix,
                        "HQ_Del_Track_Packa_suffix": HQ_Del_Track_Packa_suffix,
                        "HQ_Del_Track_Packa_Increment": HQ_Del_Track_Packa_Increment,
                        "HQ_Del_Track_Bill_prefix": HQ_Del_Track_Bill_prefix,
                        "HQ_Del_Track_Bill_suffix": HQ_Del_Track_Bill_suffix,
                        "HQ_Del_Track_Bill_Increment": HQ_Del_Track_Bill_Increment,
                        "HQ_Del_Track_EXIDV_prefix": HQ_Del_Track_EXIDV_prefix,
                        "HQ_Del_Track_EXIDV_suffix": HQ_Del_Track_EXIDV_suffix,
                        "HQ_Del_Track_EXIDV_Increment": HQ_Del_Track_EXIDV_Increment,
                        "HQ_Del_Track_Random_Shipment_Method": HQ_Del_Track_Random_Shipment_Method,
                        "HQ_Del_Track_Shipment_Method": HQ_Del_Track_Shipment_Method,
                        "HQ_Del_Track_Random_Shipping_Agent": HQ_Del_Track_Random_Shipping_Agent,
                        "HQ_Del_Track_Shipping_Agent": HQ_Del_Track_Shipping_Agent,
                        "HQ_Del_Track_Random_Weight": HQ_Del_Track_Random_Weight,
                        "HQ_Del_Track_Weight": HQ_Del_Track_Weight,
                        "HQ_Del_Track_Weight_UOM": HQ_Del_Track_Weight_UOM,
                        "HQ_Del_Track_Random_Volume": HQ_Del_Track_Random_Volume,
                        "HQ_Del_Track_Volume": HQ_Del_Track_Volume,
                        "HQ_Del_Track_Volume_UOM": HQ_Del_Track_Volume_UOM}
                    HQ_DEL_Track_Setup_df = pandas.DataFrame(HQ_DEL_Track_Setup_Dict, index=[0])
                    HQ_DEL_Track_Setup_df.Name = "HQ_DEL_Track_Setup_df"
                    print("\n-section summary-")
                    print(HQ_DEL_Track_Setup_df.transpose())

                    # Variable to Delete
                    Var_to_del = ["HQ_HQDTR_Register_No_Start", "HQ_HQDTR_Register_No_Increment", "HQ_Del_Track_Packa_prefix", "HQ_Del_Track_Packa_suffix", "HQ_Del_Track_Packa_Increment", "HQ_Del_Track_Bill_prefix", "HQ_Del_Track_Bill_suffix", "HQ_Del_Track_Bill_Increment", "HQ_Del_Track_EXIDV_prefix", "HQ_Del_Track_EXIDV_suffix", "HQ_Del_Track_EXIDV_Increment", "HQ_Del_Track_Random_Shipment_Method", "HQ_Del_Track_Random_Shipping_Agent", "HQ_Del_Track_Shipping_Agent_list,HQ_Del_Track_Shipping_Agent,HQ_Del_Track_Manual_Shipment_Method", "HQ_Del_Track_Shipment_Method_list", "HQ_Del_Track_Shipment_Method", "HQ_Del_Track_Random_Weight", "HQ_Del_Track_Weight", "HQ_Del_Track_Weight_UOM", "HQ_Del_Track_Random_Volume", "HQ_Del_Track_Volume", "HQ_Del_Track_Volume_UOM", "HQ_DEL_Track_Setup_Dict"]
                    for Variable in Var_to_del:      
                        try:
                            exec(f'del {Variable}')
                        except:
                            pass
                    try:
                        del Variable
                        del Var_to_del
                    except:
                        pass

                    # Section checker
                    Section_check = ""
                    while Section_check != "Y" and Section_check != "N":
                        Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
                    if Section_check == "Y":
                        break
                    elif Section_check == "N":
                        pass
            else:
                HQ_DEL_Track_Setup_Dict = {
                    "HQ_HQDTR_Register_No_Start": "0",
                    "HQ_HQDTR_Register_No_Increment": "1",
                    "HQ_Del_Track_Packa_prefix": "",
                    "HQ_Del_Track_Packa_suffix": "",
                    "HQ_Del_Track_Packa_Increment": "",
                    "HQ_Del_Track_Bill_prefix": "",
                    "HQ_Del_Track_Bill_suffix": "",
                    "HQ_Del_Track_Bill_Increment": "",
                    "HQ_Del_Track_EXIDV_prefix": "",
                    "HQ_Del_Track_EXIDV_suffix": "",
                    "HQ_Del_Track_EXIDV_Increment": "",
                    "HQ_Del_Track_Random_Shipment_Method": "",
                    "HQ_Del_Track_Shipment_Method": "",
                    "HQ_Del_Track_Random_Shipping_Agent": "",
                    "HQ_Del_Track_Shipping_Agent": "",
                    "HQ_Del_Track_Random_Weight": "",
                    "HQ_Del_Track_Weight": "",
                    "HQ_Del_Track_Weight_UOM": "",
                    "HQ_Del_Track_Random_Volume": "",
                    "HQ_Del_Track_Volume": "",
                    "HQ_Del_Track_Volume_UOM": ""}
                HQ_DEL_Track_Setup_df = pandas.DataFrame(HQ_DEL_Track_Setup_Dict, index=[0])
                HQ_DEL_Track_Setup_df.Name = "HQ_DEL_Track_Setup_df"
                print("You did't selected Delivery Tracking to be created - skipping")
        else:
            HQ_DEL_Track_Setup_Dict = {
                "HQ_HQDTR_Register_No_Start": "0",
                "HQ_HQDTR_Register_No_Increment": "1",
                "HQ_Del_Track_Packa_prefix": "",
                "HQ_Del_Track_Packa_suffix": "",
                "HQ_Del_Track_Packa_Increment": "",
                "HQ_Del_Track_Bill_prefix": "",
                "HQ_Del_Track_Bill_suffix": "",
                "HQ_Del_Track_Bill_Increment": "",
                "HQ_Del_Track_EXIDV_prefix": "",
                "HQ_Del_Track_EXIDV_suffix": "",
                "HQ_Del_Track_EXIDV_Increment": "",
                "HQ_Del_Track_Random_Shipment_Method": "",
                "HQ_Del_Track_Shipment_Method": "",
                "HQ_Del_Track_Random_Shipping_Agent": "",
                "HQ_Del_Track_Shipping_Agent": "",
                "HQ_Del_Track_Random_Weight": "",
                "HQ_Del_Track_Weight": "",
                "HQ_Del_Track_Weight_UOM": "",
                "HQ_Del_Track_Random_Volume": "",
                "HQ_Del_Track_Volume": "",
                "HQ_Del_Track_Volume_UOM": ""}
            HQ_DEL_Track_Setup_df = pandas.DataFrame(HQ_DEL_Track_Setup_Dict, index=[0])
            HQ_DEL_Track_Setup_df.Name = "HQ_DEL_Track_Setup_df"
    elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
        if Questions_df.iloc[0]["HQ_Del_Quest"] == "Y":
            print("\n-----------------------------------------------------")
            while Questions_df.iloc[0]["HQ_DeL_Track_Quest"] != "Y" and Questions_df.iloc[0]["HQ_DeL_Track_Quest"] != "N":
                Questions_df.iloc[0]["HQ_DeL_Track_Quest"] = update_string(input("Do you want to prepare Delivery Tracking data [Y/N]: ")).upper() 
            if Questions_df.iloc[0]["HQ_DeL_Track_Quest"] == "Y":
                print("---HQ Transfer Tracking---")
                while True:
                    # Package Trackin No.
                    HQ_Del_Track_Packa_prefix = update_string(input("Define prefix Package Trackin No. Series [Code]: "))
                    HQ_Del_Track_Packa_suffix = update_string(input("Define Package Trackin No. countable suffix [integer]: "))   
                    HQ_Del_Track_Packa_Increment = update_string(test_integer(input("Define Package Trackin No. Series increment [integer]: ")))
                    
                    # Setup Dataframe definition - Delivery Tracking Register
                    HQ_DEL_Track_Setup_Dict = {
                        "HQ_HQDTR_Register_No_Start": "0",
                        "HQ_HQDTR_Register_No_Increment": "1",
                        "HQ_Del_Track_Packa_prefix": HQ_Del_Track_Packa_prefix,
                        "HQ_Del_Track_Packa_suffix": HQ_Del_Track_Packa_suffix,
                        "HQ_Del_Track_Packa_Increment": HQ_Del_Track_Packa_Increment,
                        "HQ_Del_Track_Bill_prefix": "",
                        "HQ_Del_Track_Bill_suffix": "",
                        "HQ_Del_Track_Bill_Increment": "",
                        "HQ_Del_Track_EXIDV_prefix": "",
                        "HQ_Del_Track_EXIDV_suffix": "",
                        "HQ_Del_Track_EXIDV_Increment": "",
                        "HQ_Del_Track_Random_Shipment_Method": "",
                        "HQ_Del_Track_Shipment_Method": "",
                        "HQ_Del_Track_Random_Shipping_Agent": "",
                        "HQ_Del_Track_Shipping_Agent": "",
                        "HQ_Del_Track_Random_Weight": "",
                        "HQ_Del_Track_Weight": "",
                        "HQ_Del_Track_Weight_UOM": "",
                        "HQ_Del_Track_Random_Volume": "",
                        "HQ_Del_Track_Volume": "",
                        "HQ_Del_Track_Volume_UOM": ""}
                    HQ_DEL_Track_Setup_df = pandas.DataFrame(HQ_DEL_Track_Setup_Dict, index=[0])
                    HQ_DEL_Track_Setup_df.Name = "HQ_DEL_Track_Setup_df"
                    print("\n-section summary-")
                    print(HQ_DEL_Track_Setup_df.transpose())

                    # Variable to Delete
                    Var_to_del = ["HQ_HQDTR_Register_No_Start", "HQ_HQDTR_Register_No_Increment", "HQ_Del_Track_Packa_prefix", "HQ_Del_Track_Packa_suffix", "HQ_Del_Track_Packa_Increment", "HQ_Del_Track_Bill_prefix", "HQ_Del_Track_Bill_suffix", "HQ_Del_Track_Bill_Increment", "HQ_Del_Track_EXIDV_prefix", "HQ_Del_Track_EXIDV_suffix", "HQ_Del_Track_EXIDV_Increment", "HQ_Del_Track_Random_Shipment_Method", "HQ_Del_Track_Random_Shipping_Agent", "HQ_Del_Track_Shipping_Agent_list,HQ_Del_Track_Shipping_Agent,HQ_Del_Track_Manual_Shipment_Method", "HQ_Del_Track_Shipment_Method_list", "HQ_Del_Track_Shipment_Method", "HQ_Del_Track_Random_Weight", "HQ_Del_Track_Weight", "HQ_Del_Track_Weight_UOM", "HQ_Del_Track_Random_Volume", "HQ_Del_Track_Volume", "HQ_Del_Track_Volume_UOM", "HQ_DEL_Track_Setup_Dict"]
                    for Variable in Var_to_del:      
                        try:
                            exec(f'del {Variable}')
                        except:
                            pass
                    try:
                        del Variable
                        del Var_to_del
                    except:
                        pass

                    # Section checker
                    Section_check = ""
                    while Section_check != "Y" and Section_check != "N":
                        Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
                    if Section_check == "Y":
                        break
                    elif Section_check == "N":
                        pass
            else:
                HQ_DEL_Track_Setup_Dict = {
                    "HQ_HQDTR_Register_No_Start": "0",
                    "HQ_HQDTR_Register_No_Increment": "1",
                    "HQ_Del_Track_Packa_prefix": "",
                    "HQ_Del_Track_Packa_suffix": "",
                    "HQ_Del_Track_Packa_Increment": "",
                    "HQ_Del_Track_Bill_prefix": "",
                    "HQ_Del_Track_Bill_suffix": "",
                    "HQ_Del_Track_Bill_Increment": "",
                    "HQ_Del_Track_EXIDV_prefix": "",
                    "HQ_Del_Track_EXIDV_suffix": "",
                    "HQ_Del_Track_EXIDV_Increment": "",
                    "HQ_Del_Track_Random_Shipment_Method": "",
                    "HQ_Del_Track_Shipment_Method": "",
                    "HQ_Del_Track_Random_Shipping_Agent": "",
                    "HQ_Del_Track_Shipping_Agent": "",
                    "HQ_Del_Track_Random_Weight": "",
                    "HQ_Del_Track_Weight": "",
                    "HQ_Del_Track_Weight_UOM": "",
                    "HQ_Del_Track_Random_Volume": "",
                    "HQ_Del_Track_Volume": "",
                    "HQ_Del_Track_Volume_UOM": ""}
                HQ_DEL_Track_Setup_df = pandas.DataFrame(HQ_DEL_Track_Setup_Dict, index=[0])
                HQ_DEL_Track_Setup_df.Name = "HQ_DEL_Track_Setup_df"
                print("You did't selected Delivery Tracking to be created - skipping")
        else:
            HQ_DEL_Track_Setup_Dict = {
                "HQ_HQDTR_Register_No_Start": "0",
                "HQ_HQDTR_Register_No_Increment": "1",
                "HQ_Del_Track_Packa_prefix": "",
                "HQ_Del_Track_Packa_suffix": "",
                "HQ_Del_Track_Packa_Increment": "",
                "HQ_Del_Track_Bill_prefix": "",
                "HQ_Del_Track_Bill_suffix": "",
                "HQ_Del_Track_Bill_Increment": "",
                "HQ_Del_Track_EXIDV_prefix": "",
                "HQ_Del_Track_EXIDV_suffix": "",
                "HQ_Del_Track_EXIDV_Increment": "",
                "HQ_Del_Track_Random_Shipment_Method": "",
                "HQ_Del_Track_Shipment_Method": "",
                "HQ_Del_Track_Random_Shipping_Agent": "",
                "HQ_Del_Track_Shipping_Agent": "",
                "HQ_Del_Track_Random_Weight": "",
                "HQ_Del_Track_Weight": "",
                "HQ_Del_Track_Weight_UOM": "",
                "HQ_Del_Track_Random_Volume": "",
                "HQ_Del_Track_Volume": "",
                "HQ_Del_Track_Volume_UOM": ""}
            HQ_DEL_Track_Setup_df = pandas.DataFrame(HQ_DEL_Track_Setup_Dict, index=[0])
            HQ_DEL_Track_Setup_df.Name = "HQ_DEL_Track_Setup_df"
    else:
        pass

    # HQ Package Tracking Register
    if General_Setup_df.iloc[0]["Navision"] == "NUS3":
        if Questions_df.iloc[0]["HQ_Del_Quest"] == "Y":
            print("\n-----------------------------------------------------")
            while Questions_df.iloc[0]["HQ_Packg_Track_Quest"] != "Y" and Questions_df.iloc[0]["HQ_Packg_Track_Quest"] != "N":
                Questions_df.iloc[0]["HQ_Packg_Track_Quest"] = update_string(input("Do you want to prepare Package Tracking data [Y/N]: ")).upper() 
            if Questions_df.iloc[0]["HQ_Packg_Track_Quest"] == "Y":
                print("---HQ Package Tracking Register Definition---")
                while True:
                    # Register ID
                    HQ_HQPTR_Register_No_Start = update_string(test_integer(input("Which HQ Register start from (Has to be used the actual max +1 --> because of SQL auto-increment) [integer]: ")))
                    HQ_HQPTR_Register_No_Increment = "1"

                    # Packages number
                    HQ_Packg_Track_Packages_per_Delivery = update_string(test_integer(input("Set Numbes of Packages per one Delivery [integer]: ")))
                    print("Package weight and Volume will be counted as Delivery Totals / Number of package you selecte, UOM is same")
                    print("External Pakage Id = EXIDV2 from HQ Delivery Tracking Register")

                    # Setup Dataframe definition - Delivery Tracking Register
                    HQ_Packg_Track_Setup_Dict = {
                        "HQ_HQPTR_Register_No_Start": HQ_HQPTR_Register_No_Start,
                        "HQ_HQPTR_Register_No_Increment": HQ_HQPTR_Register_No_Increment,
                        "HQ_Packg_Track_Packages_per_Delivery": HQ_Packg_Track_Packages_per_Delivery}
                    HQ_Packg_Track_Setup_df = pandas.DataFrame(HQ_Packg_Track_Setup_Dict, index=[0])
                    HQ_Packg_Track_Setup_df.Name = "HQ_Packg_Track_Setup_df"
                    print("\n-section summary-")
                    print(HQ_Packg_Track_Setup_df.transpose())

                    # Variable to Delete
                    Var_to_del = ["HQ_HQPTR_Register_No_Start", "HQ_HQPTR_Register_No_Increment", "HQ_Packg_Track_Packages_per_Delivery", "HQ_Packg_Track_Setup_Dict"]
                    for Variable in Var_to_del:      
                        try:
                            exec(f'del {Variable}')
                        except:
                            pass
                    try:
                        del Variable
                        del Var_to_del
                    except:
                        pass

                    # Section checker
                    Section_check = ""
                    while Section_check != "Y" and Section_check != "N":
                        Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
                    if Section_check == "Y":
                        break
                    elif Section_check == "N":
                        pass
            else:
                HQ_Packg_Track_Setup_Dict = {
                    "HQ_HQPTR_Register_No_Start": "0",
                    "HQ_HQPTR_Register_No_Increment": "1",
                    "HQ_Packg_Track_Packages_per_Delivery": ""}
                HQ_Packg_Track_Setup_df = pandas.DataFrame(HQ_Packg_Track_Setup_Dict, index=[0])
                HQ_Packg_Track_Setup_df.Name = "HQ_Packg_Track_Setup_df"
                print("You did't selected Package Tracking to be created - skipping")
        else:
            HQ_Packg_Track_Setup_Dict = {
                "HQ_HQPTR_Register_No_Start": "0",
                "HQ_HQPTR_Register_No_Increment": "1",
                "HQ_Packg_Track_Packages_per_Delivery": ""}
            HQ_Packg_Track_Setup_df = pandas.DataFrame(HQ_Packg_Track_Setup_Dict, index=[0])
            HQ_Packg_Track_Setup_df.Name = "HQ_Packg_Track_Setup_df"
    elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
        HQ_Packg_Track_Setup_Dict = {
            "HQ_HQPTR_Register_No_Start": "0",
            "HQ_HQPTR_Register_No_Increment": "1",
            "HQ_Packg_Track_Packages_per_Delivery": ""}
        HQ_Packg_Track_Setup_df = pandas.DataFrame(HQ_Packg_Track_Setup_Dict, index=[0])
        HQ_Packg_Track_Setup_df.Name = "HQ_Packg_Track_Setup_df"
    else:
        pass

    # Record Links
    print("\n-----------------------------------------------------")
    if General_Setup_df.iloc[0]["Export_File_Type"] == "EXCEL":
        while Questions_df.iloc[0]["Record_link_Quest"] != "Y" and Questions_df.iloc[0]["Record_link_Quest"] != "N":
            Questions_df.iloc[0]["Record_link_Quest"] = update_string(input("Do you want to prepare .pdf Record Link [Y/N]: ")).upper() 
        if Questions_df.iloc[0]["Record_link_Quest"] == "Y":
            while True:
                # General
                NUS3_NOC_list = ["COREQA", "BDK", "BPL", "BHR", "BSL", "BIH", "BRS", "BR"]
                Record_Links_NOC = ""
                while Record_Links_NOC not in NUS3_NOC_list:
                    Record_Links_NOC = update_string(input(f"Which NOC you want to select? {NUS3_NOC_list} [Code]: "))
                Record_Links_Server = ""
                while Record_Links_Server != "QA" and Record_Links_Server != "PRD":
                    Record_Links_Server = update_string(input("Which DB you want to select as data source? [QA/PRD]: "))
                Record_Links_User_ID = update_string(input("Define your User ID which is in NAV [string]: "))
                Record_Links_Company = update_string(input("Define your Company which is in NAV you want to prepare links [string]: "))

                # Setup Dataframe definition - Record Links
                Record_Links_Setup_Dict = {
                    "NOC": Record_Links_NOC,
                    "Server": Record_Links_Server,
                    "Server_link": "kmnavfs02.bs.kme.intern",
                    "User ID": Record_Links_User_ID,
                    "Company": Record_Links_Company} 
                Record_Links_Setup_df = pandas.DataFrame(Record_Links_Setup_Dict, index=[0])
                Record_Links_Setup_df.Name = "Record_Links_Setup_df"
                print("\n-section summary-")
                print(Record_Links_Setup_df.transpose())

                # Variable to Delete
                Var_to_del = []
                for Variable in Var_to_del:      
                    try:
                        exec(f'del {Variable}')
                    except:
                        pass
                try:
                    del Variable
                    del Var_to_del
                except:
                    pass

                # Section checker
                Section_check = ""
                while Section_check != "Y" and Section_check != "N":
                    Section_check = update_string(input(f"Are all values correct? [Y/N]: ")).upper() 
                if Section_check == "Y":
                    break
                elif Section_check == "N":
                    pass
        else:
            # Setup Dataframe definition - Record Links
            Record_Links_Setup_Dict = {
                "NOC": "",
                "Server": "",
                "Server_link": "",
                "User ID": "",
                "Company": ""} 
            Record_Links_Setup_df = pandas.DataFrame(Record_Links_Setup_Dict, index=[0])
            Record_Links_Setup_df.Name = "Record_Links_Setup_df"
            print("You did't selected Record Links to be created - skipping")
    else:
        # Setup Dataframe definition - Record Links
        Record_Links_Setup_Dict = {
            "NOC": "",
            "Server": "",
            "Server_link": "",
            "User ID": "",
            "Company": ""} 
        Record_Links_Setup_df = pandas.DataFrame(Record_Links_Setup_Dict, index=[0])
        Record_Links_Setup_df.Name = "Record_Links_Setup_df"

    # JSON template safe
    print("\n-----------------------------------------------------")
    JSON_safe = ""
    while JSON_safe != "Y" and JSON_safe != "N":
        JSON_safe = update_string(input(f"Do You want to save setup in reloadable JSON file [Y/N]: ")).upper() 
    if JSON_safe == "Y":
        JSON_Location_current = ""
        while JSON_Location_current != "Y" and JSON_Location_current != "N":
            JSON_Location_current = update_string(input("Do you want to create JSON in the same directory as current script[Y/N]: ")).upper() 
        if JSON_Location_current == "Y":
            JSON_path = os.path.dirname(os.path.abspath(__file__))
            JSON_help_path = str(JSON_path) + str(r"/JSON/")
            JSON_help_path = JSON_help_path.replace("\\","/")
        elif JSON_Location_current == "N":
            JSON_path = str(input("Please define full path [string]: "))
            JSON_help_path= str(JSON_path) + str(r"/JSON/")
            JSON_help_path = JSON_help_path.replace("\\","/")
        JSON_Save_File_name = update_string(input("Define your file name[string]: "))

        # Creation of "/JSON/" help folder
        try:
            os.mkdir(JSON_help_path)
        except FileExistsError:
            pass

        # Preparation of each setup files
        General_Setup_df.to_json(f'{JSON_help_path}/{General_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        Questions_df.to_json(f'{JSON_help_path}/{Questions_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        PO_Document_Header_Setup_df.to_json(f'{JSON_help_path}/{PO_Document_Header_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        HQ_General_Setup_df.to_json(f'{JSON_help_path}/{HQ_General_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        HQ_Export_Setup_df.to_json(f'{JSON_help_path}/{HQ_Export_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        HQ_Confirmation_Setup_df.to_json(f'{JSON_help_path}/{HQ_Confirmation_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        HQ_PreAdvice_Setup_df.to_json(f'{JSON_help_path}/{HQ_PreAdvice_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        HQ_Delivery_Setup_df.to_json(f'{JSON_help_path}/{HQ_Delivery_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        HQ_Invoice_Setup_df.to_json(f'{JSON_help_path}/{HQ_Invoice_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        HQ_ATPR_Setup_df.to_json(f'{JSON_help_path}/{HQ_ATPR_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        HQ_SUB_Setup_df.to_json(f'{JSON_help_path}/{HQ_SUB_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        HQ_HQPAR_Setup_df.to_json(f'{JSON_help_path}/{HQ_HQPAR_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        HQ_HQSNR_Setup_df.to_json(f'{JSON_help_path}/{HQ_HQSNR_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        HQ_DEL_Track_Setup_df.to_json(f'{JSON_help_path}/{HQ_DEL_Track_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        HQ_Packg_Track_Setup_df.to_json(f'{JSON_help_path}/{HQ_Packg_Track_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")
        Record_Links_Setup_df.to_json(f'{JSON_help_path}/{HQ_Packg_Track_Setup_df.Name}.json', indent=4, date_format=General_Setup_df.iloc[0]["HQ_Date_format"], force_ascii=True, lines=True, orient="records")

        Items_Dict = {
            "Item_No": list(Items_df.iloc[:]["Item_No"]),
            "Line_Type": list(Items_df.iloc[:]["Line_Type"]),
            "Item_Line_Quantity": list(Items_df.iloc[:]["Item_Line_Quantity"]),
            "Item_Unit_of_Measure": list(Items_df.iloc[:]["Item_Unit_of_Measure"]),
            "Item_Unit_Price": list(Items_df.iloc[:]["Item_Unit_Price"]),
            "Main_BOM_Item": list(Items_df.iloc[:]["Main_BOM_Item"]),
            "Item_Connected_to_BOM": list(Items_df.iloc[:]["Item_Connected_to_BOM"]),
            "BOM_Item_Relation": list(Items_df.iloc[:]["BOM_Item_Relation"]),
            "Item_Free_Of_Charge": list(Items_df.iloc[:]["Item_Free_Of_Charge"]),
            "Item_Free_Of_Charge_Relation": list(Items_df.iloc[:]["Item_Free_Of_Charge_Relation"]),
            "Item_SN_Tracking": list(Items_df.iloc[:]["Item_SN_Tracking"]),
            "HQ_Confirmation_Line_Flag_Use": list(Items_df.iloc[:]["HQ_Confirmation_Line_Flag_Use"]),
            "HQ_Confirmation_Line_Flag": list(Items_df.iloc[:]["HQ_Confirmation_Line_Flag"]),
            "HQ_SUB_New_Item": list(Items_df.iloc[:]["HQ_SUB_New_Item"])}

        # Prepare JSON helper file
        json_object = json.dumps(Items_Dict, separators=(',', ':'), indent="") 
        Json_list = json_object.split("\n")

        # Export JSON Helper to Setup Dataframe 
        with open(f"{JSON_help_path}/Items_df.json", mode="tw") as file:
            for line in Json_list:
                founded = line.find("{") 
                if line.find("{") == 0:
                    file.write(str("    ")+line+str("\n"))
                elif line.find("],") == 0: 
                    file.write(line+str("\n"))
                elif line.find("]") == 0: 
                    file.write(line)
                elif line.find(":[") > 0:
                    file.write(str("\t\t")+line)
                elif line.find("}") == 0:
                    file.write(str("\n    ")+line)
                    file.write(str("\n"))
                else:
                    file.write(line)
        
        # Concatenation into one setup into one folder
        with open(f"{JSON_path}/{JSON_Save_File_name}.json", mode="tw") as file:
            file.write("""{"Setup": \n""")
            file.write("\t{\n")
            concatenate_dataframe("General_Setup_df", JSON_help_path)
            concatenate_dataframe("Questions_df", JSON_help_path)
            concatenate_dataframe("PO_Document_Header_Setup_df", JSON_help_path)
            concatenate_dataframe("Items_df", JSON_help_path)
            concatenate_dataframe("HQ_General_Setup_df", JSON_help_path)
            concatenate_dataframe("HQ_Export_Setup_df", JSON_help_path)
            concatenate_dataframe("HQ_Confirmation_Setup_df", JSON_help_path)
            concatenate_dataframe("HQ_PreAdvice_Setup_df", JSON_help_path)
            concatenate_dataframe("HQ_Delivery_Setup_df", JSON_help_path)
            concatenate_dataframe("HQ_Invoice_Setup_df", JSON_help_path)
            concatenate_dataframe("HQ_ATPR_Setup_df", JSON_help_path)
            concatenate_dataframe("HQ_SUB_Setup_df", JSON_help_path)
            concatenate_dataframe("HQ_HQPAR_Setup_df", JSON_help_path)
            concatenate_dataframe("HQ_HQSNR_Setup_df", JSON_help_path)
            concatenate_dataframe("HQ_DEL_Track_Setup_df", JSON_help_path)
            concatenate_dataframe("HQ_Packg_Track_Setup_df", JSON_help_path)
            concatenate_dataframe("Record_Links_Setup_df", JSON_help_path)

        # deletion of "/JSON/" help folder and content
        try:
            shutil.rmtree(JSON_help_path, ignore_errors=True)
            try:
                os.rmdir(path=JSON_help_path) 
            except:
                pass
            print("Help directory deletion ...")
        except:
            pass

    if JSON_safe == "N":
        print("Your data are not going to be safed.")

# Generator
# Purchase Header List preparation
Document_Number_Increment = 0
Export_PH_Document_Type_list = []
Export_PH_Document_No_list = []
Export_PH_Buy_from_Vendor_No_list = []
Export_PH_Document_Location_list = []
Export_PH_HQ_Logistic_Process_list = []
Export_PH_HQ_Order_Type_list = []
Export_PH_HQ_Shipping_Condition_list = []
Export_PH_Shippment_Method_list = []
Export_PH_Shipping_Agent_list = []
Export_PH_Shipping_Agent_Service_list = []

# Puchase ORder Line List Preparation
Export_PL_Document_Type_list = []
Export_PL_Document_No_list = []
Export_PL_Line_No_list = []
Export_PL_Type_list = []
Export_PL_No_list = []
Export_PL_Quantity_list = []

# HQ Item Transport Regsiter List preparation
Export_HQ_ITR_Register_No_list = []
Export_HQ_ITR_Register_Sub_No_list = []
Export_HQ_ITR_Document_No_list = []
Export_HQ_ITR_Document_Line_list = []
Export_HQ_ITR_Document_Type_list = []
Export_HQ_ITR_Vendor_Document_Type_list = []
Export_HQ_ITR_Vendor_Document_No_list = []
Export_HQ_ITR_Vendor_Shipment_No_list = []
Export_HQ_ITR_Line_Type_list = []
Export_HQ_ITR_Item_No_list = []
Export_HQ_ITR_Vendor_Item_No_list = []
Export_HQ_ITR_Ordered_Item_No_list = []
Export_HQ_ITR_Quantity_list = []
Export_HQ_ITR_Ordered_Quantity_list = []
Export_HQ_ITR_Serial_No_list = []
Export_HQ_ITR_Unit_of_Measure_list = []
Export_HQ_ITR_Currency_Code_list = []
Export_HQ_ITR_Unit_Price_list = []
Export_HQ_ITR_Line_Amount_list = []
Export_HQ_ITR_Order_Date_list = []
Export_HQ_ITR_Receipt_Date_list = []
Export_HQ_ITR_Posting_Date_list = []
Export_HQ_ITR_Exported_Line_No_list = []
Export_HQ_ITR_Vendor_Line_No_list = []
Export_HQ_ITR_Vendor_No_list = []
Export_HQ_ITR_Country_Region_of_Origin_Code_list = []
Export_HQ_ITR_Plant_No_list = []
Export_HQ_ITR_Line_Flag_list = []
Export_HQ_ITR_Communication_Process_Status_list = []
Export_HQ_ITR_ATP_Check_Cumulative_Quantity_list = []
Export_HQ_ITR_Vendor_Document_Created_Date_list = []
Export_HQ_ITR_Delivery_Start_Date_list = []
Export_HQ_ITR_Delivery_End_Date_list = []
Export_HQ_ITR_Tariff_Number_list = []
Export_HQ_ITR_Vendor_Delivery_No_list = []
Export_HQ_ITR_Quantity_to_Deliver_list = []
Export_HQ_ITR_Vendor_Invoice_No_list = []
Export_HQ_ITR_Quantity_to_Invoice_list = []
Export_HQ_ITR_Picking_Date_list = []
Export_HQ_ITR_Trans_Planning_Date_list = []
Export_HQ_ITR_Loading_Date_list = []
Export_HQ_ITR_Planned_GI_Date_list = []
Export_HQ_ITR_Delivery_Date_list = []
Export_HQ_ITR_Document_Order_Number_list = []
Export_HQ_ITR_To_Post_Auto_list = []

# HQ ATP Register List preparation
Export_HQ_ATP_HQ_ATP_Check_Register_ID_list = []
Export_HQ_ATP_HQ_ATP_Check_Sub_Register_ID_list = []
Export_HQ_ATP_HQ_Item_Transport_Register_ID_list = []
Export_HQ_ATP_Scheduled_Quantity_list = []
Export_HQ_ATP_Scheduled_Date_list = []
Export_HQ_ATP_Stock_list = []
Export_HQ_ATP_Document_No_list = []

# HQ Substitution Register List preparation
Export_HQ_SUB_HQ_Substitution_Register_ID_list = []
Export_HQ_SUB_HQ_Substitution_Sub_Reg_ID_list = []
Export_HQ_SUB_Substituted_Item_Old_Item_list = []
Export_HQ_SUB_Substitution_New_Item_list = []
Export_HQ_SUB_Document_No_list = []

# HQ PreAdvice Register List preparation
Export_HQ_PREA_HQ_PreAdvice_Register_ID_list = []
Export_HQ_PREA_HQ_PreAdvice_Sub_Register_ID_list = []
Export_HQ_PREA_HQ_Item_Transport_Register_ID_list = []
Export_HQ_PREA_HQ_Item_Transport_Reg_Sub_ID_list = []
Export_HQ_PREA_Document_No_list = []
Export_HQ_PREA_PreAdvice_Document_No_list = []
Export_HQ_PREA_PreAdvice_Date_list = []

# HQ Serial Number REgsiter List preparation
Export_HQ_SNR_HQ_SN_Register_ID_list = []
Export_HQ_SNR_HQ_SN_Sub_Register_ID_list = []
Export_HQ_SNR_HQ_Item_Transport_Register_ID_list = []
Export_HQ_SNR_HQ_Item_Transport_Sub_Reg_ID_list = []
Export_HQ_SNR_Purchase_Order_No_list = []
Export_HQ_SNR_Item_No_list = []
Export_HQ_SNR_Serial_Number_list = []
Export_HQ_SNR_Vendor_Document_No_list = []

# HQ Delivery Tracking Register
Export_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID_list = []
Export_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID_list = []
Export_HQ_HQDTR_Delivery_No_list = []
Export_HQ_HQDTR_Purchase_Order_No_list = []
Export_HQ_HQDTR_Package_Tracking_No_list = []
Export_HQ_HQDTR_Package_Type_list = []
Export_HQ_HQDTR_BEU_Bill_of_Landing_list = []
Export_HQ_HQDTR_Tracking_Page_list = []
Export_HQ_HQDTR_Incoterms_list = []
Export_HQ_HQDTR_Total_Weight_list = []
Export_HQ_HQDTR_Weight_Unit_list = []
Export_HQ_HQDTR_Volume_list = []
Export_HQ_HQDTR_Volume_Unit_list = []
Export_HQ_HQDTR_Shipping_Agent_Code_list = []
Export_HQ_HQDTR_EXIDV2_list = []

# HQ Package Tracking Register
Export_HQ_HQPTR_HQ_Package_Tracking_Register_ID_list = []
Export_HQ_HQPTR_HQ_Package_Tracking_Sub_Register_ID_list = []
Export_HQ_HQPTR_Delivery_No_list = []
Export_HQ_HQPTR_Package_No_list = []
Export_HQ_HQPTR_Item_list = []
Export_HQ_HQPTR_Quantity_list = []
Export_HQ_HQPTR_Unit_of_Measure_list = []
Export_HQ_HQPTR_Purchase_Order_No_list = []
Export_HQ_HQPTR_External_Package_ID_list = []
Export_HQ_HQPTR_Total_Weight_list = []
Export_HQ_HQPTR_Weight_Unit_list = []
Export_HQ_HQPTR_Volume_list = []
Export_HQ_HQPTR_Volume_Unit_list = []
Export_HQ_HQPTR_Plant_No_list = []

# Record Links
Link_ID_list = []
Record_ID_list = []
URL1_list = []
URL2_list = []
URL3_list = []
URL4_list = []
Description_list = []
Type_list = []
Note_list = []
Created_list = []
User_ID_list = []
Company_list = []
Notify_list = []
To_User_ID_list = []

# Numbe Series Preparation - Start numbers + helpers
if General_Setup_df.iloc[0]["Navision"] == "NUS3":
    Set_HQATPR_Register_No = int(HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_Register_No_Start"])
    Set_HQSUBR_Register_No = int(HQ_SUB_Setup_df.iloc[0]["HQ_SUB_Register_No_Start"])
    Set_HQPREA_Regsiter_No = int(HQ_HQPAR_Setup_df.iloc[0]["HQ_HQPAR_Register_No_Start"])
    Set_HQSNR_Register_No = int(HQ_HQSNR_Setup_df.iloc[0]["HQ_HQSNR_Register_No_Start"])
    Set_HQDTR_Regsiter_No = int(HQ_DEL_Track_Setup_df.iloc[0]["HQ_HQDTR_Register_No_Start"])
    Set_HQPTR_Regsiter_No = int(HQ_Packg_Track_Setup_df.iloc[0]["HQ_HQPTR_Register_No_Start"])

Set_HQITR_Register_No = int(HQ_General_Setup_df.iloc[0]["HQ_HQITR_Register_No_Start"])

Document_Number_Sufix_Counter = len(PO_Document_Header_Setup_df.iloc[0]["Document_Number_suffix"])
Current_PO_Number = int(PO_Document_Header_Setup_df.iloc[0]["Document_Number_suffix"])

HQ_Confirmation_Number_Sufix_Counter = len(HQ_Confirmation_Setup_df.iloc[0]["HQ_Confirmation_suffix"])
Current_HQ_Confirmation_Number = int(HQ_Confirmation_Setup_df.iloc[0]["HQ_Confirmation_suffix"])

HQ_PreAdvice_Number_Sufix_Counter = len(HQ_PreAdvice_Setup_df.iloc[0]["HQ_PreAdvice_suffix"])
Current_HQ_PreAdvice_Number = int(HQ_PreAdvice_Setup_df.iloc[0]["HQ_PreAdvice_suffix"])

HQ_Delivery_Number_Sufix_Counter = len(HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_suffix"])
Current_HQ_Delivery_Number = int(HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_suffix"])

HQ_Invoice_Number_Sufix_Counter = len(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_suffix"])
Current_HQ_Invoice_Number = int(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_suffix"])

try:
    HQ_SN_Sufix_Counter = len(HQ_HQSNR_Setup_df.iloc[0]["HQ_SN_SN_suffix"])
    Current_SN_Number = int(HQ_HQSNR_Setup_df.iloc[0]["HQ_SN_SN_suffix"])
except:
    HQ_SN_Sufix_Counter = "0"
    Current_SN_Number = ""

HQ_Package_Sufix_Counter = len(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Packa_suffix"])
Current_Package_Number = int(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Packa_suffix"])

if General_Setup_df.iloc[0]["Navision"] == "NUS3":
    HQ_Bill_Sufix_Counter = len(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Bill_suffix"])
    Current_Bill_Number = int(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Bill_suffix"])

    HQ_EXIDV_Sufix_Counter = len(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_EXIDV_suffix"])
    Current_EXIDV_Number = int(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_EXIDV_suffix"])


# Purchase Header:
print("\n-----------------------------------------------------")
for Current_PO_Number in tqdm(range(int(PO_Document_Header_Setup_df.iloc[0]["Documents_Count"])), desc="Generating"):
    Current_PO_Suffix = str(int(PO_Document_Header_Setup_df.iloc[0]["Document_Number_suffix"]) + Current_PO_Number).zfill(Document_Number_Sufix_Counter)
    Current_HQ_Conf_Number_Suffix = str(Current_HQ_Confirmation_Number).zfill(HQ_Confirmation_Number_Sufix_Counter)
    # Prepare Current data
    Current_PH_Document_Type = str(PO_Document_Header_Setup_df.iloc[0]["Document_Type"])
    Current_PH_Document_No = str(PO_Document_Header_Setup_df.iloc[0]["Document_Number_prefix"])+str(Current_PO_Suffix)
    Current_PH_Buy_from_Vendor_No = str(PO_Document_Header_Setup_df.iloc[0]["Buy_from_Vendor_No"])
    Current_PH_Document_Location = str(PO_Document_Header_Setup_df.iloc[0]["Document_Location"])
    Current_PH_HQ_Logistic_Process = str(PO_Document_Header_Setup_df.iloc[0]["HQ_Logistic_Process"])
    Current_PH_HQ_Order_Type = random.choice(str(PO_Document_Header_Setup_df.iloc[0]["HQ_Order_Type"]).split(","))    # Random
    Current_PH_HQ_Shipping_Condition = random.choice(str(PO_Document_Header_Setup_df.iloc[0]["HQ_Shipping_Condition"]).split(","))
    Current_PH_Shipment_Methods = random.choice(str(PO_Document_Header_Setup_df.iloc[0]["Shipment_Methods"]).split(","))
    Current_PH_Shipping_Agent = random.choice(str(PO_Document_Header_Setup_df.iloc[0]["Shipping_Agent"]).split(","))
    Current_PH_Shipping_Agent_Service = random.choice(str(PO_Document_Header_Setup_df.iloc[0]["Shipping_Agent_Service"]).split(","))

    # Prepare PO Header Export List
    Export_PH_Document_Type_list.append(str(Current_PH_Document_Type))
    Export_PH_Document_No_list.append(str(Current_PH_Document_No))
    Export_PH_Buy_from_Vendor_No_list.append(str(Current_PH_Buy_from_Vendor_No))
    Export_PH_Document_Location_list.append(str(Current_PH_Document_Location))
    Export_PH_HQ_Logistic_Process_list.append(str(Current_PH_HQ_Logistic_Process))
    Export_PH_HQ_Order_Type_list.append(str(Current_PH_HQ_Order_Type))
    Export_PH_HQ_Shipping_Condition_list.append(str(Current_PH_HQ_Shipping_Condition))
    Export_PH_Shippment_Method_list.append(str(Current_PH_Shipment_Methods))
    Export_PH_Shipping_Agent_list.append(str(Current_PH_Shipping_Agent))
    Export_PH_Shipping_Agent_Service_list.append(str(Current_PH_Shipping_Agent_Service))

    # Purchase Lines
    Current_PL_Line_No = 10000
    for item in Items_df.iterrows():
        Current_Item = pandas.Series(data = item[1])
        # Prelare Current Purchase Lines
        if Current_Item["Item_Free_Of_Charge"] != "Y" and Current_Item["Item_Connected_to_BOM"] != "Y":
            Current_PL_Document_Type = Current_PH_Document_Type
            Current_PL_Document_No = Current_PH_Document_No
            Current_PL_Line_No = Current_PL_Line_No
            Current_PL_Type = Current_Item["Line_Type"] 
            Current_PL_No = Current_Item["Item_No"] 
            Current_PL_Quantity = Current_Item["Item_Line_Quantity"] 
        else: 
            continue
        # Prepare PO Line Export lists
        Export_PL_Document_Type_list.append(str(Current_PL_Document_Type))
        Export_PL_Document_No_list.append(str(Current_PL_Document_No))
        Export_PL_Line_No_list.append(str(Current_PL_Line_No))
        Export_PL_Type_list.append(str(Current_PL_Type))
        Export_PL_No_list.append(str(Current_PL_No))
        Export_PL_Quantity_list.append(str(Current_PL_Quantity))

        # Final PO Line Setup
        Current_PL_Line_No += 10000

    # Export line preparation
    Exported_Lines_dict = {
        "Purchase_Order": Export_PL_Document_No_list,
        "Item_No": Export_PL_No_list,
        "Line_Number":Export_PL_Line_No_list,
        "Line_Qty": Export_PL_Quantity_list}
    Exported_Lines_df = pandas.DataFrame(data=Exported_Lines_dict, columns=Exported_Lines_dict.keys())
    Exported_Lines_df.Name = "Exported_Lines_df"

    # Delivery Randomize / Delivery assign all lines
    DEL_Purchase_Order = []
    DEL_Item_No = []
    DEL_Line_Number = []
    DEL_Line_Qty = []
    DEL_PreAdvice_No = []
    DEL_Delivery_No = []
    DEL_Invoice_No = []
    DEL_Counter = []

    DEL_Help_df = Items_df.join(Exported_Lines_df["Line_Number"])
    DEL_Help_df.Name = "DEL_Help_df"

    Items_set = Items_df["Item_Line_Quantity"].apply(int)
    Total_Order_Qty = sum(Items_set)
    Max_Delivery_Qty = Total_Order_Qty // int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"])

    if HQ_Delivery_Setup_df.iloc[0]["Delivery_Random_Select"] == "Y":
        for counter in range(int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"])):
            Current_Delivery_Qty = 0
            Current_HQ_PreAdvice_Suffix = str(Current_HQ_PreAdvice_Number).zfill(HQ_PreAdvice_Number_Sufix_Counter)
            Current_HQ_Delivery_Suffix = str(Current_HQ_Delivery_Number).zfill(HQ_Delivery_Number_Sufix_Counter)
            Current_HQ_Invoice_Suffix = str(Current_HQ_Invoice_Number).zfill(HQ_Invoice_Number_Sufix_Counter)
            if counter != (int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"]) - 1):
                while Current_Delivery_Qty < Max_Delivery_Qty:
                    # Define random Item No from List of Items (helper object)
                    Random_df_index = random.randint(0,int(DEL_Help_df.shape[0] - 1))
                    # Is Qty > 0
                    if int(DEL_Help_df.iloc[Random_df_index]["Item_Line_Quantity"]) > 0:
                        Current_DEL_Extported_Line = DEL_Help_df.iloc[Random_df_index]["Line_Number"]
                        Current_DEL_Item_No = DEL_Help_df.iloc[Random_df_index]["Item_No"]
                        Current_DEL_Quantity = random.randint(1, int(DEL_Help_df.iloc[Random_df_index]["Item_Line_Quantity"]))
                        
                        # Compare total Qty --> Cannot be higher than "Max_Delivery_Qty"
                        if  Current_Delivery_Qty + Current_DEL_Quantity > Max_Delivery_Qty:
                            Current_DEL_Quantity = Max_Delivery_Qty - Current_Delivery_Qty
                        else:
                            pass

                        # Export data preparation
                        DEL_Purchase_Order.append(str(Current_PH_Document_No))
                        DEL_Item_No.append(str(Current_DEL_Item_No))
                        DEL_Line_Number.append(str(Current_DEL_Extported_Line))
                        DEL_Line_Qty.append(str(Current_DEL_Quantity))
                        DEL_PreAdvice_No.append(str(HQ_PreAdvice_Setup_df.iloc[0]["HQ_PreAdvice_prefix"])+str(Current_HQ_PreAdvice_Suffix))
                        DEL_Delivery_No.append(str(HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_prefix"])+str(Current_HQ_Delivery_Suffix))
                        DEL_Invoice_No.append(str(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_prefix"])+str(Current_HQ_Invoice_Suffix))
                        DEL_Counter.append(counter)

                        # Final line
                        Current_Delivery_Qty += Current_DEL_Quantity
                        DEL_Help_df.at[Random_df_index, "Item_Line_Quantity"] = int(DEL_Help_df.iloc[Random_df_index]["Item_Line_Quantity"]) - Current_DEL_Quantity
                    else:
                        continue
            else:
                # Last Delivery --> to assign rest of Qty
                mask1 = DEL_Help_df["Item_Line_Quantity"].apply(int) > 0 
                DEL_Help_df2 = DEL_Help_df.where(mask1)
                DEL_Help_df2.dropna(inplace=True)

                for row in DEL_Help_df2.iterrows():		
                    row_df = pandas.Series(data = row[1]) 
                    DEL_Purchase_Order.append(str(Current_PH_Document_No))
                    DEL_Item_No.append(str(row_df["Item_No"]))
                    DEL_Line_Number.append(str(row_df["Line_Number"]))
                    DEL_Line_Qty.append(str(row_df["Item_Line_Quantity"]))
                    DEL_PreAdvice_No.append(str(HQ_PreAdvice_Setup_df.iloc[0]["HQ_PreAdvice_prefix"])+str(Current_HQ_PreAdvice_Suffix))
                    DEL_Delivery_No.append(str(HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_prefix"])+str(Current_HQ_Delivery_Suffix))
                    DEL_Invoice_No.append(str(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_prefix"])+str(Current_HQ_Invoice_Suffix))
                    DEL_Counter.append(counter)

            # Final line
            Current_HQ_PreAdvice_Number += int(HQ_PreAdvice_Setup_df.iloc[0]["HQ_PreAdvice_Increment"])
            Current_HQ_Delivery_Number += int(HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_Increment"])
            Current_HQ_Invoice_Number += int(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_Increment"])


    elif HQ_Delivery_Setup_df.iloc[0]["Delivery_Random_Select"] == "N":
        # Assign all lines to each delivery Qty / number of Delivery
        for counter in range(int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"])):
            Current_Delivery_Qty = 0
            Current_HQ_PreAdvice_Suffix = str(Current_HQ_PreAdvice_Number).zfill(HQ_PreAdvice_Number_Sufix_Counter)
            Current_HQ_Delivery_Suffix = str(Current_HQ_Delivery_Number).zfill(HQ_Delivery_Number_Sufix_Counter)
            Current_HQ_Invoice_Suffix = str(Current_HQ_Invoice_Number).zfill(HQ_Invoice_Number_Sufix_Counter)
            if counter != (int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"]) - 1):
                for row in DEL_Help_df.iterrows():
                    row_df = pandas.Series(data = row[1]) 
                    DEL_Purchase_Order.append(str(Current_PH_Document_No))
                    DEL_Item_No.append(str(row_df["Item_No"]))
                    DEL_Line_Number.append(str(row_df["Line_Number"]))
                    DEL_Line_Qty.append(str(int(row_df["Item_Line_Quantity"])//int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"])))
                    DEL_PreAdvice_No.append(str(HQ_PreAdvice_Setup_df.iloc[0]["HQ_PreAdvice_prefix"])+str(Current_HQ_PreAdvice_Suffix))
                    DEL_Delivery_No.append(str(HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_prefix"])+str(Current_HQ_Delivery_Suffix))
                    DEL_Invoice_No.append(str(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_prefix"])+str(Current_HQ_Invoice_Suffix))
                    DEL_Counter.append(counter)
            else:
                # Last Delivery --> to assign rest of Qty
                for row in DEL_Help_df.iterrows():
                    row_df = pandas.Series(data = row[1]) 
                    DEL_Purchase_Order.append(str(Current_PH_Document_No))
                    DEL_Item_No.append(str(row_df["Item_No"]))
                    DEL_Line_Number.append(str(row_df["Line_Number"]))
                    DEL_Line_Qty.append(str(Rest_Qty(int(row_df["Item_Line_Quantity"]),int(int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"])))))
                    DEL_PreAdvice_No.append(str(HQ_PreAdvice_Setup_df.iloc[0]["HQ_PreAdvice_prefix"])+str(Current_HQ_PreAdvice_Suffix))
                    DEL_Delivery_No.append(str(HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_prefix"])+str(Current_HQ_Delivery_Suffix))
                    DEL_Invoice_No.append(str(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_prefix"])+str(Current_HQ_Invoice_Suffix))
                    DEL_Counter.append(counter)

            # Final line
            Current_HQ_PreAdvice_Number += int(HQ_PreAdvice_Setup_df.iloc[0]["HQ_PreAdvice_Increment"])
            Current_HQ_Delivery_Number += int(HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_Increment"])
            Current_HQ_Invoice_Number += int(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_Increment"])

    else:
        print("Internal Error - No Random selection of Delivery to Order")

    DEL_Item_Assign_dict = {
        "DEL_Purchase_Order": DEL_Purchase_Order,
        "DEL_Item_No": DEL_Item_No,
        "DEL_Line_Number": DEL_Line_Number,
        "DEL_Line_Qty": DEL_Line_Qty,
        "DEL_PreAdvice_No": DEL_PreAdvice_No,
        "DEL_Delivery_No": DEL_Delivery_No,
        "DEL_Invoice_No": DEL_Invoice_No,
        "DEL_Counter": DEL_Counter}
    DEL_Help_Item_Assign_df = pandas.DataFrame(data=DEL_Item_Assign_dict, columns=DEL_Item_Assign_dict.keys())
    DEL_Help_Item_Assign_df.Name = "DEL_Help_Item_Assign_df"

    # Sum Line Qty for same Item in the one Delivery
    DEL_Help_Item_Assign_df["DEL_Line_Qty"] = DEL_Help_Item_Assign_df["DEL_Line_Qty"].apply(int)
    DEL_Help_Item_Assign_df["DEL_Line_Number"] = DEL_Help_Item_Assign_df["DEL_Line_Number"].apply(int)
    DEL_Help_Item_Assign_df["Total"] = DEL_Help_Item_Assign_df.groupby(["DEL_Delivery_No", "DEL_Line_Number", "DEL_Item_No"])["DEL_Line_Qty"].transform("sum")
    DEL_Help_Item_Assign_df.drop_duplicates(subset=["DEL_Delivery_No", "DEL_Line_Number", "DEL_Item_No"] ,inplace=True)
    DEL_Help_Item_Assign_df.sort_values(["DEL_Purchase_Order", "DEL_Delivery_No", "DEL_Line_Number"], axis = 0, ascending = True, inplace = True)
    DEL_Help_Item_Assign_df["DEL_Line_Qty"] = DEL_Help_Item_Assign_df["Total"]
    DEL_Help_Item_Assign_df["DEL_Line_Qty"] = DEL_Help_Item_Assign_df["DEL_Line_Qty"].apply(str)
    DEL_Help_Item_Assign_df["DEL_Line_Number"] = DEL_Help_Item_Assign_df["DEL_Line_Number"].apply(str)
    DEL_Help_Item_Assign_df.drop(axis=1, inplace=True, labels="Total")
    DEL_Help_Item_Assign_df.reset_index(drop=True ,inplace=True)

    # Add Colums from Items_df
    DEL_Item_Assign_df = DEL_Help_Item_Assign_df.merge(Items_df,left_on="DEL_Item_No", right_on="Item_No")

    # Variable to Delete
    Var_to_del = ["Exported_Lines_dict", "DEL_Purchase_Order", "DEL_Item_No", "DEL_Line_Number", "DEL_Line_Qty", "DEL_PreAdvice_No", "DEL_Delivery_No", "DEL_Invoice_No", "DEL_Item_Assign_dict", "Current_HQ_PreAdvice_Suffix", "Current_HQ_Delivery_Suffix", "Current_HQ_Invoice_Suffix", "counter", "Current_DEL_Item_No", "Current_DEL_Extported_Line", "Current_DEL_Quantity", "Current_Delivery_Qty", "DEL_Help_Item_Assign_df", "DEL_Help_df", "DEL_Help_df2", "row", "mask1", "Items_set"]
    for Variable in Var_to_del:      
        try:
            exec(f'del {Variable}')
        except:
            pass
    try:
        del Variable
        del Var_to_del
    except:
        pass

    # HQ Item Transport Register - Export
    if General_Setup_df.iloc[0]["Navision"] == "NUS3":
        if Questions_df.iloc[0]["HQ_Export_Quest"] == "Y":
            HQ_Document_Order_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
            for item in Items_df.iterrows():
                Current_Item = pandas.Series(data = item[1])
                # Prelare Current HQ Item Transport Register lines - Export
                if Current_Item["Item_Free_Of_Charge"] != "Y" and Current_Item["Item_Connected_to_BOM"] != "Y":
                    Current_HQ_ITR_Register_No = Set_HQITR_Register_No
                    Current_HQ_ITR_Register_Sub_No = ""
                    Current_HQ_ITR_Document_No = Current_PH_Document_No

                    if Current_Item["Item_Connected_to_BOM"] == "Y":
                        BOM_Item = Current_Item["BOM_Item_Relation"]
                        mask_1 = Exported_Lines_df["Item_No"] == BOM_Item
                        mask_2 = Exported_Lines_df["Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = Exported_Lines_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Document_Line = str(Export_record.iloc[0]["Line_Number"])
                    elif Current_Item["Item_Free_Of_Charge"] == "Y":
                        Main_Item = Current_Item["Item_Free_Of_Charge_Relation"]
                        mask_1 = Exported_Lines_df["Item_No"] == Main_Item
                        mask_2 = Exported_Lines_df["Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = Exported_Lines_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Document_Line = str(Export_record.iloc[0]["Line_Number"])
                    else:
                        mask_1 = Exported_Lines_df["Item_No"] == Current_Item["Item_No"] 
                        mask_2 = Exported_Lines_df["Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = Exported_Lines_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Document_Line = str(Export_record.iloc[0]["Line_Number"])

                    Current_HQ_ITR_Document_Type = Current_PH_Document_Type
                    Current_HQ_ITR_Vendor_Document_Type = "Export"
                    Current_HQ_ITR_Vendor_Document_No = ""
                    Current_HQ_ITR_Vendor_Shipment_No = ""
                    Current_HQ_ITR_Line_Type = Current_Item["Line_Type"] 
                    Current_HQ_ITR_Item_No = Current_Item["Item_No"] 
                    Current_HQ_ITR_Vendor_Item_No = ""
                    Current_HQ_ITR_Ordered_Item_No = Current_Item["Item_No"] 
                    Current_HQ_ITR_Quantity = Current_Item["Item_Line_Quantity"] 
                    Current_HQ_ITR_Ordered_Quantity = Current_Item["Item_Line_Quantity"] 
                    Current_HQ_ITR_Serial_No = ""
                    Current_HQ_ITR_Unit_of_Measure = Current_Item["Item_Unit_of_Measure"] 
                    Current_HQ_ITR_Currency_Code = HQ_General_Setup_df.iloc[0]["HQ_Currency_Code"]
                    Current_HQ_ITR_Unit_Price = "0"
                    Current_HQ_ITR_Line_Amount = "0"
                    Current_HQ_ITR_Order_Date = HQ_General_Setup_df.iloc[0]["HQ_Order_Date"]
                    Current_HQ_ITR_Receipt_Date = ""
                    Current_HQ_ITR_Posting_Date = ""
                    Current_HQ_ITR_Exported_Line_No = Current_HQ_ITR_Document_Line
                    Current_HQ_ITR_Vendor_Line_No = HQ_Export_Setup_df.iloc[0]["HQ_Export_Vendor_Line_No"]
                    Current_HQ_ITR_Vendor_No = PO_Document_Header_Setup_df.iloc[0]["Buy_from_Vendor_No"]
                    Current_HQ_ITR_Country_Region_of_Origin_Code = ""
                    Current_HQ_ITR_Plant_No = ""
                    Current_HQ_ITR_Line_Flag = ""
                    Current_HQ_ITR_Communication_Process_Status = ""
                    Current_HQ_ITR_ATP_Check_Cumulative_Quantity = "0"
                    Current_HQ_ITR_Vendor_Document_Created_Date = ""
                    Current_HQ_ITR_Delivery_Start_Date = HQ_General_Setup_df.iloc[0]["HQ_Order_Date"]
                    Current_HQ_ITR_Delivery_End_Date = HQ_General_Setup_df.iloc[0]["HQ_Order_Date"]
                    Current_HQ_ITR_Tariff_Number = ""
                    Current_HQ_ITR_Vendor_Delivery_No = ""
                    Current_HQ_ITR_Quantity_to_Deliver = "0"
                    Current_HQ_ITR_Vendor_Invoice_No = ""
                    Current_HQ_ITR_Quantity_to_Invoice = "0"
                    Current_HQ_ITR_Picking_Date = ""
                    Current_HQ_ITR_Trans_Planning_Date = ""
                    Current_HQ_ITR_Loading_Date = ""
                    Current_HQ_ITR_Planned_GI_Date = ""
                    Current_HQ_ITR_Delivery_Date = ""
                    Current_HQ_ITR_Document_Order_Number = HQ_Document_Order_Number
                    Current_HQ_ITR_To_Post_Auto = "False"
                else: 
                    continue 
                # Prepare HQ Item Transport Register - Export lists
                Export_HQ_ITR_Register_No_list.append(str(Current_HQ_ITR_Register_No))
                Export_HQ_ITR_Register_Sub_No_list.append(str(Current_HQ_ITR_Register_Sub_No))
                Export_HQ_ITR_Document_No_list.append(str(Current_HQ_ITR_Document_No))
                Export_HQ_ITR_Document_Line_list.append(str(Current_HQ_ITR_Document_Line))
                Export_HQ_ITR_Document_Type_list.append(str(Current_HQ_ITR_Document_Type))
                Export_HQ_ITR_Vendor_Document_Type_list.append(str(Current_HQ_ITR_Vendor_Document_Type))
                Export_HQ_ITR_Vendor_Document_No_list.append(str(Current_HQ_ITR_Vendor_Document_No))
                Export_HQ_ITR_Vendor_Shipment_No_list.append(str(Current_HQ_ITR_Vendor_Shipment_No))
                Export_HQ_ITR_Line_Type_list.append(str(Current_HQ_ITR_Line_Type))
                Export_HQ_ITR_Item_No_list.append(str(Current_HQ_ITR_Item_No))
                Export_HQ_ITR_Vendor_Item_No_list.append(str(Current_HQ_ITR_Vendor_Item_No))
                Export_HQ_ITR_Ordered_Item_No_list.append(str(Current_HQ_ITR_Ordered_Item_No))
                Export_HQ_ITR_Quantity_list.append(str(Current_HQ_ITR_Quantity))
                Export_HQ_ITR_Ordered_Quantity_list.append(str(Current_HQ_ITR_Ordered_Quantity))
                Export_HQ_ITR_Serial_No_list.append(str(Current_HQ_ITR_Serial_No))
                Export_HQ_ITR_Unit_of_Measure_list.append(str(Current_HQ_ITR_Unit_of_Measure))
                Export_HQ_ITR_Currency_Code_list.append(str(Current_HQ_ITR_Currency_Code))
                Export_HQ_ITR_Unit_Price_list.append(str(Current_HQ_ITR_Unit_Price))
                Export_HQ_ITR_Line_Amount_list.append(str(Current_HQ_ITR_Line_Amount))
                Export_HQ_ITR_Order_Date_list.append(str(Current_HQ_ITR_Order_Date))
                Export_HQ_ITR_Receipt_Date_list.append(str(Current_HQ_ITR_Receipt_Date))
                Export_HQ_ITR_Posting_Date_list.append(str(Current_HQ_ITR_Posting_Date))
                Export_HQ_ITR_Exported_Line_No_list.append(str(Current_HQ_ITR_Exported_Line_No))
                Export_HQ_ITR_Vendor_Line_No_list.append(str(Current_HQ_ITR_Vendor_Line_No))
                Export_HQ_ITR_Vendor_No_list.append(str(Current_HQ_ITR_Vendor_No))
                Export_HQ_ITR_Country_Region_of_Origin_Code_list.append(str(Current_HQ_ITR_Country_Region_of_Origin_Code))
                Export_HQ_ITR_Plant_No_list.append(str(Current_HQ_ITR_Plant_No))
                Export_HQ_ITR_Line_Flag_list.append(str(Current_HQ_ITR_Line_Flag))
                Export_HQ_ITR_Communication_Process_Status_list.append(str(Current_HQ_ITR_Communication_Process_Status))
                Export_HQ_ITR_ATP_Check_Cumulative_Quantity_list.append(str(Current_HQ_ITR_ATP_Check_Cumulative_Quantity))
                Export_HQ_ITR_Vendor_Document_Created_Date_list.append(str(Current_HQ_ITR_Vendor_Document_Created_Date))
                Export_HQ_ITR_Delivery_Start_Date_list.append(str(Current_HQ_ITR_Delivery_Start_Date))
                Export_HQ_ITR_Delivery_End_Date_list.append(str(Current_HQ_ITR_Delivery_End_Date))
                Export_HQ_ITR_Tariff_Number_list.append(str(Current_HQ_ITR_Tariff_Number))
                Export_HQ_ITR_Vendor_Delivery_No_list.append(str(Current_HQ_ITR_Vendor_Delivery_No))
                Export_HQ_ITR_Quantity_to_Deliver_list.append(str(Current_HQ_ITR_Quantity_to_Deliver))
                Export_HQ_ITR_Vendor_Invoice_No_list.append(str(Current_HQ_ITR_Vendor_Invoice_No))
                Export_HQ_ITR_Quantity_to_Invoice_list.append(str(Current_HQ_ITR_Quantity_to_Invoice))
                Export_HQ_ITR_Picking_Date_list.append(str(Current_HQ_ITR_Picking_Date))
                Export_HQ_ITR_Trans_Planning_Date_list.append(str(Current_HQ_ITR_Trans_Planning_Date))
                Export_HQ_ITR_Loading_Date_list.append(str(Current_HQ_ITR_Loading_Date))
                Export_HQ_ITR_Planned_GI_Date_list.append(str(Current_HQ_ITR_Planned_GI_Date))
                Export_HQ_ITR_Delivery_Date_list.append(str(Current_HQ_ITR_Delivery_Date))
                Export_HQ_ITR_Document_Order_Number_list.append(str(Current_HQ_ITR_Document_Order_Number))
                Export_HQ_ITR_To_Post_Auto_list.append(str(Current_HQ_ITR_To_Post_Auto))

                # Final Export line setup
                Set_HQITR_Register_No += int(HQ_General_Setup_df.iloc[0]["HQ_HQITR_Register_No_Increment"])
                HQ_Document_Order_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
        else:
            pass
    elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
        # NUS2 part --> currently not relevatn
        pass
    else:
        pass

    # HQ Item Transport Register - Confirmaiton
    if General_Setup_df.iloc[0]["Navision"] == "NUS3":
        if Questions_df.iloc[0]["HQ_Conf_Quest"] == "Y":
            HQ_Document_Order_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
            HQ_Vendor_Line_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
            for item in Items_df.iterrows():
                # Prepare Current HQ Item Transport Register lines - Confirmation
                Current_Item = pandas.Series(data = item[1])
                Current_HQ_ITR_Register_No = Set_HQITR_Register_No
                Current_HQ_ITR_Register_Sub_No = ""
                Current_HQ_ITR_Document_No = Current_PH_Document_No
                Current_HQ_ITR_Document_Line = ""
                Current_HQ_ITR_Document_Type = Current_PH_Document_Type
                Current_HQ_ITR_Vendor_Document_Type = "Confirmation"
                Current_HQ_ITR_Vendor_Document_No = str(HQ_Confirmation_Setup_df.iloc[0]["HQ_Confirmation_prefix"])+str(Current_HQ_Conf_Number_Suffix)
                Current_HQ_ITR_Vendor_Shipment_No = ""
                if Current_Item["Main_BOM_Item"] == "Y":
                    Current_HQ_ITR_Line_Type = "TEXT"
                else:
                    Current_HQ_ITR_Line_Type = Current_Item["Line_Type"] 

                if Current_Item["Main_BOM_Item"] == "Y":
                    Current_HQ_ITR_Item_No = Current_Item["Item_No"]+str("-BOM")
                else:
                    Current_HQ_ITR_Item_No = Current_Item["Item_No"]

                if Current_Item["Main_BOM_Item"] == "Y":
                    Current_HQ_ITR_Vendor_Item_No = Current_Item["Item_No"]+str("-BOM")
                else:
                    if Current_Item["HQ_Confirmation_Line_Flag"] == "Substituted":
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["HQ_SUB_New_Item"]
                    else:
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["Item_No"]

                if Current_Item["Item_Free_Of_Charge"] == "Y":
                    Current_HQ_ITR_Ordered_Item_No = ""    
                else:
                    Current_HQ_ITR_Ordered_Item_No = Current_Item["Item_No"]

                if Current_Item["Main_BOM_Item"] == "Y":
                    Current_HQ_ITR_Quantity = "0"
                else:
                    Current_HQ_ITR_Quantity = Current_Item["Item_Line_Quantity"]

                if Current_Item["Item_Free_Of_Charge"] == "Y":
                    Current_HQ_ITR_Ordered_Quantity =  ""    
                else:
                    Current_HQ_ITR_Ordered_Quantity =  Current_Item["Item_Line_Quantity"]

                Current_HQ_ITR_ATP_Check_Cumulative_Quantity = Current_HQ_ITR_Quantity
                Current_HQ_ITR_Serial_No = ""
                Current_HQ_ITR_Unit_of_Measure = Current_Item["Item_Unit_of_Measure"]
                Current_HQ_ITR_Currency_Code = HQ_General_Setup_df.iloc[0]["HQ_Currency_Code"]
                Current_HQ_ITR_Unit_Price = str(test_float(Current_Item["Item_Unit_Price"]))
                Current_HQ_ITR_Line_Amount = str(test_float(int(Current_HQ_ITR_Quantity)*float(Current_HQ_ITR_Unit_Price)))
                Current_HQ_ITR_Order_Date = HQ_General_Setup_df.iloc[0]["HQ_Order_Date"]
                Current_HQ_ITR_Receipt_Date = ""
                Current_HQ_ITR_Posting_Date = ""
                if Current_Item["Item_Connected_to_BOM"] == "Y":
                    BOM_Item = Current_Item["BOM_Item_Relation"]
                    mask_1 = Exported_Lines_df["Item_No"] == BOM_Item
                    mask_2 = Exported_Lines_df["Purchase_Order"] == Current_HQ_ITR_Document_No
                    Export_record = Exported_Lines_df[mask_1 & mask_2 ]
                    Current_HQ_ITR_Exported_Line_No = str(Export_record.iloc[0]["Line_Number"])
                elif Current_Item["Item_Free_Of_Charge"] == "Y":
                    Main_Item = Current_Item["Item_Free_Of_Charge_Relation"]
                    mask_1 = Exported_Lines_df["Item_No"] == Main_Item
                    mask_2 = Exported_Lines_df["Purchase_Order"] == Current_HQ_ITR_Document_No
                    Export_record = Exported_Lines_df[mask_1 & mask_2 ]
                    Current_HQ_ITR_Exported_Line_No = str(Export_record.iloc[0]["Line_Number"])
                else:
                    mask_1 = Exported_Lines_df["Item_No"] == Current_Item["Item_No"] 
                    mask_2 = Exported_Lines_df["Purchase_Order"] == Current_HQ_ITR_Document_No
                    Export_record = Exported_Lines_df[mask_1 & mask_2 ]
                    Current_HQ_ITR_Exported_Line_No = str(Export_record.iloc[0]["Line_Number"])
                Current_HQ_ITR_Vendor_Line_No = HQ_Vendor_Line_Number
                Current_HQ_ITR_Vendor_No = PO_Document_Header_Setup_df.iloc[0]["Buy_from_Vendor_No"]
                Current_HQ_ITR_Country_Region_of_Origin_Code = ""
                Current_HQ_ITR_Plant_No = ""
                Current_HQ_ITR_Line_Flag = Current_Item["HQ_Confirmation_Line_Flag"]
                Current_HQ_ITR_Communication_Process_Status = ""
                Current_HQ_ITR_Vendor_Document_Created_Date = HQ_Confirmation_Setup_df.iloc[0]["HQ_Conf_Vendor_Document_Created_Date"]
                Current_HQ_ITR_Delivery_Start_Date = HQ_General_Setup_df.iloc[0]["HQ_Delivery_Start_and_End_Date"]
                Current_HQ_ITR_Delivery_End_Date = HQ_General_Setup_df.iloc[0]["HQ_Delivery_Start_and_End_Date"]
                Current_HQ_ITR_Tariff_Number = ""
                Current_HQ_ITR_Vendor_Delivery_No = ""
                Current_HQ_ITR_Quantity_to_Deliver = "0"
                Current_HQ_ITR_Vendor_Invoice_No = ""
                Current_HQ_ITR_Quantity_to_Invoice = "0"
                Current_HQ_ITR_Picking_Date = ""
                Current_HQ_ITR_Trans_Planning_Date = ""
                Current_HQ_ITR_Loading_Date = ""
                Current_HQ_ITR_Planned_GI_Date = ""
                Current_HQ_ITR_Delivery_Date = ""
                Current_HQ_ITR_Document_Order_Number = HQ_Document_Order_Number
                Current_HQ_ITR_To_Post_Auto = "False"

                # ATP Check Register
                if HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_Multiple_Lines_Check"] == "Y":
                    New_Schedule_Date = HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_First_Scheduled_Date"]
                    for row in range(int(Current_HQ_ITR_Quantity)):
                        # Prelare Current HQ ATP Check Register lines
                        Current_HQ_ATP_HQ_ATP_Check_Register_ID = Set_HQATPR_Register_No
                        Current_HQ_ATP_HQ_ATP_Check_Sub_Register_ID = ""
                        Current_HQ_ATP_HQ_Item_Transport_Register_ID = Current_HQ_ITR_Register_No
                        Current_HQ_ATP_Scheduled_Quantity = "1"
                        Current_HQ_ATP_Stock = random.choice(str(HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_Stock"]).split(","))
                        if Current_HQ_ATP_Stock == "BACK":
                            Current_HQ_ATP_Scheduled_Date = str(HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_Zero_Date"])
                        else:
                            Current_HQ_ATP_Scheduled_Date = str(New_Schedule_Date)
                        Current_HQ_ATP_Document_No = Current_PH_Document_No
                        # Prepare HQ ATP Register - Export lists
                        Export_HQ_ATP_HQ_ATP_Check_Register_ID_list.append(str(Current_HQ_ATP_HQ_ATP_Check_Register_ID))
                        Export_HQ_ATP_HQ_ATP_Check_Sub_Register_ID_list.append(str(Current_HQ_ATP_HQ_ATP_Check_Sub_Register_ID))
                        Export_HQ_ATP_HQ_Item_Transport_Register_ID_list.append(str(Current_HQ_ATP_HQ_Item_Transport_Register_ID))
                        Export_HQ_ATP_Scheduled_Quantity_list.append(str(Current_HQ_ATP_Scheduled_Quantity))
                        Export_HQ_ATP_Scheduled_Date_list.append(str(Current_HQ_ATP_Scheduled_Date))
                        Export_HQ_ATP_Stock_list.append(str(Current_HQ_ATP_Stock))
                        Export_HQ_ATP_Document_No_list.append(str(Current_HQ_ATP_Document_No))                   
                        # Final Set for ATP Register
                        Set_HQATPR_Register_No += int(HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_Register_No_Increment"])
                        
                        Day1 = datetime.strptime(New_Schedule_Date, "%d.%m.%Y")
                        Day2 = datetime.strftime((Day1 + timedelta(days=int(HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_Sched_Date_Increments_day"]))), General_Setup_df.iloc[0]["HQ_Date_format"])
                        New_Schedule_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])
                else:
                    # Prelare Current HQ ATP Check Register lines
                    Current_HQ_ATP_HQ_ATP_Check_Register_ID = Set_HQATPR_Register_No
                    Current_HQ_ATP_HQ_ATP_Check_Sub_Register_ID = ""
                    Current_HQ_ATP_HQ_Item_Transport_Register_ID = Current_HQ_ITR_Register_No
                    Current_HQ_ATP_Scheduled_Quantity = Current_HQ_ITR_Quantity
                    Current_HQ_ATP_Stock = random.choice(str(HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_Stock"]).split(","))
                    if Current_HQ_ATP_Stock == "BACK":
                        Current_HQ_ATP_Scheduled_Date = HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_Zero_Date"]
                    else:
                        Current_HQ_ATP_Scheduled_Date = HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_First_Scheduled_Date"]
                    Current_HQ_ATP_Document_No = Current_PH_Document_No
                    # Prepare HQ ATP Register - Export lists
                    Export_HQ_ATP_HQ_ATP_Check_Register_ID_list.append(str(Current_HQ_ATP_HQ_ATP_Check_Register_ID))
                    Export_HQ_ATP_HQ_ATP_Check_Sub_Register_ID_list.append(str(Current_HQ_ATP_HQ_ATP_Check_Sub_Register_ID))
                    Export_HQ_ATP_HQ_Item_Transport_Register_ID_list.append(str(Current_HQ_ATP_HQ_Item_Transport_Register_ID))
                    Export_HQ_ATP_Scheduled_Quantity_list.append(str(Current_HQ_ATP_Scheduled_Quantity))
                    Export_HQ_ATP_Scheduled_Date_list.append(str(Current_HQ_ATP_Scheduled_Date))
                    Export_HQ_ATP_Stock_list.append(str(Current_HQ_ATP_Stock))
                    Export_HQ_ATP_Document_No_list.append(str(Current_HQ_ATP_Document_No))
                    # Final Set for ATP Register
                    Set_HQATPR_Register_No += int(HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_Register_No_Increment"])

                # Substitution Register
                if Current_Item["HQ_Confirmation_Line_Flag"] == "Substituted":
                    # Prelare Current HQ Substitution Register lines
                    Current_HQ_SUB_HQ_Substitution_Register_ID = Set_HQSUBR_Register_No
                    Current_HQ_SUB_HQ_Substitution_Sub_Reg_ID = "0"
                    Current_HQ_SUB_Substituted_Item_Old_Item = Current_Item["Item_No"]
                    Current_HQ_SUB_Substitution_New_Item = Current_Item["HQ_SUB_New_Item"]
                    Current_HQ_SUB_Document_No = Current_PH_Document_No
                    # Prepare HQ Substitution Register - Export lists
                    Export_HQ_SUB_HQ_Substitution_Register_ID_list.append(str(Current_HQ_SUB_HQ_Substitution_Register_ID))
                    Export_HQ_SUB_HQ_Substitution_Sub_Reg_ID_list.append(str(Current_HQ_SUB_HQ_Substitution_Sub_Reg_ID))
                    Export_HQ_SUB_Substituted_Item_Old_Item_list.append(str(Current_HQ_SUB_Substituted_Item_Old_Item))
                    Export_HQ_SUB_Substitution_New_Item_list.append(str(Current_HQ_SUB_Substitution_New_Item))
                    Export_HQ_SUB_Document_No_list.append(str(Current_HQ_SUB_Document_No))
                    # Final Set for Substitution Register
                    Set_HQSUBR_Register_No += int(HQ_SUB_Setup_df.iloc[0]["HQ_SUB_Register_No_Increment"])
                else:
                    pass

                # Prepare HQ Item Transport Register - Export lists
                Export_HQ_ITR_Register_No_list.append(str(Current_HQ_ITR_Register_No))
                Export_HQ_ITR_Register_Sub_No_list.append(str(Current_HQ_ITR_Register_Sub_No))
                Export_HQ_ITR_Document_No_list.append(str(Current_HQ_ITR_Document_No))
                Export_HQ_ITR_Document_Line_list.append(str(Current_HQ_ITR_Document_Line))
                Export_HQ_ITR_Document_Type_list.append(str(Current_HQ_ITR_Document_Type))
                Export_HQ_ITR_Vendor_Document_Type_list.append(str(Current_HQ_ITR_Vendor_Document_Type))
                Export_HQ_ITR_Vendor_Document_No_list.append(str(Current_HQ_ITR_Vendor_Document_No))
                Export_HQ_ITR_Vendor_Shipment_No_list.append(str(Current_HQ_ITR_Vendor_Shipment_No))
                Export_HQ_ITR_Line_Type_list.append(str(Current_HQ_ITR_Line_Type))
                Export_HQ_ITR_Item_No_list.append(str(Current_HQ_ITR_Item_No))
                Export_HQ_ITR_Vendor_Item_No_list.append(str(Current_HQ_ITR_Vendor_Item_No))
                Export_HQ_ITR_Ordered_Item_No_list.append(str(Current_HQ_ITR_Ordered_Item_No))
                Export_HQ_ITR_Quantity_list.append(str(Current_HQ_ITR_Quantity))
                Export_HQ_ITR_Ordered_Quantity_list.append(str(Current_HQ_ITR_Ordered_Quantity))
                Export_HQ_ITR_Serial_No_list.append(str(Current_HQ_ITR_Serial_No))
                Export_HQ_ITR_Unit_of_Measure_list.append(str(Current_HQ_ITR_Unit_of_Measure))
                Export_HQ_ITR_Currency_Code_list.append(str(Current_HQ_ITR_Currency_Code))
                Export_HQ_ITR_Unit_Price_list.append(str(Current_HQ_ITR_Unit_Price))
                Export_HQ_ITR_Line_Amount_list.append(str(Current_HQ_ITR_Line_Amount))
                Export_HQ_ITR_Order_Date_list.append(str(Current_HQ_ITR_Order_Date))
                Export_HQ_ITR_Receipt_Date_list.append(str(Current_HQ_ITR_Receipt_Date))
                Export_HQ_ITR_Posting_Date_list.append(str(Current_HQ_ITR_Posting_Date))
                Export_HQ_ITR_Exported_Line_No_list.append(str(Current_HQ_ITR_Exported_Line_No))
                Export_HQ_ITR_Vendor_Line_No_list.append(str(Current_HQ_ITR_Vendor_Line_No))
                Export_HQ_ITR_Vendor_No_list.append(str(Current_HQ_ITR_Vendor_No))
                Export_HQ_ITR_Country_Region_of_Origin_Code_list.append(str(Current_HQ_ITR_Country_Region_of_Origin_Code))
                Export_HQ_ITR_Plant_No_list.append(str(Current_HQ_ITR_Plant_No))
                Export_HQ_ITR_Line_Flag_list.append(str(Current_HQ_ITR_Line_Flag))
                Export_HQ_ITR_Communication_Process_Status_list.append(str(Current_HQ_ITR_Communication_Process_Status))
                Export_HQ_ITR_ATP_Check_Cumulative_Quantity_list.append(str(Current_HQ_ITR_ATP_Check_Cumulative_Quantity))
                Export_HQ_ITR_Vendor_Document_Created_Date_list.append(str(Current_HQ_ITR_Vendor_Document_Created_Date))
                Export_HQ_ITR_Delivery_Start_Date_list.append(str(Current_HQ_ITR_Delivery_Start_Date))
                Export_HQ_ITR_Delivery_End_Date_list.append(str(Current_HQ_ITR_Delivery_End_Date))
                Export_HQ_ITR_Tariff_Number_list.append(str(Current_HQ_ITR_Tariff_Number))
                Export_HQ_ITR_Vendor_Delivery_No_list.append(str(Current_HQ_ITR_Vendor_Delivery_No))
                Export_HQ_ITR_Quantity_to_Deliver_list.append(str(Current_HQ_ITR_Quantity_to_Deliver))
                Export_HQ_ITR_Vendor_Invoice_No_list.append(str(Current_HQ_ITR_Vendor_Invoice_No))
                Export_HQ_ITR_Quantity_to_Invoice_list.append(str(Current_HQ_ITR_Quantity_to_Invoice))
                Export_HQ_ITR_Picking_Date_list.append(str(Current_HQ_ITR_Picking_Date))
                Export_HQ_ITR_Trans_Planning_Date_list.append(str(Current_HQ_ITR_Trans_Planning_Date))
                Export_HQ_ITR_Loading_Date_list.append(str(Current_HQ_ITR_Loading_Date))
                Export_HQ_ITR_Planned_GI_Date_list.append(str(Current_HQ_ITR_Planned_GI_Date))
                Export_HQ_ITR_Delivery_Date_list.append(str(Current_HQ_ITR_Delivery_Date))
                Export_HQ_ITR_Document_Order_Number_list.append(str(Current_HQ_ITR_Document_Order_Number))
                Export_HQ_ITR_To_Post_Auto_list.append(str(Current_HQ_ITR_To_Post_Auto))

                # Final Export line setup
                Set_HQITR_Register_No += int(HQ_General_Setup_df.iloc[0]["HQ_HQITR_Register_No_Increment"])
                HQ_Document_Order_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
        else:
            pass
    elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
        # NUS2 part
        if Questions_df.iloc[0]["HQ_Conf_Quest"] == "Y":
            HQ_Vendor_Line_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
            for item in Items_df.iterrows():
                # Prepare Current HQ Item Transport Register lines - Confirmation
                Current_Item = pandas.Series(data = item[1])
                Current_HQ_ITR_Register_No = Set_HQITR_Register_No
                Current_HQ_ITR_Register_Sub_No = ""
                Current_HQ_ITR_Document_No = Current_PH_Document_No
                Current_HQ_ITR_Document_Line = ""
                Current_HQ_ITR_Document_Type = Current_PH_Document_Type
                Current_HQ_ITR_Vendor_Document_Type = "Confirmation"
                Current_HQ_ITR_Vendor_Document_No = str(HQ_Confirmation_Setup_df.iloc[0]["HQ_Confirmation_prefix"])+str(Current_HQ_Conf_Number_Suffix)
                Current_HQ_ITR_Vendor_Shipment_No = ""
                Current_HQ_ITR_Line_Type = Current_Item["Line_Type"] 

                if Current_Item["Main_BOM_Item"] == "Y":
                    Current_HQ_ITR_Item_No = Current_Item["Item_No"]+str("-BOM")
                else:
                    Current_HQ_ITR_Item_No = Current_Item["Item_No"]

                if Current_Item["Main_BOM_Item"] == "Y":
                    Current_HQ_ITR_Vendor_Item_No = Current_Item["Item_No"]+str("-BOM")
                else:
                    if Current_Item["HQ_Confirmation_Line_Flag"] == "Substituted":
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["HQ_SUB_New_Item"]
                    else:
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["Item_No"]

                if Current_Item["Item_Free_Of_Charge"] == "Y":
                    Current_HQ_ITR_Ordered_Item_No = ""    
                else:
                    Current_HQ_ITR_Ordered_Item_No = Current_Item["Item_No"]

                if Current_Item["Main_BOM_Item"] == "Y":
                    Current_HQ_ITR_Quantity = "0"
                else:
                    Current_HQ_ITR_Quantity = Current_Item["Item_Line_Quantity"]

                if Current_Item["Item_Free_Of_Charge"] == "Y":
                    Current_HQ_ITR_Ordered_Quantity =  ""    
                else:
                    Current_HQ_ITR_Ordered_Quantity =  Current_Item["Item_Line_Quantity"]

                Current_HQ_ITR_Serial_No = ""
                Current_HQ_ITR_Unit_of_Measure = Current_Item["Item_Unit_of_Measure"]
                Current_HQ_ITR_Currency_Code = HQ_General_Setup_df.iloc[0]["HQ_Currency_Code"]
                Current_HQ_ITR_Unit_Price = str(test_float(Current_Item["Item_Unit_Price"]))
                Current_HQ_ITR_Line_Amount = str(test_float(int(Current_HQ_ITR_Quantity)*float(Current_HQ_ITR_Unit_Price)))
                Current_HQ_ITR_Order_Date = HQ_General_Setup_df.iloc[0]["HQ_Order_Date"]
                Current_HQ_ITR_Posting_Date = ""
                if Current_Item["Item_Connected_to_BOM"] == "Y":
                    BOM_Item = Current_Item["BOM_Item_Relation"]
                    mask_1 = Exported_Lines_df["Item_No"] == BOM_Item
                    mask_2 = Exported_Lines_df["Purchase_Order"] == Current_HQ_ITR_Document_No
                    Export_record = Exported_Lines_df[mask_1 & mask_2 ]
                    Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["Line_Number"])//100)
                elif Current_Item["Item_Free_Of_Charge"] == "Y":
                    Main_Item = Current_Item["Item_Free_Of_Charge_Relation"]
                    mask_1 = Exported_Lines_df["Item_No"] == Main_Item
                    mask_2 = Exported_Lines_df["Purchase_Order"] == Current_HQ_ITR_Document_No
                    Export_record = Exported_Lines_df[mask_1 & mask_2 ]
                    Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["Line_Number"])//100)
                else:
                    mask_1 = Exported_Lines_df["Item_No"] == Current_Item["Item_No"] 
                    mask_2 = Exported_Lines_df["Purchase_Order"] == Current_HQ_ITR_Document_No
                    Export_record = Exported_Lines_df[mask_1 & mask_2 ]
                    Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["Line_Number"])//100)
                Current_HQ_ITR_Vendor_Line_No = HQ_Vendor_Line_Number
                Current_HQ_ITR_Vendor_No = PO_Document_Header_Setup_df.iloc[0]["Buy_from_Vendor_No"]
                Current_HQ_ITR_Country_Region_of_Origin_Code = ""
                Current_HQ_ITR_Plant_No = ""
                Current_HQ_ITR_Line_Flag = Current_Item["HQ_Confirmation_Line_Flag"]
                Current_HQ_ITR_Communication_Process_Status = ""
                Current_HQ_ITR_Vendor_Document_Created_Date = HQ_Confirmation_Setup_df.iloc[0]["HQ_Conf_Vendor_Document_Created_Date"]
                Current_HQ_ATP_Stock = random.choice(str(HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_Stock"]).split(","))
                if Current_HQ_ATP_Stock == "BACK":
                    Current_HQ_ITR_Receipt_Date = HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_Zero_Date"]
                else:
                    Current_HQ_ITR_Receipt_Date = HQ_ATPR_Setup_df.iloc[0]["HQ_ATP_First_Scheduled_Date"]

                # Prepare HQ Item Transport Register - Export lists
                Export_HQ_ITR_Register_No_list.append(str(Current_HQ_ITR_Register_No))
                Export_HQ_ITR_Register_Sub_No_list.append(str(Current_HQ_ITR_Register_Sub_No))
                Export_HQ_ITR_Document_No_list.append(str(Current_HQ_ITR_Document_No))
                Export_HQ_ITR_Document_Line_list.append(str(Current_HQ_ITR_Document_Line))
                Export_HQ_ITR_Document_Type_list.append(str(Current_HQ_ITR_Document_Type))
                Export_HQ_ITR_Vendor_Document_Type_list.append(str(Current_HQ_ITR_Vendor_Document_Type))
                Export_HQ_ITR_Vendor_Document_No_list.append(str(Current_HQ_ITR_Vendor_Document_No))
                Export_HQ_ITR_Vendor_Shipment_No_list.append(str(Current_HQ_ITR_Vendor_Shipment_No))
                Export_HQ_ITR_Line_Type_list.append(str(Current_HQ_ITR_Line_Type))
                Export_HQ_ITR_Item_No_list.append(str(Current_HQ_ITR_Item_No))
                Export_HQ_ITR_Vendor_Item_No_list.append(str(Current_HQ_ITR_Vendor_Item_No))
                Export_HQ_ITR_Ordered_Item_No_list.append(str(Current_HQ_ITR_Ordered_Item_No))
                Export_HQ_ITR_Quantity_list.append(str(Current_HQ_ITR_Quantity))
                Export_HQ_ITR_Ordered_Quantity_list.append(str(Current_HQ_ITR_Ordered_Quantity))
                Export_HQ_ITR_Serial_No_list.append(str(Current_HQ_ITR_Serial_No))
                Export_HQ_ITR_Unit_of_Measure_list.append(str(Current_HQ_ITR_Unit_of_Measure))
                Export_HQ_ITR_Currency_Code_list.append(str(Current_HQ_ITR_Currency_Code))
                Export_HQ_ITR_Unit_Price_list.append(str(Current_HQ_ITR_Unit_Price))
                Export_HQ_ITR_Line_Amount_list.append(str(Current_HQ_ITR_Line_Amount))
                Export_HQ_ITR_Order_Date_list.append(str(Current_HQ_ITR_Order_Date))
                Export_HQ_ITR_Receipt_Date_list.append(str(Current_HQ_ITR_Receipt_Date))
                Export_HQ_ITR_Posting_Date_list.append(str(Current_HQ_ITR_Posting_Date))
                Export_HQ_ITR_Exported_Line_No_list.append(str(Current_HQ_ITR_Exported_Line_No))
                Export_HQ_ITR_Vendor_Line_No_list.append(str(Current_HQ_ITR_Vendor_Line_No))
                Export_HQ_ITR_Vendor_No_list.append(str(Current_HQ_ITR_Vendor_No))
                Export_HQ_ITR_Country_Region_of_Origin_Code_list.append(str(Current_HQ_ITR_Country_Region_of_Origin_Code))
                Export_HQ_ITR_Plant_No_list.append(str(Current_HQ_ITR_Plant_No))
                Export_HQ_ITR_Line_Flag_list.append(str(Current_HQ_ITR_Line_Flag))
                Export_HQ_ITR_Communication_Process_Status_list.append(str(Current_HQ_ITR_Communication_Process_Status))
                Export_HQ_ITR_Vendor_Document_Created_Date_list.append(str(Current_HQ_ITR_Vendor_Document_Created_Date))
                Export_HQ_ATP_Stock_list.append(str(Current_HQ_ATP_Stock))

                # Final Export line setup
                Set_HQITR_Register_No += int(HQ_General_Setup_df.iloc[0]["HQ_HQITR_Register_No_Increment"])
                HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
        pass
    else:
        pass

    # HQ Item Transport Register - PreAdvice
    if General_Setup_df.iloc[0]["Navision"] == "NUS3":
        if Questions_df.iloc[0]["HQ_PreAdv_Quest"] == "Y":
            # Multiple Delivery per Order
            New_PreA_Document_Date = HQ_PreAdvice_Setup_df.iloc[0]["HQ_PreAdvice_Vendor_Document_Created_Date"]
            for counter in range(int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"])):
                # Define proper lines in DEL_Item_Assign_df related only to the one delivery
                mask = DEL_Item_Assign_df["DEL_Counter"] == counter
                DEL_Detail_Item_Assign_df = DEL_Item_Assign_df[mask]
                
                HQ_Document_Order_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
                HQ_Vendor_Line_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
                for item in DEL_Detail_Item_Assign_df.iterrows():
                    Current_Item = pandas.Series(data = item[1])
                    if Current_Item["Main_BOM_Item"] == "Y" or Current_Item["HQ_Confirmation_Line_Flag"] == "Cancelled" or Current_Item["HQ_Confirmation_Line_Flag"] == "Finished":
                        HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                        continue
                    # Prepare Current HQ Item Transport Register lines - PreAdvice
                    Current_HQ_ITR_Register_No = Set_HQITR_Register_No
                    Current_HQ_ITR_Register_Sub_No = ""
                    Current_HQ_ITR_Document_No = Current_PH_Document_No
                    Current_HQ_ITR_Document_Line = ""
                    Current_HQ_ITR_Document_Type = Current_PH_Document_Type
                    Current_HQ_ITR_Vendor_Document_Type = "PreAdvice"
                    Current_HQ_ITR_Vendor_Document_No = Current_Item["DEL_PreAdvice_No"] 
                    Current_HQ_ITR_Vendor_Shipment_No = ""
                    Current_HQ_ITR_Line_Type = Current_Item["Line_Type"] 
                    if Current_Item["HQ_Confirmation_Line_Flag"] == "Substituted":
                        Current_HQ_ITR_Item_No = Current_Item["HQ_SUB_New_Item"]
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["HQ_SUB_New_Item"]
                    else:
                        Current_HQ_ITR_Item_No = Current_Item["Item_No"]
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["Item_No"]
                    Current_HQ_ITR_Ordered_Item_No = ""
                    Current_HQ_ITR_Quantity = Current_Item["DEL_Line_Qty"] 
                    if Current_Item["Item_Free_Of_Charge"] == "Y":
                        Current_HQ_ITR_Ordered_Quantity =  ""    
                    else:
                        Current_HQ_ITR_Ordered_Quantity =  Current_Item["Item_Line_Quantity"]
                    Current_HQ_ITR_Serial_No = ""
                    Current_HQ_ITR_Unit_of_Measure = Current_Item["Item_Unit_of_Measure"]
                    Current_HQ_ITR_Currency_Code = ""
                    Current_HQ_ITR_Unit_Price = ""
                    Current_HQ_ITR_Line_Amount = ""
                    Current_HQ_ITR_Order_Date = HQ_General_Setup_df.iloc[0]["HQ_Order_Date"]
                    Current_HQ_ITR_Posting_Date = ""
                    if Current_Item["Item_Connected_to_BOM"] == "Y":
                        BOM_Item = Current_Item["BOM_Item_Relation"]
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == BOM_Item
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(Export_record.iloc[0]["DEL_Line_Number"])
                    elif Current_Item["Item_Free_Of_Charge"] == "Y":
                        Main_Item = Current_Item["Item_Free_Of_Charge_Relation"]
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Main_Item
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(Export_record.iloc[0]["DEL_Line_Number"])
                    else:
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Current_Item["Item_No"] 
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(Export_record.iloc[0]["DEL_Line_Number"])
                    Current_HQ_ITR_Vendor_Line_No = HQ_Vendor_Line_Number
                    Current_HQ_ITR_Vendor_No = PO_Document_Header_Setup_df.iloc[0]["Buy_from_Vendor_No"]
                    Current_HQ_ITR_Country_Region_of_Origin_Code = ""
                    Current_HQ_ITR_Plant_No = ""
                    Current_HQ_ITR_Line_Flag = ""
                    Current_HQ_ITR_Communication_Process_Status = ""
                    Current_HQ_ITR_ATP_Check_Cumulative_Quantity = "0"
                    Current_HQ_ITR_Vendor_Document_Created_Date = str(New_PreA_Document_Date)
                    Current_HQ_ITR_Tariff_Number = ""
                    Current_HQ_ITR_Vendor_Delivery_No = ""
                    Current_HQ_ITR_Quantity_to_Deliver = "0"
                    Current_HQ_ITR_Vendor_Invoice_No = ""
                    Current_HQ_ITR_Quantity_to_Invoice = "0"
                    # Date movements for Delivery Datest
                    Day1 = datetime.strptime(Current_HQ_ITR_Vendor_Document_Created_Date, "%d.%m.%Y")
                    Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                    Current_HQ_ITR_Picking_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])

                    Day1 = datetime.strptime(Current_HQ_ITR_Picking_Date, "%d.%m.%Y")
                    Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                    Current_HQ_ITR_Trans_Planning_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])

                    Day1 = datetime.strptime(Current_HQ_ITR_Trans_Planning_Date, "%d.%m.%Y")
                    Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                    Current_HQ_ITR_Loading_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])
                    
                    Day1 = datetime.strptime(Current_HQ_ITR_Loading_Date, "%d.%m.%Y")
                    Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                    Current_HQ_ITR_Planned_GI_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])

                    Day1 = datetime.strptime(Current_HQ_ITR_Planned_GI_Date, "%d.%m.%Y")
                    Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                    Current_HQ_ITR_Delivery_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])
                    Current_HQ_ITR_Delivery_Start_Date = str(Current_HQ_ITR_Delivery_Date)
                    Current_HQ_ITR_Delivery_End_Date = str(Current_HQ_ITR_Delivery_Date)
                    Current_HQ_ITR_Receipt_Date = str(Current_HQ_ITR_Delivery_Date)
                    Current_HQ_ITR_Document_Order_Number = HQ_Document_Order_Number
                    Current_HQ_ITR_To_Post_Auto = "False"      

                    # Prepare HQ Item Transport Register - Export lists
                    Export_HQ_ITR_Register_No_list.append(str(Current_HQ_ITR_Register_No))
                    Export_HQ_ITR_Register_Sub_No_list.append(str(Current_HQ_ITR_Register_Sub_No))
                    Export_HQ_ITR_Document_No_list.append(str(Current_HQ_ITR_Document_No))
                    Export_HQ_ITR_Document_Line_list.append(str(Current_HQ_ITR_Document_Line))
                    Export_HQ_ITR_Document_Type_list.append(str(Current_HQ_ITR_Document_Type))
                    Export_HQ_ITR_Vendor_Document_Type_list.append(str(Current_HQ_ITR_Vendor_Document_Type))
                    Export_HQ_ITR_Vendor_Document_No_list.append(str(Current_HQ_ITR_Vendor_Document_No))
                    Export_HQ_ITR_Vendor_Shipment_No_list.append(str(Current_HQ_ITR_Vendor_Shipment_No))
                    Export_HQ_ITR_Line_Type_list.append(str(Current_HQ_ITR_Line_Type))
                    Export_HQ_ITR_Item_No_list.append(str(Current_HQ_ITR_Item_No))
                    Export_HQ_ITR_Vendor_Item_No_list.append(str(Current_HQ_ITR_Vendor_Item_No))
                    Export_HQ_ITR_Ordered_Item_No_list.append(str(Current_HQ_ITR_Ordered_Item_No))
                    Export_HQ_ITR_Quantity_list.append(str(Current_HQ_ITR_Quantity))
                    Export_HQ_ITR_Ordered_Quantity_list.append(str(Current_HQ_ITR_Ordered_Quantity))
                    Export_HQ_ITR_Serial_No_list.append(str(Current_HQ_ITR_Serial_No))
                    Export_HQ_ITR_Unit_of_Measure_list.append(str(Current_HQ_ITR_Unit_of_Measure))
                    Export_HQ_ITR_Currency_Code_list.append(str(Current_HQ_ITR_Currency_Code))
                    Export_HQ_ITR_Unit_Price_list.append(str(Current_HQ_ITR_Unit_Price))
                    Export_HQ_ITR_Line_Amount_list.append(str(Current_HQ_ITR_Line_Amount))
                    Export_HQ_ITR_Order_Date_list.append(str(Current_HQ_ITR_Order_Date))
                    Export_HQ_ITR_Receipt_Date_list.append(str(Current_HQ_ITR_Receipt_Date))
                    Export_HQ_ITR_Posting_Date_list.append(str(Current_HQ_ITR_Posting_Date))
                    Export_HQ_ITR_Exported_Line_No_list.append(str(Current_HQ_ITR_Exported_Line_No))
                    Export_HQ_ITR_Vendor_Line_No_list.append(str(Current_HQ_ITR_Vendor_Line_No))
                    Export_HQ_ITR_Vendor_No_list.append(str(Current_HQ_ITR_Vendor_No))
                    Export_HQ_ITR_Country_Region_of_Origin_Code_list.append(str(Current_HQ_ITR_Country_Region_of_Origin_Code))
                    Export_HQ_ITR_Plant_No_list.append(str(Current_HQ_ITR_Plant_No))
                    Export_HQ_ITR_Line_Flag_list.append(str(Current_HQ_ITR_Line_Flag))
                    Export_HQ_ITR_Communication_Process_Status_list.append(str(Current_HQ_ITR_Communication_Process_Status))
                    Export_HQ_ITR_ATP_Check_Cumulative_Quantity_list.append(str(Current_HQ_ITR_ATP_Check_Cumulative_Quantity))
                    Export_HQ_ITR_Vendor_Document_Created_Date_list.append(str(Current_HQ_ITR_Vendor_Document_Created_Date))
                    Export_HQ_ITR_Delivery_Start_Date_list.append(str(Current_HQ_ITR_Delivery_Start_Date))
                    Export_HQ_ITR_Delivery_End_Date_list.append(str(Current_HQ_ITR_Delivery_End_Date))
                    Export_HQ_ITR_Tariff_Number_list.append(str(Current_HQ_ITR_Tariff_Number))
                    Export_HQ_ITR_Vendor_Delivery_No_list.append(str(Current_HQ_ITR_Vendor_Delivery_No))
                    Export_HQ_ITR_Quantity_to_Deliver_list.append(str(Current_HQ_ITR_Quantity_to_Deliver))
                    Export_HQ_ITR_Vendor_Invoice_No_list.append(str(Current_HQ_ITR_Vendor_Invoice_No))
                    Export_HQ_ITR_Quantity_to_Invoice_list.append(str(Current_HQ_ITR_Quantity_to_Invoice))
                    Export_HQ_ITR_Picking_Date_list.append(str(Current_HQ_ITR_Picking_Date))
                    Export_HQ_ITR_Trans_Planning_Date_list.append(str(Current_HQ_ITR_Trans_Planning_Date))
                    Export_HQ_ITR_Loading_Date_list.append(str(Current_HQ_ITR_Loading_Date))
                    Export_HQ_ITR_Planned_GI_Date_list.append(str(Current_HQ_ITR_Planned_GI_Date))
                    Export_HQ_ITR_Delivery_Date_list.append(str(Current_HQ_ITR_Delivery_Date))
                    Export_HQ_ITR_Document_Order_Number_list.append(str(Current_HQ_ITR_Document_Order_Number))
                    Export_HQ_ITR_To_Post_Auto_list.append(str(Current_HQ_ITR_To_Post_Auto))

                    # HQ PreAdvice Register (1 line in PreAdvice = 1line in HQ Item Transport Register)
                    Current_HQ_PREA_HQ_PreAdvice_Register_ID = Set_HQPREA_Regsiter_No
                    Current_HQ_PREA_HQ_PreAdvice_Sub_Register_ID = ""
                    Current_HQ_PREA_HQ_Item_Transport_Register_ID = Current_HQ_ITR_Register_No
                    Current_HQ_PREA_HQ_Item_Transport_Reg_Sub_ID = ""
                    Current_HQ_PREA_Document_No = Current_PH_Document_No
                    Current_HQ_PREA_PreAdvice_Document_No = Current_HQ_ITR_Vendor_Document_No
                    Current_HQ_PREA_PreAdvice_Date = Current_HQ_ITR_Delivery_Date

                    Export_HQ_PREA_HQ_PreAdvice_Register_ID_list.append(str(Current_HQ_PREA_HQ_PreAdvice_Register_ID))
                    Export_HQ_PREA_HQ_PreAdvice_Sub_Register_ID_list.append(str(Current_HQ_PREA_HQ_PreAdvice_Sub_Register_ID))
                    Export_HQ_PREA_HQ_Item_Transport_Register_ID_list.append(str(Current_HQ_PREA_HQ_Item_Transport_Register_ID))
                    Export_HQ_PREA_HQ_Item_Transport_Reg_Sub_ID_list.append(str(Current_HQ_PREA_HQ_Item_Transport_Reg_Sub_ID))
                    Export_HQ_PREA_Document_No_list.append(str(Current_HQ_PREA_Document_No))
                    Export_HQ_PREA_PreAdvice_Document_No_list.append(str(Current_HQ_PREA_PreAdvice_Document_No))
                    Export_HQ_PREA_PreAdvice_Date_list.append(str(Current_HQ_PREA_PreAdvice_Date))

                    # Final Export line setup
                    Set_HQITR_Register_No += int(HQ_General_Setup_df.iloc[0]["HQ_HQITR_Register_No_Increment"])
                    Set_HQPREA_Regsiter_No += int(HQ_HQPAR_Setup_df.iloc[0]["HQ_HQPAR_Register_No_Increment"])
                    HQ_Document_Order_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                    HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])

                Day1 = datetime.strptime(New_PreA_Document_Date, "%d.%m.%Y")
                Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                New_PreA_Document_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])
        else:
            pass
    elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
        if Questions_df.iloc[0]["HQ_PreAdv_Quest"] == "Y":
            # Multiple Delivery per Order
            New_PreA_Document_Date = HQ_PreAdvice_Setup_df.iloc[0]["HQ_PreAdvice_Vendor_Document_Created_Date"]
            for counter in range(int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"])):
                # Define proper lines in DEL_Item_Assign_df related only to the one delivery
                mask = DEL_Item_Assign_df["DEL_Counter"] == counter
                DEL_Detail_Item_Assign_df = DEL_Item_Assign_df[mask]

                HQ_Vendor_Line_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
                for item in DEL_Detail_Item_Assign_df.iterrows():
                    Current_Item = pandas.Series(data = item[1])
                    if Current_Item["Main_BOM_Item"] == "Y" or Current_Item["HQ_Confirmation_Line_Flag"] == "Cancelled" or Current_Item["HQ_Confirmation_Line_Flag"] == "Finished":
                        HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                        continue
                    # Prepare Current HQ Item Transport Register lines - PreAdvice
                    Current_HQ_ITR_Register_No = Set_HQITR_Register_No
                    Current_HQ_ITR_Register_Sub_No = ""
                    Current_HQ_ITR_Document_No = Current_PH_Document_No
                    Current_HQ_ITR_Document_Line = ""
                    Current_HQ_ITR_Document_Type = Current_PH_Document_Type
                    Current_HQ_ITR_Vendor_Document_Type = "Pre-Advice"
                    Current_HQ_ITR_Vendor_Document_No = Current_Item["DEL_PreAdvice_No"]
                    Current_HQ_ITR_Vendor_Shipment_No = ""
                    Current_HQ_ITR_Line_Type = Current_Item["Line_Type"] 
                    if Current_Item["HQ_Confirmation_Line_Flag"] == "Substituted":
                        Current_HQ_ITR_Item_No = Current_Item["HQ_SUB_New_Item"]
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["HQ_SUB_New_Item"]
                    else:
                        Current_HQ_ITR_Item_No = Current_Item["Item_No"]
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["Item_No"]
                    Current_HQ_ITR_Ordered_Item_No = ""
                    Current_HQ_ITR_Quantity = Current_Item["DEL_Line_Qty"]
                    if Current_Item["Item_Free_Of_Charge"] == "Y":
                        Current_HQ_ITR_Ordered_Quantity =  ""    
                    else:
                        Current_HQ_ITR_Ordered_Quantity =  Current_Item["Item_Line_Quantity"]
                    Current_HQ_ITR_Serial_No = ""
                    Current_HQ_ITR_Unit_of_Measure = Current_Item["Item_Unit_of_Measure"]
                    Current_HQ_ITR_Currency_Code = ""
                    Current_HQ_ITR_Unit_Price = ""
                    Current_HQ_ITR_Line_Amount = ""
                    Current_HQ_ITR_Order_Date = HQ_General_Setup_df.iloc[0]["HQ_Order_Date"]
                    Current_HQ_ITR_Posting_Date = ""
                    if Current_Item["Item_Connected_to_BOM"] == "Y":
                        BOM_Item = Current_Item["BOM_Item_Relation"]
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == BOM_Item
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["DEL_Line_Number"])//100)
                    elif Current_Item["Item_Free_Of_Charge"] == "Y":
                        Main_Item = Current_Item["Item_Free_Of_Charge_Relation"]
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Main_Item
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["DEL_Line_Number"])//100)
                    else:
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Current_Item["Item_No"] 
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["DEL_Line_Number"])//100)
                    Current_HQ_ITR_Vendor_Line_No = HQ_Vendor_Line_Number
                    Current_HQ_ITR_Vendor_No = PO_Document_Header_Setup_df.iloc[0]["Buy_from_Vendor_No"]
                    Current_HQ_ITR_Country_Region_of_Origin_Code = ""
                    Current_HQ_ITR_Plant_No = ""
                    Current_HQ_ITR_Line_Flag = ""
                    Current_HQ_ITR_Communication_Process_Status = ""
                    Current_HQ_ITR_Vendor_Document_Created_Date = str(New_PreA_Document_Date)
                    Current_HQ_ITR_Receipt_Date = str(New_PreA_Document_Date)
                    Current_HQ_ATP_Stock = ""

                    # Prepare HQ Item Transport Register - Export lists
                    Export_HQ_ITR_Register_No_list.append(str(Current_HQ_ITR_Register_No))
                    Export_HQ_ITR_Register_Sub_No_list.append(str(Current_HQ_ITR_Register_Sub_No))
                    Export_HQ_ITR_Document_No_list.append(str(Current_HQ_ITR_Document_No))
                    Export_HQ_ITR_Document_Line_list.append(str(Current_HQ_ITR_Document_Line))
                    Export_HQ_ITR_Document_Type_list.append(str(Current_HQ_ITR_Document_Type))
                    Export_HQ_ITR_Vendor_Document_Type_list.append(str(Current_HQ_ITR_Vendor_Document_Type))
                    Export_HQ_ITR_Vendor_Document_No_list.append(str(Current_HQ_ITR_Vendor_Document_No))
                    Export_HQ_ITR_Vendor_Shipment_No_list.append(str(Current_HQ_ITR_Vendor_Shipment_No))
                    Export_HQ_ITR_Line_Type_list.append(str(Current_HQ_ITR_Line_Type))
                    Export_HQ_ITR_Item_No_list.append(str(Current_HQ_ITR_Item_No))
                    Export_HQ_ITR_Vendor_Item_No_list.append(str(Current_HQ_ITR_Vendor_Item_No))
                    Export_HQ_ITR_Ordered_Item_No_list.append(str(Current_HQ_ITR_Ordered_Item_No))
                    Export_HQ_ITR_Quantity_list.append(str(Current_HQ_ITR_Quantity))
                    Export_HQ_ITR_Ordered_Quantity_list.append(str(Current_HQ_ITR_Ordered_Quantity))
                    Export_HQ_ITR_Serial_No_list.append(str(Current_HQ_ITR_Serial_No))
                    Export_HQ_ITR_Unit_of_Measure_list.append(str(Current_HQ_ITR_Unit_of_Measure))
                    Export_HQ_ITR_Currency_Code_list.append(str(Current_HQ_ITR_Currency_Code))
                    Export_HQ_ITR_Unit_Price_list.append(str(Current_HQ_ITR_Unit_Price))
                    Export_HQ_ITR_Line_Amount_list.append(str(Current_HQ_ITR_Line_Amount))
                    Export_HQ_ITR_Order_Date_list.append(str(Current_HQ_ITR_Order_Date))
                    Export_HQ_ITR_Receipt_Date_list.append(str(Current_HQ_ITR_Receipt_Date))
                    Export_HQ_ITR_Posting_Date_list.append(str(Current_HQ_ITR_Posting_Date))
                    Export_HQ_ITR_Exported_Line_No_list.append(str(Current_HQ_ITR_Exported_Line_No))
                    Export_HQ_ITR_Vendor_Line_No_list.append(str(Current_HQ_ITR_Vendor_Line_No))
                    Export_HQ_ITR_Vendor_No_list.append(str(Current_HQ_ITR_Vendor_No))
                    Export_HQ_ITR_Country_Region_of_Origin_Code_list.append(str(Current_HQ_ITR_Country_Region_of_Origin_Code))
                    Export_HQ_ITR_Plant_No_list.append(str(Current_HQ_ITR_Plant_No))
                    Export_HQ_ITR_Line_Flag_list.append(str(Current_HQ_ITR_Line_Flag))
                    Export_HQ_ITR_Communication_Process_Status_list.append(str(Current_HQ_ITR_Communication_Process_Status))
                    Export_HQ_ITR_Vendor_Document_Created_Date_list.append(str(Current_HQ_ITR_Vendor_Document_Created_Date))
                    Export_HQ_ATP_Stock_list.append(str(Current_HQ_ATP_Stock))
                    # Final Export line setup
                    Set_HQITR_Register_No += int(HQ_General_Setup_df.iloc[0]["HQ_HQITR_Register_No_Increment"])
                    HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])

                Day1 = datetime.strptime(New_PreA_Document_Date, "%d.%m.%Y")
                Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                New_PreA_Document_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])
        pass
    else:
        pass

    # HQ Item Transport Register - Delivery
    if General_Setup_df.iloc[0]["Navision"] == "NUS3":
        if Questions_df.iloc[0]["HQ_Del_Quest"] == "Y":
            # Multiple Delivery per Order
            New_Delivery_Document_Date = HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_Vendor_Document_Created_Date"]
            for counter in range(int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"])):
                # Define proper lines in DEL_Item_Assign_df related only to the one delivery
                mask = DEL_Item_Assign_df["DEL_Counter"] == counter
                DEL_Detail_Item_Assign_df = DEL_Item_Assign_df[mask]

                HQ_Document_Order_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
                HQ_Vendor_Line_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
                HQ_DTR_Per_Delivery_counter = 1
                for item in DEL_Detail_Item_Assign_df.iterrows():
                    Current_Item = pandas.Series(data = item[1])
                    if Current_Item["Main_BOM_Item"] == "Y" or Current_Item["HQ_Confirmation_Line_Flag"] == "Cancelled" or Current_Item["HQ_Confirmation_Line_Flag"] == "Finished":
                        HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                        continue
                    # Prepare Current HQ Item Transport Register lines - Delivery
                    Current_HQ_ITR_Register_No = Set_HQITR_Register_No
                    Current_HQ_ITR_Register_Sub_No = ""
                    Current_HQ_ITR_Document_No = Current_PH_Document_No
                    Current_HQ_ITR_Document_Line = ""
                    Current_HQ_ITR_Document_Type = Current_PH_Document_Type
                    Current_HQ_ITR_Vendor_Document_Type = "Delivery"
                    Current_HQ_ITR_Vendor_Document_No = Current_Item["DEL_Delivery_No"]
                    Current_HQ_ITR_Vendor_Shipment_No = ""
                    Current_HQ_ITR_Line_Type = Current_Item["Line_Type"] 
                    if Current_Item["HQ_Confirmation_Line_Flag"] == "Substituted":
                        Current_HQ_ITR_Item_No = Current_Item["HQ_SUB_New_Item"]
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["HQ_SUB_New_Item"]
                    else:
                        Current_HQ_ITR_Item_No = Current_Item["Item_No"]
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["Item_No"]
                    Current_HQ_ITR_Ordered_Item_No = ""
                    Current_HQ_ITR_Quantity = Current_Item["DEL_Line_Qty"]
                    if Current_Item["Item_Free_Of_Charge"] == "Y":
                        Current_HQ_ITR_Ordered_Quantity =  ""    
                    else:
                        Current_HQ_ITR_Ordered_Quantity =  Current_Item["Item_Line_Quantity"]
                    if Current_Item["Item_SN_Tracking"] == "Y":
                        Current_HQ_ITR_Serial_No = "USED"
                    else:
                        Current_HQ_ITR_Serial_No = ""
                    Current_HQ_ITR_Unit_of_Measure = Current_Item["Item_Unit_of_Measure"]
                    Current_HQ_ITR_Currency_Code = ""
                    Current_HQ_ITR_Unit_Price = ""
                    Current_HQ_ITR_Line_Amount = ""
                    Current_HQ_ITR_Order_Date = HQ_General_Setup_df.iloc[0]["HQ_Order_Date"]
                    Current_HQ_ITR_Posting_Date = ""
                    if Current_Item["Item_Connected_to_BOM"] == "Y":
                        BOM_Item = Current_Item["BOM_Item_Relation"]
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == BOM_Item
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(Export_record.iloc[0]["DEL_Line_Number"])
                    elif Current_Item["Item_Free_Of_Charge"] == "Y":
                        Main_Item = Current_Item["Item_Free_Of_Charge_Relation"]
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Main_Item
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(Export_record.iloc[0]["DEL_Line_Number"])
                    else:
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Current_Item["Item_No"] 
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(Export_record.iloc[0]["DEL_Line_Number"])
                    Current_HQ_ITR_Vendor_Line_No = HQ_Vendor_Line_Number
                    Current_HQ_ITR_Vendor_No = PO_Document_Header_Setup_df.iloc[0]["Buy_from_Vendor_No"]
                    Current_HQ_ITR_Country_Region_of_Origin_Code = ""
                    Current_HQ_ITR_Plant_No = ""
                    Current_HQ_ITR_Line_Flag = ""
                    Current_HQ_ITR_Communication_Process_Status = HQ_Delivery_Setup_df.iloc[0]["HQ_Delviery_E_Com_Process_Stat"]
                    Current_HQ_ITR_ATP_Check_Cumulative_Quantity = "0"
                    Current_HQ_ITR_Vendor_Document_Created_Date = str(New_Delivery_Document_Date)
                    Current_HQ_ITR_Tariff_Number = ""
                    Current_HQ_ITR_Vendor_Delivery_No = Current_HQ_ITR_Vendor_Document_No
                    Current_HQ_ITR_Quantity_to_Deliver = Current_HQ_ITR_Quantity
                    Current_HQ_ITR_Vendor_Invoice_No = ""
                    Current_HQ_ITR_Quantity_to_Invoice = "0"
                    # Date movements for Delivery Datest
                    Day1 = datetime.strptime(Current_HQ_ITR_Vendor_Document_Created_Date, "%d.%m.%Y")
                    Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                    Current_HQ_ITR_Picking_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])

                    Day1 = datetime.strptime(Current_HQ_ITR_Picking_Date, "%d.%m.%Y")
                    Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                    Current_HQ_ITR_Trans_Planning_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])

                    Day1 = datetime.strptime(Current_HQ_ITR_Trans_Planning_Date, "%d.%m.%Y")
                    Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                    Current_HQ_ITR_Loading_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])
                    
                    Day1 = datetime.strptime(Current_HQ_ITR_Loading_Date, "%d.%m.%Y")
                    Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                    Current_HQ_ITR_Planned_GI_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])

                    Day1 = datetime.strptime(Current_HQ_ITR_Planned_GI_Date, "%d.%m.%Y")
                    Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                    Current_HQ_ITR_Delivery_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])
                    Current_HQ_ITR_Delivery_Start_Date = str(Current_HQ_ITR_Delivery_Date)
                    Current_HQ_ITR_Delivery_End_Date = str(Current_HQ_ITR_Delivery_Date)
                    Current_HQ_ITR_Receipt_Date = str(Current_HQ_ITR_Delivery_Date)
                    Current_HQ_ITR_Document_Order_Number = HQ_Document_Order_Number
                    Current_HQ_ITR_To_Post_Auto = HQ_Delivery_Setup_df.iloc[0]["HQ_Delviery_To_Post_Auto"]

                    # Prepare HQ Item Transport Register - Export lists
                    Export_HQ_ITR_Register_No_list.append(str(Current_HQ_ITR_Register_No))
                    Export_HQ_ITR_Register_Sub_No_list.append(str(Current_HQ_ITR_Register_Sub_No))
                    Export_HQ_ITR_Document_No_list.append(str(Current_HQ_ITR_Document_No))
                    Export_HQ_ITR_Document_Line_list.append(str(Current_HQ_ITR_Document_Line))
                    Export_HQ_ITR_Document_Type_list.append(str(Current_HQ_ITR_Document_Type))
                    Export_HQ_ITR_Vendor_Document_Type_list.append(str(Current_HQ_ITR_Vendor_Document_Type))
                    Export_HQ_ITR_Vendor_Document_No_list.append(str(Current_HQ_ITR_Vendor_Document_No))
                    Export_HQ_ITR_Vendor_Shipment_No_list.append(str(Current_HQ_ITR_Vendor_Shipment_No))
                    Export_HQ_ITR_Line_Type_list.append(str(Current_HQ_ITR_Line_Type))
                    Export_HQ_ITR_Item_No_list.append(str(Current_HQ_ITR_Item_No))
                    Export_HQ_ITR_Vendor_Item_No_list.append(str(Current_HQ_ITR_Vendor_Item_No))
                    Export_HQ_ITR_Ordered_Item_No_list.append(str(Current_HQ_ITR_Ordered_Item_No))
                    Export_HQ_ITR_Quantity_list.append(str(Current_HQ_ITR_Quantity))
                    Export_HQ_ITR_Ordered_Quantity_list.append(str(Current_HQ_ITR_Ordered_Quantity))
                    Export_HQ_ITR_Serial_No_list.append(str(Current_HQ_ITR_Serial_No))
                    Export_HQ_ITR_Unit_of_Measure_list.append(str(Current_HQ_ITR_Unit_of_Measure))
                    Export_HQ_ITR_Currency_Code_list.append(str(Current_HQ_ITR_Currency_Code))
                    Export_HQ_ITR_Unit_Price_list.append(str(Current_HQ_ITR_Unit_Price))
                    Export_HQ_ITR_Line_Amount_list.append(str(Current_HQ_ITR_Line_Amount))
                    Export_HQ_ITR_Order_Date_list.append(str(Current_HQ_ITR_Order_Date))
                    Export_HQ_ITR_Receipt_Date_list.append(str(Current_HQ_ITR_Receipt_Date))
                    Export_HQ_ITR_Posting_Date_list.append(str(Current_HQ_ITR_Posting_Date))
                    Export_HQ_ITR_Exported_Line_No_list.append(str(Current_HQ_ITR_Exported_Line_No))
                    Export_HQ_ITR_Vendor_Line_No_list.append(str(Current_HQ_ITR_Vendor_Line_No))
                    Export_HQ_ITR_Vendor_No_list.append(str(Current_HQ_ITR_Vendor_No))
                    Export_HQ_ITR_Country_Region_of_Origin_Code_list.append(str(Current_HQ_ITR_Country_Region_of_Origin_Code))
                    Export_HQ_ITR_Plant_No_list.append(str(Current_HQ_ITR_Plant_No))
                    Export_HQ_ITR_Line_Flag_list.append(str(Current_HQ_ITR_Line_Flag))
                    Export_HQ_ITR_Communication_Process_Status_list.append(str(Current_HQ_ITR_Communication_Process_Status))
                    Export_HQ_ITR_ATP_Check_Cumulative_Quantity_list.append(str(Current_HQ_ITR_ATP_Check_Cumulative_Quantity))
                    Export_HQ_ITR_Vendor_Document_Created_Date_list.append(str(Current_HQ_ITR_Vendor_Document_Created_Date))
                    Export_HQ_ITR_Delivery_Start_Date_list.append(str(Current_HQ_ITR_Delivery_Start_Date))
                    Export_HQ_ITR_Delivery_End_Date_list.append(str(Current_HQ_ITR_Delivery_End_Date))
                    Export_HQ_ITR_Tariff_Number_list.append(str(Current_HQ_ITR_Tariff_Number))
                    Export_HQ_ITR_Vendor_Delivery_No_list.append(str(Current_HQ_ITR_Vendor_Delivery_No))
                    Export_HQ_ITR_Quantity_to_Deliver_list.append(str(Current_HQ_ITR_Quantity_to_Deliver))
                    Export_HQ_ITR_Vendor_Invoice_No_list.append(str(Current_HQ_ITR_Vendor_Invoice_No))
                    Export_HQ_ITR_Quantity_to_Invoice_list.append(str(Current_HQ_ITR_Quantity_to_Invoice))
                    Export_HQ_ITR_Picking_Date_list.append(str(Current_HQ_ITR_Picking_Date))
                    Export_HQ_ITR_Trans_Planning_Date_list.append(str(Current_HQ_ITR_Trans_Planning_Date))
                    Export_HQ_ITR_Loading_Date_list.append(str(Current_HQ_ITR_Loading_Date))
                    Export_HQ_ITR_Planned_GI_Date_list.append(str(Current_HQ_ITR_Planned_GI_Date))
                    Export_HQ_ITR_Delivery_Date_list.append(str(Current_HQ_ITR_Delivery_Date))
                    Export_HQ_ITR_Document_Order_Number_list.append(str(Current_HQ_ITR_Document_Order_Number))
                    Export_HQ_ITR_To_Post_Auto_list.append(str(Current_HQ_ITR_To_Post_Auto))

                    # HQ SN Regsiter
                    if Current_Item["Item_SN_Tracking"] == "Y":
                        for SN in range(int(Current_HQ_ITR_Quantity)):
                            Current_SN_Suffix = str(Current_SN_Number).zfill(HQ_SN_Sufix_Counter)
                            # Prepare Current HQ Serial Number Register lines
                            Current_HQ_SNR_HQ_SN_Register_ID = Set_HQSNR_Register_No
                            Current_HQ_SNR_HQ_SN_Sub_Register_ID = ""
                            Current_HQ_SNR_HQ_Item_Transport_Register_ID = Current_HQ_ITR_Register_No
                            Current_HQ_SNR_HQ_Item_Transport_Sub_Reg_ID = ""
                            Current_HQ_SNR_Purchase_Order_No = Current_PH_Document_No
                            Current_HQ_SNR_Item_No = Current_HQ_ITR_Item_No     # Substituci nemusím řešit protože bere Item ktery je už substituován
                            Current_HQ_SNR_Serial_Number = str(HQ_HQSNR_Setup_df.iloc[0]["HQ_SN_SN_prefix"])+str(Current_SN_Suffix)
                            Current_HQ_SNR_Vendor_Document_No = Current_HQ_ITR_Vendor_Document_No
                            # Prepare HQ ISerial Number Register - Export lists
                            Export_HQ_SNR_HQ_SN_Register_ID_list.append(str(Current_HQ_SNR_HQ_SN_Register_ID))
                            Export_HQ_SNR_HQ_SN_Sub_Register_ID_list.append(str(Current_HQ_SNR_HQ_SN_Sub_Register_ID))
                            Export_HQ_SNR_HQ_Item_Transport_Register_ID_list.append(str(Current_HQ_SNR_HQ_Item_Transport_Register_ID))
                            Export_HQ_SNR_HQ_Item_Transport_Sub_Reg_ID_list.append(str(Current_HQ_SNR_HQ_Item_Transport_Sub_Reg_ID))
                            Export_HQ_SNR_Purchase_Order_No_list.append(str(Current_HQ_SNR_Purchase_Order_No))
                            Export_HQ_SNR_Item_No_list.append(str(Current_HQ_SNR_Item_No))
                            Export_HQ_SNR_Serial_Number_list.append(str(Current_HQ_SNR_Serial_Number))
                            Export_HQ_SNR_Vendor_Document_No_list.append(str(Current_HQ_SNR_Vendor_Document_No))
                            # Final HQ SN REgister set
                            Set_HQSNR_Register_No += int(HQ_HQSNR_Setup_df.iloc[0]["HQ_HQSNR_Register_No_Increment"])
                            Current_SN_Number += int(HQ_HQSNR_Setup_df.iloc[0]["HQ_SN_SN_Increment"])
                    else:
                        pass

                    # HQ Delivery Tracking Regsiter
                    Current_Package_Suffix = str(Current_Package_Number).zfill(HQ_Package_Sufix_Counter)
                    Current_Bill_Suffix = str(Current_Bill_Number).zfill(HQ_Bill_Sufix_Counter)
                    Current_EXIDV_Suffix = str(Current_EXIDV_Number).zfill(HQ_EXIDV_Sufix_Counter)
                    if Questions_df.iloc[0]["HQ_DeL_Track_Quest"] == "Y" and HQ_DTR_Per_Delivery_counter == 1:
                        Current_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID = Set_HQDTR_Regsiter_No
                        Current_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID = ""
                        Current_HQ_HQDTR_Delivery_No = Current_HQ_ITR_Vendor_Document_No
                        Current_HQ_HQDTR_Purchase_Order_No = Current_PH_Document_No
                        Current_HQ_HQDTR_Package_Tracking_No = str(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Packa_prefix"])+str(Current_Package_Suffix)
                        Current_HQ_HQDTR_Package_Type = "Dummy Verpackungsmaterial"
                        Current_HQ_HQDTR_BEU_Bill_of_Landing = str(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Bill_prefix"])+str(Current_Bill_Suffix)
                        Current_HQ_HQDTR_Tracking_Page = ""
                        Current_HQ_HQDTR_Incoterms = random.choice(str(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Shipment_Method"]).split(","))
                        if HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Random_Weight"] == "Y":
                            Current_HQ_HQDTR_Total_Weight = random.randrange(0, int(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Weight"]))
                        else:
                            Current_HQ_HQDTR_Total_Weight = HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Weight"]
                        Current_HQ_HQDTR_Weight_Unit = HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Weight_UOM"]
                        if HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Random_Volume"] == "Y":
                            Current_HQ_HQDTR_Volume = random.randrange(0, int(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Volume"]))
                        else:
                            Current_HQ_HQDTR_Volume = HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Volume"]
                        Current_HQ_HQDTR_Volume_Unit = HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Volume_UOM"]
                        Current_HQ_HQDTR_Shipping_Agent_Code = random.choice(str(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Shipping_Agent"]).split(","))
                        Current_HQ_HQDTR_EXIDV2 = str(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_EXIDV_prefix"])+str(Current_EXIDV_Suffix)

                        # Prepare HQ Delivery Tracking Register - Export lists
                        Export_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID_list.append(str(Current_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID))
                        Export_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID_list.append(str(Current_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID))
                        Export_HQ_HQDTR_Delivery_No_list.append(str(Current_HQ_HQDTR_Delivery_No))
                        Export_HQ_HQDTR_Purchase_Order_No_list.append(str(Current_HQ_HQDTR_Purchase_Order_No))
                        Export_HQ_HQDTR_Package_Tracking_No_list.append(str(Current_HQ_HQDTR_Package_Tracking_No))
                        Export_HQ_HQDTR_Package_Type_list.append(str(Current_HQ_HQDTR_Package_Type))
                        Export_HQ_HQDTR_BEU_Bill_of_Landing_list.append(str(Current_HQ_HQDTR_BEU_Bill_of_Landing))
                        Export_HQ_HQDTR_Tracking_Page_list.append(str(Current_HQ_HQDTR_Tracking_Page))
                        Export_HQ_HQDTR_Incoterms_list.append(str(Current_HQ_HQDTR_Incoterms))
                        Export_HQ_HQDTR_Total_Weight_list.append(str(Current_HQ_HQDTR_Total_Weight))
                        Export_HQ_HQDTR_Weight_Unit_list.append(str(Current_HQ_HQDTR_Weight_Unit))
                        Export_HQ_HQDTR_Volume_list.append(str(Current_HQ_HQDTR_Volume))
                        Export_HQ_HQDTR_Volume_Unit_list.append(str(Current_HQ_HQDTR_Volume_Unit))
                        Export_HQ_HQDTR_Shipping_Agent_Code_list.append(str(Current_HQ_HQDTR_Shipping_Agent_Code))
                        Export_HQ_HQDTR_EXIDV2_list.append(str(Current_HQ_HQDTR_EXIDV2))
                        # Final HQ Package Tracking Register set
                        Set_HQDTR_Regsiter_No += int(HQ_DEL_Track_Setup_df.iloc[0]["HQ_HQDTR_Register_No_Increment"])
                    else:
                        pass

                    # HQ Package Tracking Register
                    if Questions_df.iloc[0]["HQ_Packg_Track_Quest"] == "Y":
                        Current_HQ_HQPTR_HQ_Package_Tracking_Register_ID = Set_HQPTR_Regsiter_No
                        Current_HQ_HQPTR_HQ_Package_Tracking_Sub_Register_ID = ""
                        Current_HQ_HQPTR_Delivery_No = Current_HQ_ITR_Vendor_Document_No
                        Current_HQ_HQPTR_Package_No = Current_HQ_HQDTR_Package_Tracking_No
                        Current_HQ_HQPTR_Item = Current_HQ_ITR_Item_No
                        Current_HQ_HQPTR_Quantity = Current_HQ_ITR_Quantity
                        Current_HQ_HQPTR_Unit_of_Measure = Current_HQ_ITR_Unit_of_Measure
                        Current_HQ_HQPTR_Purchase_Order_No = Current_PH_Document_No
                        Current_HQ_HQPTR_External_Package_ID = Current_HQ_HQDTR_EXIDV2
                        Current_HQ_HQPTR_Total_Weight = Current_HQ_HQDTR_Total_Weight
                        Current_HQ_HQPTR_Weight_Unit = Current_HQ_HQDTR_Weight_Unit
                        Current_HQ_HQPTR_Volume = Current_HQ_HQDTR_Volume
                        Current_HQ_HQPTR_Volume_Unit = Current_HQ_HQDTR_Volume_Unit
                        Current_HQ_HQPTR_Plant_No = random.choice(str(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_Plant"]).split(","))
                        # Prepare HQ Package Tracking Register - Export lists
                        Export_HQ_HQPTR_HQ_Package_Tracking_Register_ID_list.append(str(Current_HQ_HQPTR_HQ_Package_Tracking_Register_ID))
                        Export_HQ_HQPTR_HQ_Package_Tracking_Sub_Register_ID_list.append(str(Current_HQ_HQPTR_HQ_Package_Tracking_Sub_Register_ID))
                        Export_HQ_HQPTR_Delivery_No_list.append(str(Current_HQ_HQPTR_Delivery_No))
                        Export_HQ_HQPTR_Package_No_list.append(str(Current_HQ_HQPTR_Package_No))
                        Export_HQ_HQPTR_Item_list.append(str(Current_HQ_HQPTR_Item))
                        Export_HQ_HQPTR_Quantity_list.append(str(Current_HQ_HQPTR_Quantity))
                        Export_HQ_HQPTR_Unit_of_Measure_list.append(str(Current_HQ_HQPTR_Unit_of_Measure))
                        Export_HQ_HQPTR_Purchase_Order_No_list.append(str(Current_HQ_HQPTR_Purchase_Order_No))
                        Export_HQ_HQPTR_External_Package_ID_list.append(str(Current_HQ_HQPTR_External_Package_ID))
                        Export_HQ_HQPTR_Total_Weight_list.append(str(Current_HQ_HQPTR_Total_Weight))
                        Export_HQ_HQPTR_Weight_Unit_list.append(str(Current_HQ_HQPTR_Weight_Unit))
                        Export_HQ_HQPTR_Volume_list.append(str(Current_HQ_HQPTR_Volume))
                        Export_HQ_HQPTR_Volume_Unit_list.append(str(Current_HQ_HQPTR_Volume_Unit))
                        Export_HQ_HQPTR_Plant_No_list.append(str(Current_HQ_HQPTR_Plant_No))
                        # Final HQ Package Tracking Register set
                        Set_HQPTR_Regsiter_No += int(HQ_Packg_Track_Setup_df.iloc[0]["HQ_HQPTR_Register_No_Increment"])
                    else:
                        pass

                    # Final Export line setup
                    Set_HQITR_Register_No += int(HQ_General_Setup_df.iloc[0]["HQ_HQITR_Register_No_Increment"])
                    HQ_Document_Order_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                    HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                    HQ_DTR_Per_Delivery_counter += 1

                Current_Package_Number += int(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Packa_Increment"])
                Current_Bill_Number += int(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Bill_Increment"])
                Current_EXIDV_Number += int(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_EXIDV_Increment"])
                Day1 = datetime.strptime(New_Delivery_Document_Date, "%d.%m.%Y")
                Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                New_Delivery_Document_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])
        else:
            pass
    elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
        if Questions_df.iloc[0]["HQ_Del_Quest"] == "Y":
            # Multiple Delivery per Order
            New_Delivery_Document_Date = HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_Vendor_Document_Created_Date"]
            for counter in range(int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"])):
                # Define proper lines in DEL_Item_Assign_df related only to the one delivery
                mask = DEL_Item_Assign_df["DEL_Counter"] == counter
                DEL_Detail_Item_Assign_df = DEL_Item_Assign_df[mask]

                HQ_Document_Order_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
                HQ_Vendor_Line_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
                HQ_DTR_Per_Delivery_counter = 1
                for item in DEL_Detail_Item_Assign_df.iterrows():
                    Current_Item = pandas.Series(data = item[1])
                    if Current_Item["Main_BOM_Item"] == "Y" or Current_Item["HQ_Confirmation_Line_Flag"] == "Cancelled" or Current_Item["HQ_Confirmation_Line_Flag"] == "Finished":
                        HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                        continue
                               
                    Current_Quantity = int(Current_Item["DEL_Line_Qty"])    # Helps define number of lines with SN
                    if Current_Item["Item_SN_Tracking"] == "Y":
                        for SN in range(int(Current_Quantity)):
                            Current_SN_Suffix = str(Current_SN_Number).zfill(HQ_SN_Sufix_Counter)
                            # Prepare Current HQ Item Transport Register lines - Delivery
                            Current_HQ_ITR_Register_No = Set_HQITR_Register_No
                            Current_HQ_ITR_Register_Sub_No = ""
                            Current_HQ_ITR_Document_No = Current_PH_Document_No
                            Current_HQ_ITR_Document_Line = ""
                            Current_HQ_ITR_Document_Type = Current_PH_Document_Type
                            Current_HQ_ITR_Vendor_Document_Type = "Delivery"
                            Current_HQ_ITR_Vendor_Document_No = Current_Item["DEL_Delivery_No"]
                            Current_HQ_ITR_Vendor_Shipment_No = ""
                            Current_HQ_ITR_Line_Type = Current_Item["Line_Type"] 
                            if Current_Item["HQ_Confirmation_Line_Flag"] == "Substituted":
                                Current_HQ_ITR_Item_No = Current_Item["HQ_SUB_New_Item"]
                                Current_HQ_ITR_Vendor_Item_No = Current_Item["HQ_SUB_New_Item"]
                            else:
                                Current_HQ_ITR_Item_No = Current_Item["Item_No"]
                                Current_HQ_ITR_Vendor_Item_No = Current_Item["Item_No"]
                            Current_HQ_ITR_Ordered_Item_No = ""
                            Current_HQ_ITR_Quantity = "1"
                            if Current_Item["Item_Free_Of_Charge"] == "Y":
                                Current_HQ_ITR_Ordered_Quantity =  ""    
                            else:
                                Current_HQ_ITR_Ordered_Quantity =  Current_Item["Item_Line_Quantity"]
                            Current_HQ_ITR_Serial_No = str(HQ_HQSNR_Setup_df.iloc[0]["HQ_SN_SN_prefix"])+str(Current_SN_Suffix)
                            Current_HQ_ITR_Unit_of_Measure = Current_Item["Item_Unit_of_Measure"]
                            Current_HQ_ITR_Currency_Code = ""
                            Current_HQ_ITR_Unit_Price = ""
                            Current_HQ_ITR_Line_Amount = ""
                            Current_HQ_ITR_Order_Date = HQ_General_Setup_df.iloc[0]["HQ_Order_Date"]
                            Current_HQ_ITR_Posting_Date = ""
                            if Current_Item["Item_Connected_to_BOM"] == "Y":
                                BOM_Item = Current_Item["BOM_Item_Relation"]
                                mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == BOM_Item
                                mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                                Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                                Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["DEL_Line_Number"])//100)
                            elif Current_Item["Item_Free_Of_Charge"] == "Y":
                                Main_Item = Current_Item["Item_Free_Of_Charge_Relation"]
                                mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Main_Item
                                mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                                Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                                Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["DEL_Line_Number"])//100)
                            else:
                                mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Current_Item["Item_No"] 
                                mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                                Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                                Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["DEL_Line_Number"])//100)
                            Current_HQ_ITR_Vendor_Line_No = HQ_Vendor_Line_Number
                            Current_HQ_ITR_Vendor_No = PO_Document_Header_Setup_df.iloc[0]["Buy_from_Vendor_No"]
                            Current_HQ_ITR_Country_Region_of_Origin_Code = ""
                            Current_HQ_ITR_Plant_No = ""
                            Current_HQ_ITR_Line_Flag = ""
                            Current_HQ_ITR_Communication_Process_Status = HQ_Delivery_Setup_df.iloc[0]["HQ_Delviery_E_Com_Process_Stat"]
                            Current_HQ_ITR_Vendor_Document_Created_Date = str(New_Delivery_Document_Date)
                            Current_HQ_ITR_Receipt_Date = HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_Vendor_Document_Created_Date"]
                            Current_HQ_ATP_Stock = ""

                            # Prepare HQ Item Transport Register - Export lists
                            Export_HQ_ITR_Register_No_list.append(str(Current_HQ_ITR_Register_No))
                            Export_HQ_ITR_Register_Sub_No_list.append(str(Current_HQ_ITR_Register_Sub_No))
                            Export_HQ_ITR_Document_No_list.append(str(Current_HQ_ITR_Document_No))
                            Export_HQ_ITR_Document_Line_list.append(str(Current_HQ_ITR_Document_Line))
                            Export_HQ_ITR_Document_Type_list.append(str(Current_HQ_ITR_Document_Type))
                            Export_HQ_ITR_Vendor_Document_Type_list.append(str(Current_HQ_ITR_Vendor_Document_Type))
                            Export_HQ_ITR_Vendor_Document_No_list.append(str(Current_HQ_ITR_Vendor_Document_No))
                            Export_HQ_ITR_Vendor_Shipment_No_list.append(str(Current_HQ_ITR_Vendor_Shipment_No))
                            Export_HQ_ITR_Line_Type_list.append(str(Current_HQ_ITR_Line_Type))
                            Export_HQ_ITR_Item_No_list.append(str(Current_HQ_ITR_Item_No))
                            Export_HQ_ITR_Vendor_Item_No_list.append(str(Current_HQ_ITR_Vendor_Item_No))
                            Export_HQ_ITR_Ordered_Item_No_list.append(str(Current_HQ_ITR_Ordered_Item_No))
                            Export_HQ_ITR_Quantity_list.append(str(Current_HQ_ITR_Quantity))
                            Export_HQ_ITR_Ordered_Quantity_list.append(str(Current_HQ_ITR_Ordered_Quantity))
                            Export_HQ_ITR_Serial_No_list.append(str(Current_HQ_ITR_Serial_No))
                            Export_HQ_ITR_Unit_of_Measure_list.append(str(Current_HQ_ITR_Unit_of_Measure))
                            Export_HQ_ITR_Currency_Code_list.append(str(Current_HQ_ITR_Currency_Code))
                            Export_HQ_ITR_Unit_Price_list.append(str(Current_HQ_ITR_Unit_Price))
                            Export_HQ_ITR_Line_Amount_list.append(str(Current_HQ_ITR_Line_Amount))
                            Export_HQ_ITR_Order_Date_list.append(str(Current_HQ_ITR_Order_Date))
                            Export_HQ_ITR_Receipt_Date_list.append(str(Current_HQ_ITR_Receipt_Date))
                            Export_HQ_ITR_Posting_Date_list.append(str(Current_HQ_ITR_Posting_Date))
                            Export_HQ_ITR_Exported_Line_No_list.append(str(Current_HQ_ITR_Exported_Line_No))
                            Export_HQ_ITR_Vendor_Line_No_list.append(str(Current_HQ_ITR_Vendor_Line_No))
                            Export_HQ_ITR_Vendor_No_list.append(str(Current_HQ_ITR_Vendor_No))
                            Export_HQ_ITR_Country_Region_of_Origin_Code_list.append(str(Current_HQ_ITR_Country_Region_of_Origin_Code))
                            Export_HQ_ITR_Plant_No_list.append(str(Current_HQ_ITR_Plant_No))
                            Export_HQ_ITR_Line_Flag_list.append(str(Current_HQ_ITR_Line_Flag))
                            Export_HQ_ITR_Communication_Process_Status_list.append(str(Current_HQ_ITR_Communication_Process_Status))
                            Export_HQ_ITR_Vendor_Document_Created_Date_list.append(str(Current_HQ_ITR_Vendor_Document_Created_Date))
                            Export_HQ_ATP_Stock_list.append(str(Current_HQ_ATP_Stock))

                            # HQ Transfer Tracking No. (just reused "HQ Delivery Tracking register variables")
                            Current_Package_Suffix = str(Current_Package_Number).zfill(HQ_Package_Sufix_Counter)
                            if HQ_Document_Order_Number == 10:
                                Current_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID = Current_HQ_ITR_Register_No
                                Current_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID = ""
                                Current_HQ_HQDTR_Purchase_Order_No = Current_PH_Document_No
                                Current_HQ_HQDTR_Package_Tracking_No = str(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Packa_prefix"])+str(Current_Package_Suffix)
                                # Prepare HQ Delivery Tracking Register - Export lists
                                Export_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID_list.append(str(Current_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID))
                                Export_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID_list.append(str(Current_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID))
                                Export_HQ_HQDTR_Purchase_Order_No_list.append(str(Current_HQ_HQDTR_Purchase_Order_No))
                                Export_HQ_HQDTR_Package_Tracking_No_list.append(str(Current_HQ_HQDTR_Package_Tracking_No))                          

                            # Final Export line setup in loop of SNs
                            Set_HQITR_Register_No += int(HQ_General_Setup_df.iloc[0]["HQ_HQITR_Register_No_Increment"])
                            Current_SN_Number += int(HQ_HQSNR_Setup_df.iloc[0]["HQ_SN_SN_Increment"])
                            HQ_Document_Order_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                    else:
                        # Prepare Current HQ Item Transport Register lines - Delivery
                        Current_HQ_ITR_Register_No = Set_HQITR_Register_No
                        Current_HQ_ITR_Register_Sub_No = ""
                        Current_HQ_ITR_Document_No = Current_PH_Document_No
                        Current_HQ_ITR_Document_Line = ""
                        Current_HQ_ITR_Document_Type = Current_PH_Document_Type
                        Current_HQ_ITR_Vendor_Document_Type = "Delivery"
                        Current_HQ_ITR_Vendor_Document_No = Current_Item["DEL_Delivery_No"]
                        Current_HQ_ITR_Vendor_Shipment_No = ""
                        Current_HQ_ITR_Line_Type = Current_Item["Line_Type"] 
                        if Current_Item["HQ_Confirmation_Line_Flag"] == "Substituted":
                            Current_HQ_ITR_Item_No = Current_Item["HQ_SUB_New_Item"]
                            Current_HQ_ITR_Vendor_Item_No = Current_Item["HQ_SUB_New_Item"]
                        else:
                            Current_HQ_ITR_Item_No = Current_Item["Item_No"]
                            Current_HQ_ITR_Vendor_Item_No = Current_Item["Item_No"]
                        Current_HQ_ITR_Ordered_Item_No = ""
                        Current_HQ_ITR_Quantity = Current_Item["DEL_Line_Qty"]
                        if Current_Item["Item_Free_Of_Charge"] == "Y":
                            Current_HQ_ITR_Ordered_Quantity =  ""    
                        else:
                            Current_HQ_ITR_Ordered_Quantity =  Current_Item["Item_Line_Quantity"]
                        Current_HQ_ITR_Serial_No = ""
                        Current_HQ_ITR_Unit_of_Measure = Current_Item["Item_Unit_of_Measure"]
                        Current_HQ_ITR_Currency_Code = ""
                        Current_HQ_ITR_Unit_Price = ""
                        Current_HQ_ITR_Line_Amount = ""
                        Current_HQ_ITR_Order_Date = HQ_General_Setup_df.iloc[0]["HQ_Order_Date"]
                        Current_HQ_ITR_Posting_Date = ""
                        if Current_Item["Item_Connected_to_BOM"] == "Y":
                            BOM_Item = Current_Item["BOM_Item_Relation"]
                            mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == BOM_Item
                            mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                            Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                            Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["DEL_Line_Number"])//100)
                        elif Current_Item["Item_Free_Of_Charge"] == "Y":
                            Main_Item = Current_Item["Item_Free_Of_Charge_Relation"]
                            mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Main_Item
                            mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                            Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                            Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["DEL_Line_Number"])//100)
                        else:
                            mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Current_Item["Item_No"] 
                            mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                            Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                            Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["DEL_Line_Number"])//100)
                        Current_HQ_ITR_Vendor_Line_No = HQ_Vendor_Line_Number
                        Current_HQ_ITR_Vendor_No = PO_Document_Header_Setup_df.iloc[0]["Buy_from_Vendor_No"]
                        Current_HQ_ITR_Country_Region_of_Origin_Code = ""
                        Current_HQ_ITR_Plant_No = ""
                        Current_HQ_ITR_Line_Flag = ""
                        Current_HQ_ITR_Communication_Process_Status = HQ_Delivery_Setup_df.iloc[0]["HQ_Delviery_E_Com_Process_Stat"]
                        Current_HQ_ITR_Vendor_Document_Created_Date = str(New_Delivery_Document_Date)
                        Current_HQ_ITR_Receipt_Date = HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_Vendor_Document_Created_Date"]
                        Current_HQ_ATP_Stock = ""
                        # Prepare HQ Item Transport Register - Export lists
                        Export_HQ_ITR_Register_No_list.append(str(Current_HQ_ITR_Register_No))
                        Export_HQ_ITR_Register_Sub_No_list.append(str(Current_HQ_ITR_Register_Sub_No))
                        Export_HQ_ITR_Document_No_list.append(str(Current_HQ_ITR_Document_No))
                        Export_HQ_ITR_Document_Line_list.append(str(Current_HQ_ITR_Document_Line))
                        Export_HQ_ITR_Document_Type_list.append(str(Current_HQ_ITR_Document_Type))
                        Export_HQ_ITR_Vendor_Document_Type_list.append(str(Current_HQ_ITR_Vendor_Document_Type))
                        Export_HQ_ITR_Vendor_Document_No_list.append(str(Current_HQ_ITR_Vendor_Document_No))
                        Export_HQ_ITR_Vendor_Shipment_No_list.append(str(Current_HQ_ITR_Vendor_Shipment_No))
                        Export_HQ_ITR_Line_Type_list.append(str(Current_HQ_ITR_Line_Type))
                        Export_HQ_ITR_Item_No_list.append(str(Current_HQ_ITR_Item_No))
                        Export_HQ_ITR_Vendor_Item_No_list.append(str(Current_HQ_ITR_Vendor_Item_No))
                        Export_HQ_ITR_Ordered_Item_No_list.append(str(Current_HQ_ITR_Ordered_Item_No))
                        Export_HQ_ITR_Quantity_list.append(str(Current_HQ_ITR_Quantity))
                        Export_HQ_ITR_Ordered_Quantity_list.append(str(Current_HQ_ITR_Ordered_Quantity))
                        Export_HQ_ITR_Serial_No_list.append(str(Current_HQ_ITR_Serial_No))
                        Export_HQ_ITR_Unit_of_Measure_list.append(str(Current_HQ_ITR_Unit_of_Measure))
                        Export_HQ_ITR_Currency_Code_list.append(str(Current_HQ_ITR_Currency_Code))
                        Export_HQ_ITR_Unit_Price_list.append(str(Current_HQ_ITR_Unit_Price))
                        Export_HQ_ITR_Line_Amount_list.append(str(Current_HQ_ITR_Line_Amount))
                        Export_HQ_ITR_Order_Date_list.append(str(Current_HQ_ITR_Order_Date))
                        Export_HQ_ITR_Receipt_Date_list.append(str(Current_HQ_ITR_Receipt_Date))
                        Export_HQ_ITR_Posting_Date_list.append(str(Current_HQ_ITR_Posting_Date))
                        Export_HQ_ITR_Exported_Line_No_list.append(str(Current_HQ_ITR_Exported_Line_No))
                        Export_HQ_ITR_Vendor_Line_No_list.append(str(Current_HQ_ITR_Vendor_Line_No))
                        Export_HQ_ITR_Vendor_No_list.append(str(Current_HQ_ITR_Vendor_No))
                        Export_HQ_ITR_Country_Region_of_Origin_Code_list.append(str(Current_HQ_ITR_Country_Region_of_Origin_Code))
                        Export_HQ_ITR_Plant_No_list.append(str(Current_HQ_ITR_Plant_No))
                        Export_HQ_ITR_Line_Flag_list.append(str(Current_HQ_ITR_Line_Flag))
                        Export_HQ_ITR_Communication_Process_Status_list.append(str(Current_HQ_ITR_Communication_Process_Status))
                        Export_HQ_ITR_Vendor_Document_Created_Date_list.append(str(Current_HQ_ITR_Vendor_Document_Created_Date))
                        Export_HQ_ATP_Stock_list.append(str(Current_HQ_ATP_Stock))

                        # HQ Transfer Tracking No. (just reused "HQ Delivery Tracking register variables")
                        Current_Package_Suffix = str(Current_Package_Number).zfill(HQ_Package_Sufix_Counter)
                        if HQ_Document_Order_Number == 10:
                            Current_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID = Current_HQ_ITR_Register_No
                            Current_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID = ""
                            Current_HQ_HQDTR_Purchase_Order_No = Current_PH_Document_No
                            Current_HQ_HQDTR_Package_Tracking_No = str(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Packa_prefix"])+str(Current_Package_Suffix)
                            # Prepare HQ Delivery Tracking Register - Export lists
                            Export_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID_list.append(str(Current_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID))
                            Export_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID_list.append(str(Current_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID))
                            Export_HQ_HQDTR_Purchase_Order_No_list.append(str(Current_HQ_HQDTR_Purchase_Order_No))
                            Export_HQ_HQDTR_Package_Tracking_No_list.append(str(Current_HQ_HQDTR_Package_Tracking_No))                          

                    # Final Export line setup
                    Set_HQITR_Register_No += int(HQ_General_Setup_df.iloc[0]["HQ_HQITR_Register_No_Increment"])
                    HQ_Document_Order_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                    HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                    HQ_DTR_Per_Delivery_counter += 1

                Current_Package_Number += int(HQ_DEL_Track_Setup_df.iloc[0]["HQ_Del_Track_Packa_Increment"])
                Day1 = datetime.strptime(New_Delivery_Document_Date, "%d.%m.%Y")
                Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                New_Delivery_Document_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])
        else:
            pass
    else:
        pass

    # HQ Item Transport Register - Invoice
    if General_Setup_df.iloc[0]["Navision"] == "NUS3":
        if Questions_df.iloc[0]["HQ_Inv_Quest"] == "Y" and Questions_df.iloc[0]["HQ_Del_Quest"] == "Y":
            # Multiple Delivery per Order
            New_Invoice_Posting_Date = HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_Vendor_Document_Created_Date"]
            for counter in range(int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"])):
                # Define proper lines in DEL_Item_Assign_df related only to the one delivery
                mask = DEL_Item_Assign_df["DEL_Counter"] == counter
                DEL_Detail_Item_Assign_df = DEL_Item_Assign_df[mask]
                
                New_Invoice_Plant = random.choice(str(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_Plant"]).split(","))
                HQ_Document_Order_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
                HQ_Vendor_Line_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
                for item in DEL_Detail_Item_Assign_df.iterrows():
                    # Prepare Current HQ Item Transport Register lines - Invoice
                    Current_Item = pandas.Series(data = item[1])
                    if Current_Item["Main_BOM_Item"] == "Y" or Current_Item["HQ_Confirmation_Line_Flag"] == "Cancelled" or Current_Item["HQ_Confirmation_Line_Flag"] == "Finished":
                        HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                        continue
                    Current_HQ_ITR_Register_No = Set_HQITR_Register_No
                    Current_HQ_ITR_Register_Sub_No = ""
                    Current_HQ_ITR_Document_No = Current_PH_Document_No
                    Current_HQ_ITR_Document_Line = ""
                    Current_HQ_ITR_Document_Type = Current_PH_Document_Type
                    Current_HQ_ITR_Vendor_Document_Type = "Invoice"
                    Current_HQ_ITR_Vendor_Document_No = Current_Item["DEL_Invoice_No"]
                    # Delivery Conection
                    Current_HQ_ITR_Vendor_Shipment_No = Current_Item["DEL_Delivery_No"]

                    Current_HQ_ITR_Line_Type = Current_Item["Line_Type"] 
                    if Current_Item["HQ_Confirmation_Line_Flag"] == "Substituted":
                        Current_HQ_ITR_Item_No = Current_Item["HQ_SUB_New_Item"]
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["HQ_SUB_New_Item"]
                    else:
                        Current_HQ_ITR_Item_No = Current_Item["Item_No"]
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["Item_No"]
                    Current_HQ_ITR_Ordered_Item_No = ""
                    Current_HQ_ITR_Quantity = Current_Item["DEL_Line_Qty"]
                    if Current_Item["Item_Free_Of_Charge"] == "Y":
                        Current_HQ_ITR_Ordered_Quantity =  ""    
                    else:
                        Current_HQ_ITR_Ordered_Quantity =  Current_Item["Item_Line_Quantity"]
                    Current_HQ_ITR_Serial_No = ""
                    Current_HQ_ITR_Unit_of_Measure = Current_Item["Item_Unit_of_Measure"]
                    Current_HQ_ITR_Currency_Code = HQ_General_Setup_df.iloc[0]["HQ_Currency_Code"]
                    Current_HQ_ITR_Unit_Price = str(test_float(Current_Item["Item_Unit_Price"]))
                    Current_HQ_ITR_Line_Amount = str(test_float(int(Current_HQ_ITR_Quantity)*float(Current_HQ_ITR_Unit_Price)))
                    Current_HQ_ITR_Order_Date = HQ_General_Setup_df.iloc[0]["HQ_Order_Date"]
                    Current_HQ_ITR_Receipt_Date = HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_Receipt_Date"]
                    Current_HQ_ITR_Posting_Date = str(New_Invoice_Posting_Date)
                    if Current_Item["Item_Connected_to_BOM"] == "Y":
                        BOM_Item = Current_Item["BOM_Item_Relation"]
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == BOM_Item
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(Export_record.iloc[0]["DEL_Line_Number"])
                    elif Current_Item["Item_Free_Of_Charge"] == "Y":
                        Main_Item = Current_Item["Item_Free_Of_Charge_Relation"]
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Main_Item
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(Export_record.iloc[0]["DEL_Line_Number"])
                    else:
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Current_Item["Item_No"] 
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(Export_record.iloc[0]["DEL_Line_Number"])
                    Current_HQ_ITR_Vendor_Line_No = HQ_Vendor_Line_Number
                    Current_HQ_ITR_Vendor_No = PO_Document_Header_Setup_df.iloc[0]["Buy_from_Vendor_No"]
                    Current_HQ_ITR_Country_Region_of_Origin_Code = random.choice(str(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_Count_Reg_Origin"]).split(","))
                    Current_HQ_ITR_Plant_No = str(New_Invoice_Plant)
                    Current_HQ_ITR_Line_Flag = ""
                    Current_HQ_ITR_Communication_Process_Status = ""
                    Current_HQ_ITR_ATP_Check_Cumulative_Quantity = "0"
                    Current_HQ_ITR_Vendor_Document_Created_Date = str(Current_HQ_ITR_Posting_Date)
                    Current_HQ_ITR_Delivery_Start_Date = HQ_General_Setup_df.iloc[0]["HQ_Delivery_Start_and_End_Date"]
                    Current_HQ_ITR_Delivery_End_Date = HQ_General_Setup_df.iloc[0]["HQ_Delivery_Start_and_End_Date"]
                    Current_HQ_ITR_Tariff_Number = random.choice(str(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_Tariff"]).split(","))
                    Current_HQ_ITR_Vendor_Delivery_No = Current_HQ_ITR_Vendor_Shipment_No
                    Current_HQ_ITR_Quantity_to_Deliver = "0"
                    Current_HQ_ITR_Vendor_Invoice_No = Current_HQ_ITR_Vendor_Document_No
                    Current_HQ_ITR_Quantity_to_Invoice = Current_HQ_ITR_Quantity
                    Current_HQ_ITR_Picking_Date = ""
                    Current_HQ_ITR_Trans_Planning_Date = ""
                    Current_HQ_ITR_Loading_Date = ""
                    Current_HQ_ITR_Planned_GI_Date = ""
                    Current_HQ_ITR_Delivery_Date = ""
                    Current_HQ_ITR_Document_Order_Number = HQ_Document_Order_Number
                    Current_HQ_ITR_To_Post_Auto = HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_To_Post_Auto"]

                    # Prepare HQ Item Transport Register - Export lists
                    Export_HQ_ITR_Register_No_list.append(str(Current_HQ_ITR_Register_No))
                    Export_HQ_ITR_Register_Sub_No_list.append(str(Current_HQ_ITR_Register_Sub_No))
                    Export_HQ_ITR_Document_No_list.append(str(Current_HQ_ITR_Document_No))
                    Export_HQ_ITR_Document_Line_list.append(str(Current_HQ_ITR_Document_Line))
                    Export_HQ_ITR_Document_Type_list.append(str(Current_HQ_ITR_Document_Type))
                    Export_HQ_ITR_Vendor_Document_Type_list.append(str(Current_HQ_ITR_Vendor_Document_Type))
                    Export_HQ_ITR_Vendor_Document_No_list.append(str(Current_HQ_ITR_Vendor_Document_No))
                    Export_HQ_ITR_Vendor_Shipment_No_list.append(str(Current_HQ_ITR_Vendor_Shipment_No))
                    Export_HQ_ITR_Line_Type_list.append(str(Current_HQ_ITR_Line_Type))
                    Export_HQ_ITR_Item_No_list.append(str(Current_HQ_ITR_Item_No))
                    Export_HQ_ITR_Vendor_Item_No_list.append(str(Current_HQ_ITR_Vendor_Item_No))
                    Export_HQ_ITR_Ordered_Item_No_list.append(str(Current_HQ_ITR_Ordered_Item_No))
                    Export_HQ_ITR_Quantity_list.append(str(Current_HQ_ITR_Quantity))
                    Export_HQ_ITR_Ordered_Quantity_list.append(str(Current_HQ_ITR_Ordered_Quantity))
                    Export_HQ_ITR_Serial_No_list.append(str(Current_HQ_ITR_Serial_No))
                    Export_HQ_ITR_Unit_of_Measure_list.append(str(Current_HQ_ITR_Unit_of_Measure))
                    Export_HQ_ITR_Currency_Code_list.append(str(Current_HQ_ITR_Currency_Code))
                    Export_HQ_ITR_Unit_Price_list.append(str(Current_HQ_ITR_Unit_Price))
                    Export_HQ_ITR_Line_Amount_list.append(str(Current_HQ_ITR_Line_Amount))
                    Export_HQ_ITR_Order_Date_list.append(str(Current_HQ_ITR_Order_Date))
                    Export_HQ_ITR_Receipt_Date_list.append(str(Current_HQ_ITR_Receipt_Date))
                    Export_HQ_ITR_Posting_Date_list.append(str(Current_HQ_ITR_Posting_Date))
                    Export_HQ_ITR_Exported_Line_No_list.append(str(Current_HQ_ITR_Exported_Line_No))
                    Export_HQ_ITR_Vendor_Line_No_list.append(str(Current_HQ_ITR_Vendor_Line_No))
                    Export_HQ_ITR_Vendor_No_list.append(str(Current_HQ_ITR_Vendor_No))
                    Export_HQ_ITR_Country_Region_of_Origin_Code_list.append(str(Current_HQ_ITR_Country_Region_of_Origin_Code))
                    Export_HQ_ITR_Plant_No_list.append(str(Current_HQ_ITR_Plant_No))
                    Export_HQ_ITR_Line_Flag_list.append(str(Current_HQ_ITR_Line_Flag))
                    Export_HQ_ITR_Communication_Process_Status_list.append(str(Current_HQ_ITR_Communication_Process_Status))
                    Export_HQ_ITR_ATP_Check_Cumulative_Quantity_list.append(str(Current_HQ_ITR_ATP_Check_Cumulative_Quantity))
                    Export_HQ_ITR_Vendor_Document_Created_Date_list.append(str(Current_HQ_ITR_Vendor_Document_Created_Date))
                    Export_HQ_ITR_Delivery_Start_Date_list.append(str(Current_HQ_ITR_Delivery_Start_Date))
                    Export_HQ_ITR_Delivery_End_Date_list.append(str(Current_HQ_ITR_Delivery_End_Date))
                    Export_HQ_ITR_Tariff_Number_list.append(str(Current_HQ_ITR_Tariff_Number))
                    Export_HQ_ITR_Vendor_Delivery_No_list.append(str(Current_HQ_ITR_Vendor_Delivery_No))
                    Export_HQ_ITR_Quantity_to_Deliver_list.append(str(Current_HQ_ITR_Quantity_to_Deliver))
                    Export_HQ_ITR_Vendor_Invoice_No_list.append(str(Current_HQ_ITR_Vendor_Invoice_No))
                    Export_HQ_ITR_Quantity_to_Invoice_list.append(str(Current_HQ_ITR_Quantity_to_Invoice))
                    Export_HQ_ITR_Picking_Date_list.append(str(Current_HQ_ITR_Picking_Date))
                    Export_HQ_ITR_Trans_Planning_Date_list.append(str(Current_HQ_ITR_Trans_Planning_Date))
                    Export_HQ_ITR_Loading_Date_list.append(str(Current_HQ_ITR_Loading_Date))
                    Export_HQ_ITR_Planned_GI_Date_list.append(str(Current_HQ_ITR_Planned_GI_Date))
                    Export_HQ_ITR_Delivery_Date_list.append(str(Current_HQ_ITR_Delivery_Date))
                    Export_HQ_ITR_Document_Order_Number_list.append(str(Current_HQ_ITR_Document_Order_Number))
                    Export_HQ_ITR_To_Post_Auto_list.append(str(Current_HQ_ITR_To_Post_Auto))

                    # Final Export line setup
                    Set_HQITR_Register_No += int(HQ_General_Setup_df.iloc[0]["HQ_HQITR_Register_No_Increment"])
                    HQ_Document_Order_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                    HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])

                # Record Links:
                if Questions_df.iloc[0]["Record_link_Quest"] == "Y":
                    Link_ID_list.append(str(""))
                    Record_ID_list.append(str("<Binary Data>"))
                    URL1_list.append(str(f"""//{Record_Links_Setup_df.iloc[0]["Server_link"]}/NUS_{Record_Links_Setup_df.iloc[0]["Server"]}/{Record_Links_Setup_df.iloc[0]["NOC"]}/HQ/Archive/PDF/{Current_HQ_ITR_Vendor_Document_No}.pdf"""))
                    URL2_list.append(str(""))
                    URL3_list.append(str(""))
                    URL4_list.append(str(""))
                    Description_list.append(str(Current_HQ_ITR_Vendor_Document_No))
                    Type_list.append(str("Link"))
                    Note_list.append(str(""))
                    Created_list.append(str(""))
                    User_ID_list.append(str(Record_Links_Setup_df.iloc[0]["User ID"]))
                    Company_list.append(str(Record_Links_Setup_df.iloc[0]["Company"]))
                    Notify_list.append(str("No"))
                    To_User_ID_list.append(str(""))

                Day1 = datetime.strptime(New_Invoice_Posting_Date, "%d.%m.%Y")
                Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                New_Invoice_Posting_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])

            # Clean Variables to speed up searching in dataframe
            Var_to_del = ["Exported_Lines_df"]
            for Variable in Var_to_del:      
                try:
                    exec(f'del {Variable}')
                except:
                    pass
            try:
                del Variable
                del Var_to_del
            except:
                pass
        else:
            pass
    elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
        if Questions_df.iloc[0]["HQ_Inv_Quest"] == "Y" and Questions_df.iloc[0]["HQ_Del_Quest"] == "Y":
            # Multiple Delivery per Order
            New_Invoice_Posting_Date = HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_Vendor_Document_Created_Date"]
            for counter in range(int(HQ_Delivery_Setup_df.iloc[0]["Delivery_Count_per_Order"])):
                # Define proper lines in DEL_Item_Assign_df related only to the one delivery
                mask = DEL_Item_Assign_df["DEL_Counter"] == counter
                DEL_Detail_Item_Assign_df = DEL_Item_Assign_df[mask]

                HQ_Vendor_Line_Number = int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Start"])
                for item in DEL_Detail_Item_Assign_df.iterrows():
                    # Prepare Current HQ Item Transport Register lines - Invoice
                    Current_Item = pandas.Series(data = item[1])
                    if Current_Item["Main_BOM_Item"] == "Y" or Current_Item["HQ_Confirmation_Line_Flag"] == "Cancelled" or Current_Item["HQ_Confirmation_Line_Flag"] == "Finished":
                        HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                        continue
                    Current_HQ_ITR_Register_No = Set_HQITR_Register_No
                    Current_HQ_ITR_Register_Sub_No = ""
                    Current_HQ_ITR_Document_No = Current_PH_Document_No
                    Current_HQ_ITR_Document_Line = ""
                    Current_HQ_ITR_Document_Type = Current_PH_Document_Type
                    Current_HQ_ITR_Vendor_Document_Type = "Invoice"
                    Current_HQ_ITR_Vendor_Document_No = Current_Item["DEL_Invoice_No"]
                    # Delivery Conection
                    Current_HQ_ITR_Vendor_Shipment_No = Current_Item["DEL_Delivery_No"]

                    Current_HQ_ITR_Line_Type = Current_Item["Line_Type"] 
                    if Current_Item["HQ_Confirmation_Line_Flag"] == "Substituted":
                        Current_HQ_ITR_Item_No = Current_Item["HQ_SUB_New_Item"]
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["HQ_SUB_New_Item"]
                    else:
                        Current_HQ_ITR_Item_No = Current_Item["Item_No"]
                        Current_HQ_ITR_Vendor_Item_No = Current_Item["Item_No"]
                    Current_HQ_ITR_Ordered_Item_No = ""
                    Current_HQ_ITR_Quantity = Current_Item["DEL_Line_Qty"]
                    if Current_Item["Item_Free_Of_Charge"] == "Y":
                        Current_HQ_ITR_Ordered_Quantity =  ""    
                    else:
                        Current_HQ_ITR_Ordered_Quantity =  Current_Item["Item_Line_Quantity"]
                    Current_HQ_ITR_Serial_No = ""
                    Current_HQ_ITR_Unit_of_Measure = Current_Item["Item_Unit_of_Measure"]
                    Current_HQ_ITR_Currency_Code = HQ_General_Setup_df.iloc[0]["HQ_Currency_Code"]
                    Current_HQ_ITR_Unit_Price = str(test_float(Current_Item["Item_Unit_Price"]))
                    Current_HQ_ITR_Line_Amount = str(test_float(int(Current_HQ_ITR_Quantity)*float(Current_HQ_ITR_Unit_Price)))
                    Current_HQ_ITR_Order_Date = HQ_General_Setup_df.iloc[0]["HQ_Order_Date"]
                    Current_HQ_ITR_Receipt_Date = HQ_Delivery_Setup_df.iloc[0]["HQ_Delivery_Receipt_Date"]
                    Current_HQ_ITR_Posting_Date = str(New_Invoice_Posting_Date)
                    if Current_Item["Item_Connected_to_BOM"] == "Y":
                        BOM_Item = Current_Item["BOM_Item_Relation"]
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == BOM_Item
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["DEL_Line_Number"])//100)
                    elif Current_Item["Item_Free_Of_Charge"] == "Y":
                        Main_Item = Current_Item["Item_Free_Of_Charge_Relation"]
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Main_Item
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["DEL_Line_Number"])//100)
                    else:
                        mask_1 = DEL_Detail_Item_Assign_df["DEL_Item_No"] == Current_Item["Item_No"] 
                        mask_2 = DEL_Detail_Item_Assign_df["DEL_Purchase_Order"] == Current_HQ_ITR_Document_No
                        Export_record = DEL_Detail_Item_Assign_df[mask_1 & mask_2 ]
                        Current_HQ_ITR_Exported_Line_No = str(int(Export_record.iloc[0]["DEL_Line_Number"])//100)
                    Current_HQ_ITR_Vendor_Line_No = HQ_Vendor_Line_Number
                    Current_HQ_ITR_Vendor_No = PO_Document_Header_Setup_df.iloc[0]["Buy_from_Vendor_No"]
                    Current_HQ_ITR_Country_Region_of_Origin_Code = random.choice(str(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_Count_Reg_Origin"]).split(","))
                    Current_HQ_ITR_Plant_No = random.choice(str(HQ_Invoice_Setup_df.iloc[0]["HQ_Invoice_Plant"]).split(","))
                    Current_HQ_ITR_Line_Flag = ""
                    Current_HQ_ITR_Communication_Process_Status = ""
                    Current_HQ_ITR_Vendor_Document_Created_Date = str(New_Invoice_Posting_Date)
                    Current_HQ_ATP_Stock = ""
                    # Prepare HQ Item Transport Register - Export lists
                    Export_HQ_ITR_Register_No_list.append(str(Current_HQ_ITR_Register_No))
                    Export_HQ_ITR_Register_Sub_No_list.append(str(Current_HQ_ITR_Register_Sub_No))
                    Export_HQ_ITR_Document_No_list.append(str(Current_HQ_ITR_Document_No))
                    Export_HQ_ITR_Document_Line_list.append(str(Current_HQ_ITR_Document_Line))
                    Export_HQ_ITR_Document_Type_list.append(str(Current_HQ_ITR_Document_Type))
                    Export_HQ_ITR_Vendor_Document_Type_list.append(str(Current_HQ_ITR_Vendor_Document_Type))
                    Export_HQ_ITR_Vendor_Document_No_list.append(str(Current_HQ_ITR_Vendor_Document_No))
                    Export_HQ_ITR_Vendor_Shipment_No_list.append(str(Current_HQ_ITR_Vendor_Shipment_No))
                    Export_HQ_ITR_Line_Type_list.append(str(Current_HQ_ITR_Line_Type))
                    Export_HQ_ITR_Item_No_list.append(str(Current_HQ_ITR_Item_No))
                    Export_HQ_ITR_Vendor_Item_No_list.append(str(Current_HQ_ITR_Vendor_Item_No))
                    Export_HQ_ITR_Ordered_Item_No_list.append(str(Current_HQ_ITR_Ordered_Item_No))
                    Export_HQ_ITR_Quantity_list.append(str(Current_HQ_ITR_Quantity))
                    Export_HQ_ITR_Ordered_Quantity_list.append(str(Current_HQ_ITR_Ordered_Quantity))
                    Export_HQ_ITR_Serial_No_list.append(str(Current_HQ_ITR_Serial_No))
                    Export_HQ_ITR_Unit_of_Measure_list.append(str(Current_HQ_ITR_Unit_of_Measure))
                    Export_HQ_ITR_Currency_Code_list.append(str(Current_HQ_ITR_Currency_Code))
                    Export_HQ_ITR_Unit_Price_list.append(str(Current_HQ_ITR_Unit_Price))
                    Export_HQ_ITR_Line_Amount_list.append(str(Current_HQ_ITR_Line_Amount))
                    Export_HQ_ITR_Order_Date_list.append(str(Current_HQ_ITR_Order_Date))
                    Export_HQ_ITR_Receipt_Date_list.append(str(Current_HQ_ITR_Receipt_Date))
                    Export_HQ_ITR_Posting_Date_list.append(str(Current_HQ_ITR_Posting_Date))
                    Export_HQ_ITR_Exported_Line_No_list.append(str(Current_HQ_ITR_Exported_Line_No))
                    Export_HQ_ITR_Vendor_Line_No_list.append(str(Current_HQ_ITR_Vendor_Line_No))
                    Export_HQ_ITR_Vendor_No_list.append(str(Current_HQ_ITR_Vendor_No))
                    Export_HQ_ITR_Country_Region_of_Origin_Code_list.append(str(Current_HQ_ITR_Country_Region_of_Origin_Code))
                    Export_HQ_ITR_Plant_No_list.append(str(Current_HQ_ITR_Plant_No))
                    Export_HQ_ITR_Line_Flag_list.append(str(Current_HQ_ITR_Line_Flag))
                    Export_HQ_ITR_Communication_Process_Status_list.append(str(Current_HQ_ITR_Communication_Process_Status))
                    Export_HQ_ITR_Vendor_Document_Created_Date_list.append(str(Current_HQ_ITR_Vendor_Document_Created_Date))
                    Export_HQ_ATP_Stock_list.append(str(Current_HQ_ATP_Stock))
                    # Final Export line setup
                    Set_HQITR_Register_No += int(HQ_General_Setup_df.iloc[0]["HQ_HQITR_Register_No_Increment"])
                    HQ_Vendor_Line_Number += int(HQ_General_Setup_df.iloc[0]["HQ_Document_Order_No_Increment"])
                Day1 = datetime.strptime(New_Invoice_Posting_Date, "%d.%m.%Y")
                Day2 = datetime.strftime((Day1 + timedelta(days=1)), General_Setup_df.iloc[0]["HQ_Date_format"])
                New_Invoice_Posting_Date = test_date(Day2, General_Setup_df.iloc[0]["HQ_Date_format"])

        pass
    else:
        pass

    # Final PO Header Setup
    Current_PO_Number += int(PO_Document_Header_Setup_df.iloc[0]["Document_Number_Increment"])
    Current_HQ_Confirmation_Number += int(HQ_Confirmation_Setup_df.iloc[0]["HQ_Confirmation_Increment"])
    # Clean Variables to speed up searching in dataframe
    # Variable to Delete
    Var_to_del = ["DEL_Item_Assign_df"]
    for Variable in Var_to_del:      
        try:
            exec(f'del {Variable}')
        except:
            pass
    try:
        del Variable
        del Var_to_del
    except:
        pass

# Variable to Delete
Var_to_del = ["Current_HQ_ATP_Document_No", "Current_HQ_ATP_HQ_ATP_Check_Register_ID", "Current_HQ_ATP_HQ_ATP_Check_Sub_Register_ID", "Current_HQ_ATP_HQ_Item_Transport_Register_ID", "Current_HQ_ATP_Scheduled_Date", "Current_HQ_ATP_Scheduled_Quantity", "Current_HQ_ATP_Stock", "Current_HQ_HQDTR_BEU_Bill_of_Landing", "Current_HQ_HQDTR_Delivery_No", "Current_HQ_HQDTR_EXIDV2", "Current_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID", "Current_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID", "Current_HQ_HQDTR_Incoterms", "Current_HQ_HQDTR_Package_Tracking_No", "Current_HQ_HQDTR_Package_Type", "Current_HQ_HQDTR_Purchase_Order_No", "Current_HQ_HQDTR_Shipping_Agent_Code", "Current_HQ_HQDTR_Total_Weight", "Current_HQ_HQDTR_Tracking_Page", "Current_HQ_HQDTR_Volume", "Current_HQ_HQDTR_Volume_Unit", "Current_HQ_HQDTR_Weight_Unit", "Current_HQ_HQPTR_Delivery_No", "Current_HQ_HQPTR_External_Package_ID", "Current_HQ_HQPTR_HQ_Package_Tracking_Register_ID", "Current_HQ_HQPTR_HQ_Package_Tracking_Sub_Register_ID", "Current_HQ_HQPTR_Item", "Current_HQ_HQPTR_Package_No", "Current_HQ_HQPTR_Plant_No", "Current_HQ_HQPTR_Purchase_Order_No", "Current_HQ_HQPTR_Quantity", "Current_HQ_HQPTR_Total_Weight", "Current_HQ_HQPTR_Unit_of_Measure", "Current_HQ_HQPTR_Volume", "Current_HQ_HQPTR_Volume_Unit", "Current_HQ_HQPTR_Weight_Unit", "Current_HQ_ITR_ATP_Check_Cumulative_Quantity", "Current_HQ_ITR_Communication_Process_Status", "Current_HQ_ITR_Country_Region_of_Origin_Code", "Current_HQ_ITR_Currency_Code", "Current_HQ_ITR_Delivery_Date", "Current_HQ_ITR_Delivery_End_Date", "Current_HQ_ITR_Delivery_Start_Date", "Current_HQ_ITR_Document_No", "Current_HQ_ITR_Document_Order_Number", "Current_HQ_ITR_Document_Type", "Current_HQ_ITR_Item_No", "Current_HQ_ITR_Line_Amount", "Current_HQ_ITR_Line_Flag", "Current_HQ_ITR_Line_Type", "Current_HQ_ITR_Loading_Date", "Current_HQ_ITR_Order_Date", "Current_HQ_ITR_Ordered_Item_No", "Current_HQ_ITR_Ordered_Quantity", "Current_HQ_ITR_Picking_Date", "Current_HQ_ITR_Planned_GI_Date", "Current_HQ_ITR_Plant_No", "Current_HQ_ITR_Posting_Date", "Current_HQ_ITR_Quantity", "Current_HQ_ITR_Quantity_to_Deliver", "Current_HQ_ITR_Quantity_to_Invoice", "Current_HQ_ITR_Receipt_Date", "Current_HQ_ITR_Register_No", "Current_HQ_ITR_Register_Sub_No", "Current_HQ_ITR_Serial_No", "Current_HQ_ITR_Tariff_Number", "Current_HQ_ITR_To_Post_Auto", "Current_HQ_ITR_Trans_Planning_Date", "Current_HQ_ITR_Unit_of_Measure", "Current_HQ_ITR_Unit_Price", "Current_HQ_ITR_Vendor_Delivery_No", "Current_HQ_ITR_Vendor_Document_Created_Date", "Current_HQ_ITR_Vendor_Document_No", "Current_HQ_ITR_Vendor_Document_Type", "Current_HQ_ITR_Vendor_Invoice_No", "Current_HQ_ITR_Vendor_Item_No", "Current_HQ_ITR_Vendor_Line_No", "Current_HQ_ITR_Vendor_No", "Current_HQ_ITR_Vendor_Shipment_No", "Current_HQ_PREA_Document_No", "Current_HQ_PREA_HQ_Item_Transport_Reg_Sub_ID", "Current_HQ_PREA_HQ_Item_Transport_Register_ID", "Current_HQ_PREA_HQ_PreAdvice_Register_ID", "Current_HQ_PREA_HQ_PreAdvice_Sub_Register_ID", "Current_HQ_PREA_PreAdvice_Date", "Current_HQ_PREA_PreAdvice_Document_No", "Current_HQ_SNR_HQ_Item_Transport_Register_ID", "Current_HQ_SNR_HQ_Item_Transport_Sub_Reg_ID", "Current_HQ_SNR_HQ_SN_Register_ID", "Current_HQ_SNR_HQ_SN_Sub_Register_ID", "Current_HQ_SNR_Item_No", "Current_HQ_SNR_Purchase_Order_No", "Current_HQ_SNR_Serial_Number", "Current_HQ_SNR_Vendor_Document_No", "Current_HQ_SUB_Document_No", "Current_HQ_SUB_HQ_Substitution_Register_ID", "Current_HQ_SUB_HQ_Substitution_Sub_Reg_ID", "Current_HQ_SUB_Substituted_Item_Old_Item", "Current_HQ_SUB_Substitution_New_Item", "Current_PH_Buy_from_Vendor_No", "Current_PH_Document_No", "Current_PH_Document_Type", "Current_PH_HQ_Logistic_Process", "Current_PH_HQ_Order_Type", "Current_PH_HQ_Shipping_Condition", "Current_PH_Shipping_Agent", "Current_PH_Shipping_Agent_Service", "Current_PL_Document_No", "Current_PL_Document_Type", "Current_PL_Line_No", "Current_PL_No", "Current_PL_Quantity", "Current_PL_Type"]
for Variable in Var_to_del:      
    try:
        exec(f'del {Variable}')
    except:
        pass
try:
    del Variable
    del Var_to_del
except:
    pass

print("Data Preparation ...")
# Export DataFrame Preparations NUS3
if General_Setup_df.iloc[0]["Navision"] == "NUS3":
    Export_Purchase_Header_dict = {
        "Document Type": Export_PH_Document_Type_list,
        "No.": Export_PH_Document_No_list,
        "Buy-from Vendor No.": Export_PH_Buy_from_Vendor_No_list,
        "Shipment Method Code": Export_PH_Shippment_Method_list,
        "Location Code": Export_PH_Document_Location_list,
        "HQ Order Type": Export_PH_HQ_Order_Type_list,
        "HQ Shipping Condition": Export_PH_HQ_Shipping_Condition_list,
        "HQ Logistic Process": Export_PH_HQ_Logistic_Process_list,
        "Shipping Agent Code": Export_PH_Shipping_Agent_list,
        "Shipping Agent Service Code": Export_PH_Shipping_Agent_Service_list}

    Export_Purchase_Lines_dict = {
        "Document Type": Export_PL_Document_Type_list,
        "Document No.": Export_PL_Document_No_list,
        "Line No.": Export_PL_Line_No_list,
        "Type": Export_PL_Type_list,
        "No.": Export_PL_No_list,
        "Quantity": Export_PL_Quantity_list}

    Export_HQ_ITR_dict = {
        "Register No.": Export_HQ_ITR_Register_No_list,
        "Register Sub No.": Export_HQ_ITR_Register_Sub_No_list,
        "Document No.": Export_HQ_ITR_Document_No_list,
        "Document Line No.": Export_HQ_ITR_Document_Line_list,
        "Document Type": Export_HQ_ITR_Document_Type_list,
        "Vendor Document Type": Export_HQ_ITR_Vendor_Document_Type_list,
        "Vendor Document No.": Export_HQ_ITR_Vendor_Document_No_list,
        "Vendor Shipment No.": Export_HQ_ITR_Vendor_Shipment_No_list,
        "Line Type": Export_HQ_ITR_Line_Type_list,
        "Item No.": Export_HQ_ITR_Item_No_list,
        "Vendor Item No.": Export_HQ_ITR_Vendor_Item_No_list,
        "Ordered Item No.": Export_HQ_ITR_Ordered_Item_No_list,
        "Quantity": Export_HQ_ITR_Quantity_list,
        "Ordered Quantity": Export_HQ_ITR_Ordered_Quantity_list,
        "Serial No.": Export_HQ_ITR_Serial_No_list,
        "Unit of Measure": Export_HQ_ITR_Unit_of_Measure_list,
        "Currency Code": Export_HQ_ITR_Currency_Code_list,
        "Unit Price": Export_HQ_ITR_Unit_Price_list,
        "Line Amount": Export_HQ_ITR_Line_Amount_list,
        "Order Date": Export_HQ_ITR_Order_Date_list,
        "Receipt Date": Export_HQ_ITR_Receipt_Date_list,
        "Posting Date": Export_HQ_ITR_Posting_Date_list,
        "Exported Line No.": Export_HQ_ITR_Exported_Line_No_list,
        "Vendor Line No.": Export_HQ_ITR_Vendor_Line_No_list,
        "Vendor No.": Export_HQ_ITR_Vendor_No_list,
        "Country/Region of Origin Code": Export_HQ_ITR_Country_Region_of_Origin_Code_list,
        "Plant No.": Export_HQ_ITR_Plant_No_list,
        "Line Flag": Export_HQ_ITR_Line_Flag_list,
        "Communication Process Status": Export_HQ_ITR_Communication_Process_Status_list,
        "ATP Check Cumulative Quantity": Export_HQ_ITR_ATP_Check_Cumulative_Quantity_list,
        "Vendor Document Created Date": Export_HQ_ITR_Vendor_Document_Created_Date_list,
        "Delivery Start Date": Export_HQ_ITR_Delivery_Start_Date_list,
        "Delivery End Date": Export_HQ_ITR_Delivery_End_Date_list,
        "Tariff Number": Export_HQ_ITR_Tariff_Number_list,
        "Vendor Delivery No.": Export_HQ_ITR_Vendor_Delivery_No_list,
        "Quantity to Deliver": Export_HQ_ITR_Quantity_to_Deliver_list,
        "Vendor Invoice No.": Export_HQ_ITR_Vendor_Invoice_No_list,
        "Quantity to Invoice": Export_HQ_ITR_Quantity_to_Invoice_list,
        "Picking Date": Export_HQ_ITR_Picking_Date_list,
        "Trans. Planning Date": Export_HQ_ITR_Trans_Planning_Date_list,
        "Loading Date": Export_HQ_ITR_Loading_Date_list,
        "Planned GI Date": Export_HQ_ITR_Planned_GI_Date_list,
        "Delivery Date": Export_HQ_ITR_Delivery_Date_list,
        "Document Order Number": Export_HQ_ITR_Document_Order_Number_list,
        "To Post - Auto": Export_HQ_ITR_To_Post_Auto_list}

    Export_HQ_ATPR_dict = {
        "HQ ATP Check Register ID": Export_HQ_ATP_HQ_ATP_Check_Register_ID_list,
        "HQ ATP Check Sub. Register ID": Export_HQ_ATP_HQ_ATP_Check_Sub_Register_ID_list,
        "HQ Item Transport Register ID": Export_HQ_ATP_HQ_Item_Transport_Register_ID_list,
        "Scheduled Quantity": Export_HQ_ATP_Scheduled_Quantity_list,
        "Scheduled Date": Export_HQ_ATP_Scheduled_Date_list,
        "Stock": Export_HQ_ATP_Stock_list,
        "Document No.": Export_HQ_ATP_Document_No_list}

    Export_HQ_SUBR_dict = {
        "HQ Substitution Register ID": Export_HQ_SUB_HQ_Substitution_Register_ID_list,
        "HQ Substitution Sub. Reg. ID": Export_HQ_SUB_HQ_Substitution_Sub_Reg_ID_list,
        "Substituted Item - Old Item": Export_HQ_SUB_Substituted_Item_Old_Item_list,
        "Substitution - New Item": Export_HQ_SUB_Substitution_New_Item_list,
        "Document No.": Export_HQ_SUB_Document_No_list}

    Export_HQ_PREAR_dict = {
        "HQ PreAdvice Register ID": Export_HQ_PREA_HQ_PreAdvice_Register_ID_list,
        "HQ PreAdvice Sub. Register ID": Export_HQ_PREA_HQ_PreAdvice_Sub_Register_ID_list,
        "HQ Item Transport Register ID": Export_HQ_PREA_HQ_Item_Transport_Register_ID_list,
        "HQ Item Transport Reg. Sub ID": Export_HQ_PREA_HQ_Item_Transport_Reg_Sub_ID_list,
        "Document No.": Export_HQ_PREA_Document_No_list,
        "PreAdvice Document No.": Export_HQ_PREA_PreAdvice_Document_No_list,
        "PreAdvice Date": Export_HQ_PREA_PreAdvice_Date_list}

    Export_HQ_SNR_dict = {
        "HQ SN Register ID": Export_HQ_SNR_HQ_SN_Register_ID_list,
        "HQ SN Sub. Register ID": Export_HQ_SNR_HQ_SN_Sub_Register_ID_list,
        "HQ Item Transport Register ID": Export_HQ_SNR_HQ_Item_Transport_Register_ID_list,
        "HQ Item Transport Sub. Reg. ID": Export_HQ_SNR_HQ_Item_Transport_Sub_Reg_ID_list,
        "Purchase Order No.": Export_HQ_SNR_Purchase_Order_No_list,
        "Item No.": Export_HQ_SNR_Item_No_list,
        "Serial Number": Export_HQ_SNR_Serial_Number_list,
        "Vendor Document No.": Export_HQ_SNR_Vendor_Document_No_list}

    Export_HQ_DTR_dict = {
        "HQ Delivery Tracking Register ID": Export_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID_list,
        "HQ Delivery Tracking Sub. Register ID": Export_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID_list,
        "Delivery No.": Export_HQ_HQDTR_Delivery_No_list,
        "Purchase Order No.": Export_HQ_HQDTR_Purchase_Order_No_list,
        "Package Tracking No.": Export_HQ_HQDTR_Package_Tracking_No_list,
        "Package Type": Export_HQ_HQDTR_Package_Type_list,
        "BEU Bill of Landing": Export_HQ_HQDTR_BEU_Bill_of_Landing_list,
        "Tracking Page": Export_HQ_HQDTR_Tracking_Page_list,
        "Incoterms": Export_HQ_HQDTR_Incoterms_list,
        "Total Weight": Export_HQ_HQDTR_Total_Weight_list,
        "Weight Unit": Export_HQ_HQDTR_Weight_Unit_list,
        "Volume": Export_HQ_HQDTR_Volume_list,
        "Volume Unit": Export_HQ_HQDTR_Volume_Unit_list,
        "Shipping Agent Code": Export_HQ_HQDTR_Shipping_Agent_Code_list,
        "EXIDV2": Export_HQ_HQDTR_EXIDV2_list}

    Export_HQ_PTR_dict = {
        "HQ Package Tracking Register ID": Export_HQ_HQPTR_HQ_Package_Tracking_Register_ID_list,
        "HQ Package Tracking Sub. Register ID": Export_HQ_HQPTR_HQ_Package_Tracking_Sub_Register_ID_list,
        "Delivery No.": Export_HQ_HQPTR_Delivery_No_list,
        "Package No.": Export_HQ_HQPTR_Package_No_list,
        "Item": Export_HQ_HQPTR_Item_list,
        "Quantity": Export_HQ_HQPTR_Quantity_list,
        "Unit of Measure": Export_HQ_HQPTR_Unit_of_Measure_list,
        "Purchase Order No.": Export_HQ_HQPTR_Purchase_Order_No_list,
        "External Package ID": Export_HQ_HQPTR_External_Package_ID_list,
        "Total Weight": Export_HQ_HQPTR_Total_Weight_list,
        "Weight Unit": Export_HQ_HQPTR_Weight_Unit_list,
        "Volume": Export_HQ_HQPTR_Volume_list,
        "Volume Unit": Export_HQ_HQPTR_Volume_Unit_list,
        "Plant No.": Export_HQ_HQPTR_Plant_No_list}

    Export_Record_Link_dict = {
        "Link ID": Link_ID_list,
        "Record ID": Record_ID_list,
        "URL1": URL1_list,
        "URL2": URL2_list,
        "URL3": URL3_list,
        "URL4": URL4_list,
        "Description": Description_list,
        "Type": Type_list,
        "Note": Note_list,
        "Created": Created_list,
        "User ID": User_ID_list,
        "Company": Company_list,
        "Notify": Notify_list,
        "To User ID": To_User_ID_list}

elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
    #! Dodělat správné názvy polí !!!!!
    Export_Purchase_Header_dict = {
        "Document Type": Export_PH_Document_Type_list,
        "No.": Export_PH_Document_No_list,
        "Buy-from Vendor No.": Export_PH_Buy_from_Vendor_No_list,
        "Shipment Method Code": Export_PH_Shippment_Method_list,
        "Location Code": Export_PH_Document_Location_list,
        "HQ Order Type": Export_PH_HQ_Order_Type_list,
        "HQ Shipping Condition": Export_PH_HQ_Shipping_Condition_list,
        "HQ Logistic Process": Export_PH_HQ_Logistic_Process_list,
        "Shipping Agent Code": Export_PH_Shipping_Agent_list,
        "Shipping Agent Service Code": Export_PH_Shipping_Agent_Service_list}

    Export_Purchase_Lines_dict = {
        "Document Type": Export_PL_Document_Type_list,
        "Document No.": Export_PL_Document_No_list,
        "Line No.": Export_PL_Line_No_list,
        "Type": Export_PL_Type_list,
        "No.": Export_PL_No_list,
        "Quantity": Export_PL_Quantity_list}

    Export_HQ_ITR_dict = {
        "Register No.": Export_HQ_ITR_Register_No_list,
        "Register Sub No.": Export_HQ_ITR_Register_Sub_No_list,
        "Document No.": Export_HQ_ITR_Document_No_list,
        "Document Line No.": Export_HQ_ITR_Document_Line_list,
        "Document Type": Export_HQ_ITR_Document_Type_list,
        "Vendor Document Type": Export_HQ_ITR_Vendor_Document_Type_list,
        "Vendor Document No.": Export_HQ_ITR_Vendor_Document_No_list,
        "Vendor Shipment No.": Export_HQ_ITR_Vendor_Shipment_No_list,
        "Line Type": Export_HQ_ITR_Line_Type_list,
        "Item No.": Export_HQ_ITR_Item_No_list,
        "Vendor Item No.": Export_HQ_ITR_Vendor_Item_No_list,
        "Ordered Item No.": Export_HQ_ITR_Ordered_Item_No_list,
        "Quantity": Export_HQ_ITR_Quantity_list,
        "Ordered Quantity": Export_HQ_ITR_Ordered_Quantity_list,
        "Serial No.": Export_HQ_ITR_Serial_No_list,
        "Unit of Measure": Export_HQ_ITR_Unit_of_Measure_list,
        "Currency Code": Export_HQ_ITR_Currency_Code_list,
        "Unit Price": Export_HQ_ITR_Unit_Price_list,
        "Line Amount": Export_HQ_ITR_Line_Amount_list,
        "Order Date": Export_HQ_ITR_Order_Date_list,
        "Receipt Date": Export_HQ_ITR_Receipt_Date_list,
        "Posting Date": Export_HQ_ITR_Posting_Date_list,
        "Exported Line No.": Export_HQ_ITR_Exported_Line_No_list,
        "Vendor Line No.": Export_HQ_ITR_Vendor_Line_No_list,
        "Vendor No.": Export_HQ_ITR_Vendor_No_list,
        "Country/Region of Origin Code": Export_HQ_ITR_Country_Region_of_Origin_Code_list,
        "Plant No.": Export_HQ_ITR_Plant_No_list,
        "Line Flag": Export_HQ_ITR_Line_Flag_list,
        "Communication Process Status": Export_HQ_ITR_Communication_Process_Status_list,
        "Vendor Document Created Date": Export_HQ_ITR_Vendor_Document_Created_Date_list,
        "Stock": Export_HQ_ATP_Stock_list}

    Export_HQ_Trans_Tracking_dict = {
        "Register No.": Export_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID_list,
        "Register Sub No.": Export_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID_list,
        "Tracking No.": Export_HQ_HQDTR_Package_Tracking_No_list,
        "Purchase Order No.": Export_HQ_HQDTR_Purchase_Order_No_list} 
else:
    pass

# Variable to Delete
Var_to_del = ["Export_HQ_ATP_Document_No_list", "Export_HQ_ATP_HQ_ATP_Check_Register_ID_list", "Export_HQ_ATP_HQ_ATP_Check_Sub_Register_ID_list", "Export_HQ_ATP_HQ_Item_Transport_Register_ID_list", "Export_HQ_ATP_Scheduled_Date_list", "Export_HQ_ATP_Scheduled_Quantity_list", "Export_HQ_ATP_Stock_list", "Export_HQ_HQDTR_BEU_Bill_of_Landing_list", "Export_HQ_HQDTR_Delivery_No_list", "Export_HQ_HQDTR_EXIDV2_list", "Export_HQ_HQDTR_HQ_Delivery_Tracking_Register_ID_list", "Export_HQ_HQDTR_HQ_Delivery_Tracking_Sub_Register_ID_list", "Export_HQ_HQDTR_Incoterms_list", "Export_HQ_HQDTR_Package_Tracking_No_list", "Export_HQ_HQDTR_Package_Type_list", "Export_HQ_HQDTR_Purchase_Order_No_list", "Export_HQ_HQDTR_Shipping_Agent_Code_list", "Export_HQ_HQDTR_Total_Weight_list", "Export_HQ_HQDTR_Tracking_Page_list", "Export_HQ_HQDTR_Volume_list", "Export_HQ_HQDTR_Volume_Unit_list", "Export_HQ_HQDTR_Weight_Unit_list", "Export_HQ_HQPTR_Delivery_No_list", "Export_HQ_HQPTR_External_Package_ID_list", "Export_HQ_HQPTR_HQ_Package_Tracking_Register_ID_list", "Export_HQ_HQPTR_HQ_Package_Tracking_Sub_Register_ID_list", "Export_HQ_HQPTR_Item_list", "Export_HQ_HQPTR_Package_No_list", "Export_HQ_HQPTR_Plant_No_list", "Export_HQ_HQPTR_Purchase_Order_No_list", "Export_HQ_HQPTR_Quantity_list", "Export_HQ_HQPTR_Total_Weight_list", "Export_HQ_HQPTR_Unit_of_Measure_list", "Export_HQ_HQPTR_Volume_list", "Export_HQ_HQPTR_Volume_Unit_list", "Export_HQ_HQPTR_Weight_Unit_list", "Export_HQ_ITR_ATP_Check_Cumulative_Quantity_list", "Export_HQ_ITR_Communication_Process_Status_list", "Export_HQ_ITR_Country_Region_of_Origin_Code_list", "Export_HQ_ITR_Currency_Code_list", "Export_HQ_ITR_Delivery_Date_list", "Export_HQ_ITR_Delivery_End_Date_list", "Export_HQ_ITR_Delivery_Start_Date_list", "Export_HQ_ITR_Document_No_list", "Export_HQ_ITR_Document_Order_Number_list", "Export_HQ_ITR_Document_Type_list", "Export_HQ_ITR_Exported_Line_No_list", "Export_HQ_ITR_Item_No_list", "Export_HQ_ITR_Line_Amount_list", "Export_HQ_ITR_Line_Flag_list", "Export_HQ_ITR_Line_Type_list", "Export_HQ_ITR_Loading_Date_list", "Export_HQ_ITR_Order_Date_list", "Export_HQ_ITR_Ordered_Item_No_list", "Export_HQ_ITR_Ordered_Quantity_list", "Export_HQ_ITR_Picking_Date_list", "Export_HQ_ITR_Planned_GI_Date_list", "Export_HQ_ITR_Plant_No_list", "Export_HQ_ITR_Posting_Date_list", "Export_HQ_ITR_Quantity_list", "Export_HQ_ITR_Quantity_to_Deliver_list", "Export_HQ_ITR_Quantity_to_Invoice_list", "Export_HQ_ITR_Receipt_Date_list", "Export_HQ_ITR_Register_No_list", "Export_HQ_ITR_Register_Sub_No_list", "Export_HQ_ITR_Serial_No_list", "Export_HQ_ITR_Tariff_Number_list", "Export_HQ_ITR_To_Post_Auto_list", "Export_HQ_ITR_Trans_Planning_Date_list", "Export_HQ_ITR_Unit_of_Measure_list", "Export_HQ_ITR_Unit_Price_list", "Export_HQ_ITR_Vendor_Delivery_No_list", "Export_HQ_ITR_Vendor_Document_Created_Date_list", "Export_HQ_ITR_Vendor_Document_No_list", "Export_HQ_ITR_Vendor_Document_Type_list", "Export_HQ_ITR_Vendor_Invoice_No_list", "Export_HQ_ITR_Vendor_Item_No_list", "Export_HQ_ITR_Vendor_Line_No_list", "Export_HQ_ITR_Vendor_No_list", "Export_HQ_ITR_Vendor_Shipment_No_list", "Export_HQ_PREA_Document_No_list", "Export_HQ_PREA_HQ_Item_Transport_Reg_Sub_ID_list", "Export_HQ_PREA_HQ_Item_Transport_Register_ID_list", "Export_HQ_PREA_HQ_PreAdvice_Register_ID_list", "Export_HQ_PREA_HQ_PreAdvice_Sub_Register_ID_list", "Export_HQ_PREA_PreAdvice_Date_list", "Export_HQ_PREA_PreAdvice_Document_No_list", "Export_HQ_SNR_HQ_Item_Transport_Register_ID_list", "Export_HQ_SNR_HQ_Item_Transport_Sub_Reg_ID_list", "Export_HQ_SNR_HQ_SN_Register_ID_list", "Export_HQ_SNR_HQ_SN_Sub_Register_ID_list", "Export_HQ_SNR_Item_No_list", "Export_HQ_SNR_Purchase_Order_No_list", "Export_HQ_SNR_Serial_Number_list", "Export_HQ_SNR_Vendor_Document_No_list", "Export_HQ_SUB_Document_No_list", "Export_HQ_SUB_HQ_Substitution_Register_ID_list", "Export_HQ_SUB_HQ_Substitution_Sub_Reg_ID_list", "Export_HQ_SUB_Substituted_Item_Old_Item_list", "Export_HQ_SUB_Substitution_New_Item_list", "Export_PH_Buy_from_Vendor_No_list", "Export_PH_Document_No_list", "Export_PH_Document_Type_list", "Export_PH_HQ_Logistic_Process_list", "Export_PH_HQ_Order_Type_list", "Export_PH_HQ_Shipping_Condition_list", "Export_PH_Shipping_Agent_list", "Export_PH_Shipping_Agent_Service_list", "Export_PH_Shippment_Method_list", "Export_PL_Document_No_list", "Export_PL_Document_Type_list", "Export_PL_Line_No_list", "Export_PL_No_list", "Export_PL_Quantity_list", "Export_PL_Type_list"]
for Variable in Var_to_del:      
    try:
        exec(f'del {Variable}')
    except:
        pass
try:
    del Variable
    del Var_to_del
except:
    pass

# Export Dataframe preparation
if General_Setup_df.iloc[0]["Navision"] == "NUS3":
    Export_Purchase_Header_df = pandas.DataFrame(Export_Purchase_Header_dict, columns=Export_Purchase_Header_dict.keys())
    Export_Purchase_Lines_df = pandas.DataFrame(Export_Purchase_Lines_dict, columns=Export_Purchase_Lines_dict.keys())
    Export_HQ_ITR_df = pandas.DataFrame(Export_HQ_ITR_dict, columns=Export_HQ_ITR_dict.keys())
    Export_HQ_ATPR_df = pandas.DataFrame(Export_HQ_ATPR_dict, columns=Export_HQ_ATPR_dict.keys())
    Export_HQ_SUBR_df = pandas.DataFrame(Export_HQ_SUBR_dict, columns=Export_HQ_SUBR_dict.keys())
    Export_HQ_PREAR_df = pandas.DataFrame(Export_HQ_PREAR_dict, columns=Export_HQ_PREAR_dict.keys())
    Export_HQ_SNR_df = pandas.DataFrame(Export_HQ_SNR_dict, columns=Export_HQ_SNR_dict.keys())
    Export_HQ_DTR_df = pandas.DataFrame(Export_HQ_DTR_dict, columns=Export_HQ_DTR_dict.keys())
    Export_HQ_PTR_df = pandas.DataFrame(Export_HQ_PTR_dict, columns=Export_HQ_PTR_dict.keys())
    Export_Record_Link_df = pandas.DataFrame(Export_Record_Link_dict, columns=Export_Record_Link_dict.keys())
elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
    Export_Purchase_Header_df = pandas.DataFrame(Export_Purchase_Header_dict, columns=Export_Purchase_Header_dict.keys())
    Export_Purchase_Lines_df = pandas.DataFrame(Export_Purchase_Lines_dict, columns=Export_Purchase_Lines_dict.keys())
    Export_HQ_ITR_df = pandas.DataFrame(Export_HQ_ITR_dict, columns=Export_HQ_ITR_dict.keys())
    Export_HQ_TTR_df = pandas.DataFrame(Export_HQ_Trans_Tracking_dict, columns=Export_HQ_Trans_Tracking_dict.keys())
else:
    pass

# NUS3 - Load Invoice Numbers to field "Vendor Invoice No" on the record of "Vendor Document Type = Deviery" --> new development by ""
if General_Setup_df.iloc[0]["Navision"] == "NUS3":
    mask_1 = Export_HQ_ITR_df["Vendor Document Type"] == "Invoice"
    mask_2 = Export_HQ_ITR_df["Document Order Number"] == "10"
    Export_HQ_ITR_df_help = Export_HQ_ITR_df[mask_1 & mask_2]
    Export_HQ_ITR_df_help1 = Export_HQ_ITR_df_help[["Vendor Document No.", "Vendor Shipment No."]]

    for row in Export_HQ_ITR_df_help1.iterrows():
        row_df = pandas.Series(data = row[1]) 
        Export_HQ_ITR_df.loc[(Export_HQ_ITR_df["Vendor Document No."] == row_df["Vendor Shipment No."]) & (Export_HQ_ITR_df["Vendor Document Type"] == "Delivery"), "Vendor Invoice No."] = row_df["Vendor Document No."]

# Variable to Delete
Var_to_del = ["Export_Purchase_Header_dict", "Export_Purchase_Lines_dict", "Export_HQ_ITR_dict", "Export_HQ_ATPR_dict", "Export_HQ_SUBR_dict", "Export_HQ_PREAR_dict", "Export_HQ_SNR_dict", "Export_HQ_DTR_dict", "Export_HQ_PTR_dict"]
for Variable in Var_to_del:      
    try:
        exec(f'del {Variable}')
    except:
        pass
try:
    del Variable
    del Var_to_del
except:
    pass

print("Data Saving ...")
# Export_Type = Excel
if General_Setup_df.iloc[0]["Export_Type"] == "EXCEL":
    if General_Setup_df.iloc[0]["Export_File_Location_as_Current"] == "Y":
        Export_File_path = os.path.dirname(os.path.abspath(__file__))
    else:
        Export_File_path = General_Setup_df.iloc[0]["Export_File_path"]
    Export_File_name = General_Setup_df.iloc[0]["Export_File_name"]

    # Writer Dataframes
    if General_Setup_df.iloc[0]["Navision"] == "NUS3":
        # Copy Configuraiton package to destina tion folder and rename it
        original = f"{Export_File_path}/NUS3/Templates/HQ_DATA_GENERATOR_template.xlsx"
        target = f"{Export_File_path}/NUS3_{Export_File_name}.xlsx"
        shutil.copyfile(original, target)

        with pandas.ExcelWriter(f"{Export_File_path}/NUS3_{Export_File_name}.xlsx", engine= "openpyxl", mode="w") as writer:  
            Export_Purchase_Header_df.to_excel(writer, sheet_name="Purchase Header", index=False, startcol=0, startrow=2, header=True, engine="openpyxl")
            Export_Purchase_Lines_df.to_excel(writer, sheet_name="Purchase Line", index=False, startcol=0, startrow=2, header=True, engine="openpyxl")
            Export_HQ_ITR_df.to_excel(writer, sheet_name="HQ Item Transport Register", index=False, startcol=0, startrow=2, header=True, engine="openpyxl")
            Export_HQ_ATPR_df.to_excel(writer, sheet_name="HQ ATP Check Register", index=False, startcol=0, startrow=2, header=True, engine="openpyxl")
            Export_HQ_SUBR_df.to_excel(writer, sheet_name="HQ Substitution Register", index=False, startcol=0, startrow=2, header=True, engine="openpyxl")
            Export_HQ_PREAR_df.to_excel(writer, sheet_name="HQ PreAdvice Register", index=False, startcol=0, startrow=2, header=True, engine="openpyxl")
            Export_HQ_SNR_df.to_excel(writer, sheet_name="HQ Serial Number Register", index=False, startcol=0, startrow=2, header=True, engine="openpyxl")
            Export_HQ_DTR_df.to_excel(writer, sheet_name="HQ Delivery Tracking Register", index=False, startcol=0, startrow=2, header=True, engine="openpyxl")
            Export_HQ_PTR_df.to_excel(writer, sheet_name="HQ Package Tracking Register", index=False, startcol=0, startrow=2, header=True, engine="openpyxl")
            Export_Record_Link_df.to_excel(writer, sheet_name="Record Link", index=False, startcol=0, startrow=0, header=True, engine="openpyxl")

        print("Data Formating ...")
        # Writer addtional texts + table definition
        workbook=openpyxl.load_workbook(f"{Export_File_path}/NUS3_{Export_File_name}.xlsx")
        workbook_index = workbook.sheetnames
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

        # Purcahse Header
        worksheet = workbook.worksheets[workbook_index.index("Purchase Header")]
        worksheet["A1"] = str(General_Setup_df.iloc[0]["Conf_Package"])
        worksheet["B1"] = "Purchase Header"
        worksheet["C1"] = "38"

        Columns_Name = excel_colum_def(Export_Purchase_Header_df.shape[1])
        if Export_Purchase_Header_df.shape[0] != 0:
            Puchase_Header_tab = Table(displayName="Purchase_Header", ref=f"A3:{Columns_Name}{Export_Purchase_Header_df.shape[0] + 3}")
        else:
            Puchase_Header_tab = Table(displayName="Purchase_Header", ref=f"A3:{Columns_Name}{1 + 3}")
        Puchase_Header_tab.tableStyleInfo = style
        worksheet.add_table(Puchase_Header_tab)

        # Purchase Lines
        worksheet = workbook.worksheets[workbook_index.index("Purchase Line")]
        worksheet["A1"] = str(General_Setup_df.iloc[0]["Conf_Package"])
        worksheet["B1"] = "Purchase Line"
        worksheet["C1"] = "39"

        Columns_Name = excel_colum_def(Export_Purchase_Lines_df.shape[1])
        if Export_Purchase_Lines_df.shape[0] != 0:
            Purchase_Line_tab = Table(displayName="Purchase_Line", ref=f"A3:{Columns_Name}{Export_Purchase_Lines_df.shape[0] + 3}")
        else:
            Purchase_Line_tab = Table(displayName="Purchase_Line", ref=f"A3:{Columns_Name}{1 + 3}")
        Purchase_Line_tab.tableStyleInfo = style
        worksheet.add_table(Purchase_Line_tab)

        # HQ Item Transport Register
        worksheet = workbook.worksheets[workbook_index.index("HQ Item Transport Register")]
        worksheet["A1"] = str(General_Setup_df.iloc[0]["Conf_Package"])
        worksheet["B1"] = "HQ Item Transport Register"
        worksheet["C1"] = "51064"

        Columns_Name = excel_colum_def(Export_HQ_ITR_df.shape[1])
        if Export_HQ_ITR_df.shape[0] != 0:
            HQ_ITR_tab = Table(displayName="HQ_Item_Transport_Register", ref=f"A3:{Columns_Name}{Export_HQ_ITR_df.shape[0] + 3}")
        else:
            HQ_ITR_tab = Table(displayName="HQ_Item_Transport_Register", ref=f"A3:{Columns_Name}{1 + 3}")
        HQ_ITR_tab.tableStyleInfo = style
        worksheet.add_table(HQ_ITR_tab)

        # HQ ATP Check Register
        worksheet = workbook.worksheets[workbook_index.index("HQ ATP Check Register")]
        worksheet["A1"] = str(General_Setup_df.iloc[0]["Conf_Package"])
        worksheet["B1"] = "HQ ATP Check Register"
        worksheet["C1"] = "51060"

        Columns_Name = excel_colum_def(Export_HQ_ATPR_df.shape[1])
        if Export_HQ_ATPR_df.shape[0] != 0:
            HQ_ATPR_tab = Table(displayName="HQ_ATP_Check_Register", ref=f"A3:{Columns_Name}{Export_HQ_ATPR_df.shape[0] + 3}")
        else:
            HQ_ATPR_tab = Table(displayName="HQ_ATP_Check_Register", ref=f"A3:{Columns_Name}{1 + 3}")
        HQ_ATPR_tab.tableStyleInfo = style
        worksheet.add_table(HQ_ATPR_tab)

        # HQ Substitution Register
        worksheet = workbook.worksheets[workbook_index.index("HQ Substitution Register")]
        worksheet["A1"] = str(General_Setup_df.iloc[0]["Conf_Package"])
        worksheet["B1"] = "HQ Substitution Register"
        worksheet["C1"] = "51066"

        Columns_Name = excel_colum_def(Export_HQ_SUBR_df.shape[1])
        if Export_HQ_SUBR_df.shape[0] != 0:
            HQ_SUBR_tab = Table(displayName="HQ_Substitution_Register", ref=f"A3:{Columns_Name}{Export_HQ_SUBR_df.shape[0] + 3}")
        else:
            HQ_SUBR_tab = Table(displayName="HQ_Substitution_Register", ref=f"A3:{Columns_Name}{1 + 3}")
        HQ_SUBR_tab.tableStyleInfo = style
        worksheet.add_table(HQ_SUBR_tab)

        # HQ Pre-Advice Register
        worksheet = workbook.worksheets[workbook_index.index("HQ PreAdvice Register")]
        worksheet["A1"] = str(General_Setup_df.iloc[0]["Conf_Package"])
        worksheet["B1"] = "HQ PreAdvice Register"
        worksheet["C1"] = "51071"

        Columns_Name = excel_colum_def(Export_HQ_PREAR_df.shape[1])
        if Export_HQ_PREAR_df.shape[0] != 0:
            HQ_PRER_tab = Table(displayName="HQ_PreAdvice_Register", ref=f"A3:{Columns_Name}{Export_HQ_PREAR_df.shape[0] + 3}")
        else:
            HQ_PRER_tab = Table(displayName="HQ_PreAdvice_Register", ref=f"A3:{Columns_Name}{1 + 3}")
        HQ_PRER_tab.tableStyleInfo = style
        worksheet.add_table(HQ_PRER_tab)

        # HQ Serial Number Register
        worksheet = workbook.worksheets[workbook_index.index("HQ Serial Number Register")]
        worksheet["A1"] = str(General_Setup_df.iloc[0]["Conf_Package"])
        worksheet["B1"] = "HQ Serial Number Register"
        worksheet["C1"] = "51065"

        Columns_Name = excel_colum_def(Export_HQ_SNR_df.shape[1])
        if Export_HQ_SNR_df.shape[0] != 0:
            HQ_SNR_tab = Table(displayName="HQ_Serial_Number_Register", ref=f"A3:{Columns_Name}{Export_HQ_SNR_df.shape[0] + 3}")
        else:
            HQ_SNR_tab = Table(displayName="HQ_Serial_Number_Register", ref=f"A3:{Columns_Name}{1 + 3}")
        HQ_SNR_tab.tableStyleInfo = style
        worksheet.add_table(HQ_SNR_tab)

        # HQ Delivery Tracking Register
        worksheet = workbook.worksheets[workbook_index.index("HQ Delivery Tracking Register")]
        worksheet["A1"] = str(General_Setup_df.iloc[0]["Conf_Package"])
        worksheet["B1"] = "HQ Delivery Tracking Register"
        worksheet["C1"] = "51062"

        Columns_Name = excel_colum_def(Export_HQ_DTR_df.shape[1])
        if Export_HQ_DTR_df.shape[0] != 0:
            HQ_DTR_tab = Table(displayName="HQ_Delivery_Tracking_Register", ref=f"A3:{Columns_Name}{Export_HQ_DTR_df.shape[0] + 3}")
        else:
            HQ_DTR_tab = Table(displayName="HQ_Delivery_Tracking_Register", ref=f"A3:{Columns_Name}{1 + 3}")
        HQ_DTR_tab.tableStyleInfo = style
        worksheet.add_table(HQ_DTR_tab)

        # HQ Package Tracking Regsiter
        worksheet = workbook.worksheets[workbook_index.index("HQ Package Tracking Register")]
        worksheet["A1"] = str(General_Setup_df.iloc[0]["Conf_Package"])
        worksheet["B1"] = "HQ Package Tracking Register"
        worksheet["C1"] = "51063"

        Columns_Name = excel_colum_def(Export_HQ_PTR_df.shape[1])
        if Export_HQ_PTR_df.shape[0] != 0:
            HQ_PTR_tab = Table(displayName="HQ_Package_Tracking_Register", ref=f"A3:{Columns_Name}{Export_HQ_PTR_df.shape[0] + 3}")
        else:
            HQ_PTR_tab = Table(displayName="HQ_Package_Tracking_Register", ref=f"A3:{Columns_Name}{1 + 3}")
        HQ_PTR_tab.tableStyleInfo = style
        worksheet.add_table(HQ_PTR_tab)

        # Record Links
        worksheet = workbook.worksheets[workbook_index.index("Record Link")]
        Columns_Name = excel_colum_def(Export_Record_Link_df.shape[1])
        if Export_Record_Link_df.shape[0] != 0:
            HQ_Link_tab = Table(displayName="Export_Record_Link_df", ref=f"A1:{Columns_Name}{Export_Record_Link_df.shape[0]}")
        else:
            HQ_Link_tab = Table(displayName="Export_Record_Link_df", ref=f"A1:{Columns_Name}{1}")
        HQ_Link_tab.tableStyleInfo = style
        worksheet.add_table(HQ_Link_tab)

        workbook.save(f"{Export_File_path}/NUS3_{Export_File_name}.xlsx") 

    elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
        with pandas.ExcelWriter(f"{Export_File_path}/NUS2_{Export_File_name}.xlsx", engine= "openpyxl") as writer:  
            Export_Purchase_Header_df.to_excel(writer, sheet_name="PurchaseHeader", index=False, startcol=0, startrow=2, header=True)
            Export_Purchase_Lines_df.to_excel(writer, sheet_name="PurchaseLine", index=False, startcol=0, startrow=2, header=True)
            Export_HQ_ITR_df.to_excel(writer, sheet_name="HQItemTransportRegister", index=False, startcol=0, startrow=2, header=True)
            Export_HQ_TTR_df.to_excel(writer, sheet_name="HQTransferTrackingNo", index=False, startcol=0, startrow=2, header=True)

        print("Data Updates ...")
        # Writer addtional texts + table definition
        workbook=openpyxl.load_workbook(f"{Export_File_path}/NUS2_{Export_File_name}.xlsx")
        workbook_index = workbook.sheetnames
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

        # Purcahse Header
        worksheet = workbook.worksheets[workbook_index.index("PurchaseHeader")]
        worksheet["A1"] = "Purchase Header"
        worksheet["B1"] = "38"

        Columns_Name = excel_colum_def(Export_Purchase_Header_df.shape[1])
        if Export_Purchase_Header_df.shape[0] != 0:
            Puchase_Header_tab = Table(displayName="Purchase_Header", ref=f"A3:{Columns_Name}{Export_Purchase_Header_df.shape[0] + 3}")
        else:
            Puchase_Header_tab = Table(displayName="Purchase_Header", ref=f"A3:{Columns_Name}{1 + 3}")
        Puchase_Header_tab.tableStyleInfo = style
        worksheet.add_table(Puchase_Header_tab)

        # Purchase Lines
        worksheet = workbook.worksheets[workbook_index.index("PurchaseLine")]
        worksheet["A1"] = "Purchase Line"
        worksheet["B1"] = "39"

        Columns_Name = excel_colum_def(Export_Purchase_Lines_df.shape[1])
        if Export_Purchase_Lines_df.shape[0] != 0:
            Purchase_Line_tab = Table(displayName="Purchase_Line", ref=f"A3:{Columns_Name}{Export_Purchase_Lines_df.shape[0] + 3}")
        else:
            Purchase_Line_tab = Table(displayName="Purchase_Line", ref=f"A3:{Columns_Name}{1 + 3}")
        Purchase_Line_tab.tableStyleInfo = style
        worksheet.add_table(Purchase_Line_tab)

        # HQ Item Transport Register
        worksheet = workbook.worksheets[workbook_index.index("HQItemTransportRegister")]
        worksheet["A1"] = "HQ Item Transport Register"
        worksheet["B1"] = "51317"

        Columns_Name = excel_colum_def(Export_HQ_ITR_df.shape[1])
        if Export_HQ_ITR_df.shape[0] != 0:
            HQ_ITR_tab = Table(displayName="HQ_Item_Transport_Register", ref=f"A3:{Columns_Name}{Export_HQ_ITR_df.shape[0] + 3}")
        else:
            HQ_ITR_tab = Table(displayName="HQ_Item_Transport_Register", ref=f"A3:{Columns_Name}{1 + 3}")
        HQ_ITR_tab.tableStyleInfo = style
        worksheet.add_table(HQ_ITR_tab)

        # HQ Transfer Tracking No
        worksheet = workbook.worksheets[workbook_index.index("HQTransferTrackingNo")]
        worksheet["A1"] = "HQ Transfer Tracking No."
        worksheet["B1"] = "51322"

        Columns_Name = excel_colum_def(Export_HQ_TTR_df.shape[1])
        if Export_HQ_TTR_df.shape[0] != 0:
            HQ_TTR_tab = Table(displayName="HQ_Transfer_Tracking_No", ref=f"A3:{Columns_Name}{Export_HQ_TTR_df.shape[0] + 3}")
        else:
            HQ_TTR_tab = Table(displayName="HQ_Transfer_Tracking_No", ref=f"A3:{Columns_Name}{1 + 3}")
        HQ_TTR_tab.tableStyleInfo = style
        worksheet.add_table(HQ_TTR_tab)

        workbook.save(f"{Export_File_path}/NUS2_{Export_File_name}.xlsx") 
    else:
        pass

elif General_Setup_df.iloc[0]["Export_Type"] == "XML":
    if General_Setup_df.iloc[0]["Navision"] == "NUS3":
        #! Dodělat
        # General principle
        """
        1) Copy template into varialbe
        2) update values
        3) Safe file into import folders (file name take as pořadové čislo + Vendor Document No.)
        4) Zkontrolovat jestli existují importní složky
        """

        #-------- Confirmaiton --------#
        with open("/Lib/NUS3_Template_Confirmation.xml", "r") as f:
            temp_Header_file_lines = f.readlines()
        for line in temp_Header_file_lines:
            line = line.replace("<GENERATION_DATE>HERE</GENERATION_DATE>", f"<GENERATION_DATE></GENERATION_DATE>")
        #-------- PreAdvice --------#
        #-------- Delivery --------#
        #-------- Invoice --------#
    elif General_Setup_df.iloc[0]["Navision"] == "NUS2":
        pass
        #! Dodělat
    else:
        print("You are idiot to create all data and not to export them.")
else:
    print("No export Type selected")

print("Test Data prepared")
